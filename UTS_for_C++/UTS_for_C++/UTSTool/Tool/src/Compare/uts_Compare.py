##
# @file uts_Compare.py
#
# @brief Below is brief information of Test Comparison module.
#
# @section Description
# This module supports validating output log file and update test report.
#
# @section Notes
# - None
#
# @section Restriction
# - None.
#
# @section History
# - 2020.07.06  Create Baseline
#
# Copyright(c) 2020 Renesas Electoronics Corporation

# Imports
from os.path             import abspath, exists, basename, dirname, join
from os                  import mkdir, listdir
from configparser        import ConfigParser
from sys                 import argv
from copy                import deepcopy
from re                  import search, findall, sub, split, MULTILINE, DOTALL
from openpyxl            import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles     import Border, Side
import sys
for str_dir in listdir(dirname(dirname(abspath(__file__)))):
	sys.path.insert(0, join(dirname(dirname(abspath(__file__))), str_dir))
import uts_TestExtract
import uts_TestCode
from uts_TestExtract import testdata_finding_type_row as f_type_row
from uts_TestExtract import testdata_get_cell_value


class ConfigInfo:
	str_specpath = "" # Test spec path
	str_logpath  = "" # Test log path
	str_revision = "" # Source test revision
	str_reportsheet = "" # Target test report sheet
	list_sheet   = [] # List of target function (list of str)


class ValidateInfo:
	str_description = "" # Validate description
	str_macro       = "" # Macro value
	str_digit       = "" # Decimal value


class LogDataInfo:
	str_testcaseid    = "" # Test case ID
	str_testresult    = "" # Test result
	list_validateinfo = [] # List of ValidateInfo in a test case (list of class ValidateInfo)


class LogInfo:
	str_function     = "" # Target function
	list_logdatainfo = [] # List of LogDataInfo in a Target function (list of class LogDataInfo)


class DataSpecInfo:
	class_item       = None # uts_TestExtract.CellInfo of item
	class_type       = None # uts_TestExtract.CellInfo of type
	class_param_name = None # uts_TestExtract.CellInfo of parameter name
	class_io         = None # uts_TestExtract.CellInfo of io
	class_meaning    = None # uts_TestExtract.CellInfo of meaning


class TableInfo:
	str_function = "" # Target function
	list_pcl     = [] # Contains expected data of pcl (2-demension list of str)
	list_log     = [] # Contains expected data of log (2-demension list of str)
	list_info    = [] # Contains pcl information and the results (2-demension list of str)
	list_err     = [] # Contains log value not exist in pcl (2-demension list of str)


class TestCaseInfo:
	str_function   = "" # Target function
	str_testcaseid = "" # Test Case ID
	str_feature    = "" # Test Feature
	str_testtype   = "" # Test Type
	str_testresult = "" # Test Result
	str_logfiles   = "" # Test Log
	str_remarks    = "" # Test Remarks
	str_device	   = "" # Target Device
	int_row        = 0  # Current row in Test Report


# Macro
g_str_TESTCASEID        = "Test Case ID"
g_str_FEATURE           = "Feature/\nComponent"
g_str_TESTTYPE          = "Test \nType"
g_str_RESULT            = "Result"
g_str_LOGFILES          = "Log Files"
g_str_REMARKS           = "Remarks"
g_str_REVISION          = "Test Target SCM Revision"
g_int_TESTID_COL        = 3
g_int_TESTCASEID_COL    = 4
g_int_DATASPEC_COL      = 8
g_int_TR_TESTCASEID_COL = 4
g_int_TR_TESTCASEID_ROW = 13
g_int_TR_REVISION_COL   = 7
g_int_TR_REVISION_ROW   = 10
g_str_NULL              = "null"
g_str_NOT_NULL          = "not null"
# Debug mode
g_bool_DEBUG = False


# Require option in config file
g_dict_OPTION = {
	"PATH"              : ["InputPath"],
	"COMPARE"           : ["LogPath", "Revision"],
	"TARGET_OF_TESTING" : ["FindByPrefix", "Prefix", "SheetNameContainData"],
}


int_PCLTABLE  = 1
int_LOGTABLE  = 2
int_INFOTABLE = 3
int_ERRTABLE  = 4


## This function is used to get workbook object
# @param  str_specpath   Path to Test Specification
# @return class_workbook Workbok object
def __get_workbook(str_specpath):
	class_workbook = load_workbook(str_specpath, data_only=True)
	return class_workbook


## This function is used to get workbook sheet object
# @param  class_workbook  Workbook object
# @param  str_sheet       Sheet name
# @return class_worksheet Worksheet object
def __get_worksheet(class_workbook, str_sheet):
	try:
		return class_workbook[str_sheet]
	except:
		raise Exception("[E_UTS_PCL] Sheet {}: is not exist.".format(str_sheet))


## This function is used to get a block of test case log
# @param  str_log          Workbook object
# @return list_of_testcase List of testcase
def __get_testcasedata_from_test_log(str_log):
	str_pattern = r"(\[I\]\[\s*RUN\s*\].+?\[V\]\[\s*RESULT\s*\].+?\n)"
	list_of_testcase = findall(str_pattern, str_log, MULTILINE | DOTALL)
	return list_of_testcase


## This function is used to get information from log file
# @param  str_logpath  Path to Test log
# @return list_loginfo List of class LogInfo
def __get_list_loginfo(str_logpath):
	class_file = uts_TestExtract.testdata_get_file(str_logpath, "r")
	list_log = __get_testcasedata_from_test_log(class_file.read())
	class_file.close()

	list_loginfo = []
	# Loop for each test case log
	for str_log in list_log:
		str_log = str_log[:-1]  # Remove the last endline \n
		list_log_line = str_log.splitlines()  # Split line

		str_log_start = list_log_line.pop(0)
		str_log_end = list_log_line.pop(-1).replace(" ", "")

		class_re = search(r"^\[I\]\[\s*RUN\s*\]\s*UT_(\w+?)_(V\w+?)\.(.+)$", str_log_start)
		if class_re:
			str_function = class_re.group(1)
			str_testcaseid = class_re.group(3)
		else:
			# Log format error, continue with other log
			print("[Warning] {}: Format error.".format(str_log_start))
			continue

		# Get the testcase result
		class_re = search( r"^\[V\]\[RESULT\].+?:(\w+)$", str_log_end)
		if class_re:
			str_testresult = class_re.group(1).strip().upper()
		else:
			# Log format error, continue with other log
			print("[Warning] {}: Format error.".format(str_log_end))
			continue

		# Convert testcase result to PASSED, FAILED and N/A
		if str_testresult in list(["OK", "NG"]):
			str_testresult = str_testresult.replace("OK", "PASSED")
			str_testresult = str_testresult.replace("NG", "FAILED")
		else:
			str_testresult = "N/A"

		# Get validation data from log
		list_validateinfo = []
		# Loop for each validation log
		for str_log_validate in list_log_line:
			if ("[V]" in str_log_validate):
				str_patt = r"^\[V\]\<\S+?\>\s*(.+?)\s*\[output\].*?\<==\>\s*\[expected\]\s*(.*)\s*\((.*?)\)$"
			elif ("[E]" in str_log_validate):
				str_patt = r"^\[E\]\<\S+?\>\s*(.+?)\s*\[output\].*?\<==\>\s*\[expected\]\s*(.*)\s*\((.*?)\)$"
			else:
				str_patt = ""

			if (str_patt != ""):
				class_re = search(str_patt, str_log_validate)
				if class_re:
					str_description = class_re.group(1)
					str_macro = class_re.group(2)
					str_digit = class_re.group(3)
				else:
					print("[Warning] {}: Format error.".format(str_log_validate))
					continue

				class_validate_info = ValidateInfo()
				class_validate_info.str_description = str_description
				class_validate_info.str_macro = str_macro
				class_validate_info.str_digit = str_digit

				list_validateinfo.append(deepcopy(class_validate_info))
			else:
				print("[Warning] {}: Format error.".format(str_log_validate))

		class_loginfo = LogInfo()
		class_loginfo.str_function = str_function
		class_loginfo.list_logdatainfo = []

		class_log_datainfo = LogDataInfo()
		class_log_datainfo.str_testcaseid = str_testcaseid
		class_log_datainfo.str_testresult = str_testresult
		class_log_datainfo.list_validateinfo = list_validateinfo

		class_loginfo.list_logdatainfo.append(deepcopy(class_log_datainfo))

		for int_index in range(len(list_loginfo)):
			if (list_loginfo[int_index].str_function == class_loginfo.str_function):
				# Check duplicate Test Case ID in Test log
				for class_logdatainfo in list_loginfo[int_index].list_logdatainfo:
					if class_logdatainfo.str_testcaseid == str_testcaseid:
						raise Exception ("[E_UTS_FILE] Log file {}: duplicate Test Case ID {}.".format(str_logpath, str_testcaseid))

				list_loginfo[int_index].list_logdatainfo.append(deepcopy(class_log_datainfo))
		else:
			list_loginfo.append(deepcopy(class_loginfo))

	return list_loginfo


## This function is used to get data specification column information in Test Spec
# @param  class_worksheet    Worksheet object
# @return class_dataspecinfo class DataSpecInfo
def __get_list_dataspecinfo(class_worksheet):
	iter_col = class_worksheet[get_column_letter(g_int_DATASPEC_COL)]
	for class_cell in iter_col:
		str_value = str(class_cell.value).strip()

		class_cellinfo = uts_TestExtract.CellInfo(class_cell.row, class_cell.column, str_value)

		if (str_value == uts_TestExtract.str_ITEM):
			class_item = deepcopy(class_cellinfo)
		elif (str_value == uts_TestExtract.str_TYPE):
			class_type = deepcopy(class_cellinfo)
		elif (str_value == uts_TestExtract.str_PARAMETER_NAME):
			class_param_name = deepcopy(class_cellinfo)
		elif (str_value == uts_TestExtract.str_IO):
			class_io = deepcopy(class_cellinfo)
		elif (str_value == uts_TestExtract.str_MEANING):
			class_meaning = deepcopy(class_cellinfo)

	if (not class_item):
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: not found \"Item\"".format(class_worksheet.title, get_column_letter(g_int_DATASPEC_COL)))
	if (not class_type):
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: not found \"Type\"".format(class_worksheet.title, get_column_letter(g_int_DATASPEC_COL)))
	if (not class_param_name):
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: not found \"Parameter name\"".format(class_worksheet.title, get_column_letter(g_int_DATASPEC_COL)))
	if (not class_io):
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: not found \"I/O\"".format(class_worksheet.title, get_column_letter(g_int_DATASPEC_COL)))
	if (not class_meaning):
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: not found \"Meaning\"".format(class_worksheet.title, get_column_letter(g_int_DATASPEC_COL)))

	class_dataspecinfo = DataSpecInfo()
	class_dataspecinfo.class_item = class_item
	class_dataspecinfo.class_type = class_type
	class_dataspecinfo.class_param_name = class_param_name
	class_dataspecinfo.class_io = class_io
	class_dataspecinfo.class_meaning = class_meaning

	return class_dataspecinfo


## This function is used to form a row from table to a line, separate by comma ","
# @param  list_data Table data
# @return str_all   A line of string
def __get_line_of_table(list_data):
	str_all = ""
	for list_row in list_data:
		str_row = ""
		for str_cell in list_row:
			str_row += str(str_cell) + ","
		str_row = str_row[:-1]
		str_row += "\n"
		str_all += str_row
	return str_all


## This function is used to write data of Table to file csv
# @param  list_tableinfo List of all Table information
# @return None
def __testreport_write_to_csv(list_tableinfo):
	str_comparisonlog_folder = "comparisonlog"
	for class_tableinfo in list_tableinfo:
		if not exists(str_comparisonlog_folder):
			mkdir(str_comparisonlog_folder)
		class_file = uts_TestExtract.testdata_get_file(join(str_comparisonlog_folder, class_tableinfo.str_function + ".csv"), "w")
		class_file.write("PCL expected data\n")
		class_file.write(__get_line_of_table(class_tableinfo.list_pcl))
		class_file.write("\nLog expected data\n")
		class_file.write(__get_line_of_table(class_tableinfo.list_log))
		class_file.write("\nLog data not found in PCL\n")
		class_file.write(__get_line_of_table(class_tableinfo.list_err))
		class_file.write("\nResult\n")
		class_file.write(__get_line_of_table(class_tableinfo.list_info))
		class_file.close()


## This function is used to create missing element of Array by detect ... in data table
# @param  list_pcl  Data table
# @return list_pcl  Data table with array modified
def __handle_array_three_dot(list_pcl):
	int_skip = 0
	for int_col, str_target in enumerate(list_pcl[0]):
		if int_skip > 0:
			int_skip -= 1
			continue

		str_target = str_target.replace(" ", "")
		# Detect array element that contain MACRO
		if search(r"\[[a-z_A-Z]+\]", str_target):
			raise Exception("[E_UTS_PCL] Array element {} is not numeric.".format(str_target))
		# Skip if there is 2-demension array
		if search(r"\[\d+\]\[\d+\]", str_target):
			continue

		class_re = search(r"^(.+)\[(\d+)(\.{3}|\…)(\d+)\](.*)", str_target)
		if class_re:
			str_prestring = class_re.group(1)
			int_start_index = int(class_re.group(2))
			int_end_index = int(class_re.group(4))
			str_substring = class_re.group(5)

			int_skip = int_end_index - int_start_index + 1
			for int_temp_index in reversed(range(int_start_index, int_end_index + 1)):
				str_appendix = "{}[{}]{}".format(str_prestring, int_temp_index, str_substring)

				for int_row in range(len(list_pcl)):
					if int_row == 0:
						str_insert = str_appendix
					else:
						str_insert = list_pcl[int_row][int_col]
					list_pcl[int_row].insert(int_col + 1, str_insert)

			for int_row in range(len(list_pcl)):
				list_pcl[int_row].pop(int_col)

	return list_pcl


## This function is used to create missing element of Array in data table
# @param  list_pcl  Data table
# @return list_pcl  Data table with array modified
def __handle_array_init_value(list_pcl):
	list_no_need_insert_element = []
	int_skip = 0

	def __not_null_column(list_pcl, int_col):
		list_address = [g_str_NULL, g_str_NOT_NULL]
		for int_row in range(1, len(list_pcl)):
			if list_pcl[int_row][int_col].strip() in list_address:
				return True
		return False

	# Get element value from parent data (int_row, int_col)
	def __get_init_value(list_pcl, int_row, int_col, int_element):
		str_init = list_pcl[int_row][int_col].strip()
		if search(r"^\{", str_init) and search(r"\}$", str_init):
			str_init = sub(r"^\s*\{\s*", "", str_init)
			str_init = sub(r"\s*\}\s*$", "", str_init)
			list_init = split(r"\s*,\s*", str_init)
			if len(list_init) == 1:
				return list_init[0]
			elif len(list_init) > int_element:
				return list_init[int_element]
			else:
				return "0"
		elif search(r"^\"", str_init) and search(r"\"$", str_init):
			return "-"
		else:
			return str_init

	for int_col, str_target in enumerate(list_pcl[0]):
		if int_skip > 0:
			int_skip -= 1
			continue

		str_target = str_target.replace(" ", "")
		# Skip if there is 2-demension array
		if search(r"\[\d+\]\[\d+\]", str_target):
			continue

		class_re = search(r"^(.+)\[(\d+)\]$", str_target)
		if class_re:
			str_prestring = class_re.group(1)
			int_index = int(class_re.group(2))
			# Get array name
			str_temp = sub("\[\d+\]", "[]", str_target)
			# If index equal 0 then append to storage, and skip
			if int_index == 0:
				list_no_need_insert_element.append(str_temp)
				continue
			elif str_temp in list_no_need_insert_element:
				# There is not parent array to insert element, skip
				continue
			elif __not_null_column(list_pcl, int_col):
				continue
			else:
				int_skip += int_index
				# Create element for array
				for int_i in reversed(range(int_index)):
					str_target_new = "{}[{}]".format(str_prestring, int_i)
					for int_row in range(len(list_pcl)):
						if int_row == 0:
							list_pcl[int_row].insert(int_col + 1, str_target_new)
						else:
							list_pcl[int_row].insert(int_col + 1, __get_init_value(list_pcl, int_row, int_col, int_i))

				for int_row in range(1, len(list_pcl)):
					if search(r"^\"", list_pcl[int_row][int_col]) and search(r"\"$", list_pcl[int_row][int_col]):
						# If no string value in cell then leave
						continue
					else:
						# Parent cell set to "-"
						list_pcl[int_row][int_col] = "-"

				# For string value in cell, remove the master index
				# Ex: string[6] = "hello" -> string = "hello"
				for int_row in range(1, len(list_pcl)):
					if search(r"^\"", list_pcl[int_row][int_col]) and search(r"\"$", list_pcl[int_row][int_col]):
						list_pcl[0][int_col] = sub(r"\[\d+\]$", "", list_pcl[0][int_col])
						break

				int_sub_skip = 0
				for int_sub_col, str_sub_target in enumerate(list_pcl[0]):
					if int_sub_col > int_col + int_index:
						if int_sub_skip > 0:
							int_sub_skip -= 1
							continue

						str_patn = r"(.*){}(.*)".format(str_target.replace('[', '\[').replace(']', '\]'))
						class_re = search(str_patn, str_sub_target)
						if class_re:
							str_sub_prestring = class_re.group(1)
							if not search(r"\.$", str_sub_prestring) and str_sub_prestring != "":
								continue

							str_sub_sufstring = class_re.group(2)

							int_sub_skip = int_index
							for int_i in reversed(range(int_index)):
								str_target_new = "{}{}[{}]{}".format(str_sub_prestring, str_prestring, int_i, str_sub_sufstring)
								for int_row in range(len(list_pcl)):
									if int_row == 0:
										list_pcl[int_row].insert(int_sub_col + 1, str_target_new)
									else:
										list_pcl[int_row].insert(int_sub_col + 1, __get_init_value(list_pcl, int_row, int_sub_col, int_i))

							for int_row in range(1, len(list_pcl)):
								if "\"" in list_pcl[int_row][int_sub_col]:
									continue
								else:
									list_pcl[int_row][int_sub_col] = "-"

	return list_pcl


## This function is used to find a parent name of a structre/array
# @param  class_worksheet    Worksheet object
# @param  class_dataspecinfo Data spec column of Test spec"s information
# @param  int_cur_col        Current column in Test spec
# @return str_param_name     Name of this element
def __find_parent_param_name(class_worksheet, class_dataspecinfo, int_cur_col):
	# Get obj from Data spec
	class_type = class_dataspecinfo.class_type
	class_param = class_dataspecinfo.class_param_name

	int_cur_row, str_cur_type = f_type_row(class_worksheet, int_cur_col, 0, class_type.int_row, class_param.int_row)
	str_cur_param = testdata_get_cell_value(class_worksheet, class_param.int_row, int_cur_col)

	if int_cur_row == class_param.int_row:
		raise Exception("[E_UTS_PCL] Sheet {}: column {}: is empty type.".format(class_worksheet.title, get_column_letter(int_cur_col)))

	str_result = str_cur_param

	int_temp_row = int_cur_row - 1
	while int_temp_row >= class_type.int_row:
		int_temp_col = int_cur_col - 1
		while int_temp_col > class_type.int_col:
			str_temp_type = testdata_get_cell_value(class_worksheet, int_temp_row, int_temp_col)
			str_temp_param = testdata_get_cell_value(class_worksheet, class_param.int_row, int_temp_col)

			# Parent found
			if str_temp_type != None:
				str_result = str_temp_param + "." + str_result
				break

			int_temp_col -= 1
		int_temp_row -= 1

	return str_result


## This function is used to get the validation description used in Test log
# @param  class_worksheet            Worksheet object
# @param  int_cur_col                Current column in Test spec
# @param  class_dataspecinfo         Data spec column of Test spec"s information
# @param  str_cur_item               Current Item description
# @param  str_cur_meaning            Current Meaning description
# @return str_validation_description Validation description
def __get_validation_description(class_worksheet, int_cur_col, class_dataspecinfo, str_cur_item, str_cur_meaning):
	str_validation_description = __find_parent_param_name(class_worksheet, class_dataspecinfo, int_cur_col)

	class_re = search(r"(\*)?(\w+)(\[\d*\])?\.(\w+)(\[\d+\])?", str_validation_description)
	if class_re:
		str_star = class_re.group(1)
		str_left = class_re.group(2)
		str_index = class_re.group(3)
		str_right = class_re.group(4)
		if str_right == str_left:
			str_left = "{}{}{}.".format(str_star if str_star else "",
										 str_left,
										 str_index if str_index else "")
			str_validation_description = str_validation_description.replace(str_left, "")

	# If Local function, get meaning
	if (str_cur_item == uts_TestExtract.str_LOCAL_FUNCTION):
		str_patt = r"(\w+)\(\)"
		class_re = search(str_patt, str_cur_meaning)
		if (class_re):
			str_validation_description = class_re.group(1) + "." + str_validation_description
		else:
			raise Exception("[E_UTS_PCL] Sheet {}: Column {}: Meaning is incorrect format.".format(class_worksheet.title, get_column_letter(int_cur_col)))

	return str_validation_description


## This function is used to conver a hexa from string to a integer written in string
# @param  str_data A hexa string
# @return str_data An int string
def __convert_hexa_to_int(str_data):
	# Change str_data to decimal number
	# Example:
	# 0x000a / 0x000A / 0X000a / 0x000au / 0x000aU -> 10
	# 10 / 10u / 10U -> 10
	# 0xffffffff -> 4294967295
	# 0xffffffffffffffff -> 18446744073709551615
	str_patn_hexa_prefix = r"^0[xX](?=[0-9a-fA-F])"
	str_patn_num_suffix = r"(?<=[0-9a-fA-F])[uU]$"
	if search(str_patn_hexa_prefix, str_data):
		# When 0x is occur, remove 0x and u or U
		str_data = sub(str_patn_hexa_prefix, "", str_data)
		str_data = sub(str_patn_num_suffix, "", str_data)
		# The build-in function int() is convert hex to unsigned decimal
		str_data = str(int(str_data, 16))
	else:
		if search(r"^\d+[uU]$", str_data):
			str_data = sub(str_patn_num_suffix, "", str_data)
		else:
			str_data = str_data
	return str_data


## This function is used to get data table from Test spec
# @param  class_worksheet    Worksheet object
# @param  list_testcaseid    List of TestCaseID information
# @param  class_dataspecinfo Data spec column of Test spec"s information
# @return list_pcl           Data table
def __get_list_pcl(class_worksheet, list_testcaseid, class_dataspecinfo):
	# Create pcl table
	list_pcl = [None]*(len(list_testcaseid) + 1)

	# Create first, second column of table
	list_pcl[0] = ["testcaseid"]
	for int_row in range(1, len(list_pcl)):
		# Get testcaseid
		class_testcaseid = list_testcaseid[int_row - 1]
		str_testcaseid = class_testcaseid.str_value
		# # Add to pcl table
		list_pcl[int_row] = [str_testcaseid]

	# Check for "/O" in IO row
	str_cur_item = "None"
	str_cur_meaning = "None"

	int_next_col = class_dataspecinfo.class_io.int_col + 1
	class_cell = class_worksheet.cell(class_dataspecinfo.class_io.int_row, int_next_col)
	while class_cell.value != None:
		if (str(class_cell.value) == uts_TestExtract.str_IS_IO_OUTPUT):
			int_cur_col = class_cell.column
			# Get item value with current col and item row
			str_item = testdata_get_cell_value(class_worksheet, class_dataspecinfo.class_item.int_row, int_cur_col)
			if (str_item is not None):
				str_cur_item = str(str_item)
			# Get meaning value with current col and meaning row
			str_meaning = testdata_get_cell_value(class_worksheet, class_dataspecinfo.class_meaning.int_row, int_cur_col)
			if (str_meaning is not None):
				str_cur_meaning = str(str_meaning)

			# Get log description from PCL
			str_validation_description = __get_validation_description(class_worksheet, int_cur_col, class_dataspecinfo, str_cur_item, str_cur_meaning)
			list_data = []
			list_data.append(str_validation_description)

			for class_testcaseid in list_testcaseid:
				str_data = testdata_get_cell_value(class_worksheet, class_testcaseid.int_row, int_cur_col)
				str_data = str_data.strip()

				if search(r"^\"", str_data) and search(r"\"$", str_data):
					str_data = sub(r"\\0.*\"", "\"", str_data)

				list_data.append(str_data)
			# Append description and data to table
			for int_index in range(len(list_pcl)):
				list_pcl[int_index].append(deepcopy(list_data[int_index]))
		int_next_col += 1
		class_cell = class_worksheet.cell(class_dataspecinfo.class_io.int_row, int_next_col)

	# Append table when meet array format
	def is_remain_three_dot(list_pcl):
		for str_target in list_pcl[0]:
			if search(r"^\[\d+\]$", str_target):
				raise Exception("[E_UTS_PCL] {}: {}: array format error.".format(class_worksheet.title, str_target))

			if search(r"\[\s*\d+\s*(\.{3}|\…)\s*\d+\s*\]", str_target):
				return True

		return False

	while is_remain_three_dot(list_pcl):
		list_pcl = __handle_array_three_dot(list_pcl)

	list_pcl = __handle_array_init_value(list_pcl)

	# Conver data to hexa
	for int_row in range(1, len(list_pcl)):
		for int_col in range(len(list_pcl[0])):
			list_pcl[int_row][int_col] = __convert_hexa_to_int(list_pcl[int_row][int_col])

	return list_pcl


## This function is used to get empty data table from Test spec about testcaseid, feature, testtype, resutl
# @param  class_worksheet Worksheet object
# @param  list_testcaseid List of TestCaseID information
# @return list_info       Data table
def __get_list_info(class_worksheet, list_testcaseid):
	list_info = []

	for class_testcaseid in list_testcaseid:
		# Get testcaseid
		str_testcaseid = class_testcaseid.str_value
		# Get feature
		str_feature = testdata_get_cell_value(class_worksheet, class_testcaseid.int_row, class_testcaseid.int_col + 1)
		# Get target device
		str_device = testdata_get_cell_value(class_worksheet, class_testcaseid.int_row, class_testcaseid.int_col + 2)
		# Get testtype
		str_testtype = testdata_get_cell_value(class_worksheet, class_testcaseid.int_row, class_testcaseid.int_col + 3)
		# Add to pcl table
		list_info.append([str_testcaseid, str_feature, str_testtype, str_device, "N/A", "N/A", "N/A", "-", "-"])

	list_description = ["testcaseid", "feature", "testtype", "device", "testresult", "compareresult", "finalresult", "logfiles", "remarks"]
	list_info.insert(0, list_description)

	return list_info


## This function is used to store data table into list of data table present
# @param  list_tableinfo List of data table
# @param  str_function   Target function
# @param  list_table     Data table that want to store
# @param  bool_is_output A flag. True then write to PCL data table, False then write to Log data table
# @return list_tableinfo List of data table
def __store_list_expected(list_tableinfo, str_function, list_table, bool_is_output):
	# Check, Is sheet/function duplicated?
	int_exist_index = None
	for int_index, class_tableinfo in enumerate(list_tableinfo):
		if (class_tableinfo.str_function == str_function):
			int_exist_index = int_index
			break

	# If sheet is not duplicated then append new table obj
	# Else then save to existed function/sheet
	if (int_exist_index is None):
		class_tableinfo = TableInfo()
		class_tableinfo.str_function = str_function
		class_tableinfo.list_pcl = []
		class_tableinfo.list_log = []
		if (bool_is_output):
			class_tableinfo.list_pcl = deepcopy(list_table)
		else:
			class_tableinfo.list_log = deepcopy(list_table)
		list_tableinfo.append(deepcopy(class_tableinfo))
	else:
		if (bool_is_output):
			list_tableinfo[int_exist_index].list_pcl = deepcopy(list_table)
		else:
			list_tableinfo[int_exist_index].list_log = deepcopy(list_table)
	return list_tableinfo


## This function is used to store data table of information into list of data table present
# @param  list_tableinfo List of data table
# @param  str_function   Target function
# @param  list_table     Data table that want to store
# @return list_tableinfo List of data table
def __store_list_info(list_tableinfo, str_function, list_table):
	for int_index, class_tableinfo in enumerate(list_tableinfo):
		if (class_tableinfo.str_function == str_function):
			list_tableinfo[int_index].list_info = deepcopy(list_table)
	return list_tableinfo


## This function is used to store data table error into list of data table present
# @param  list_tableinfo List of data table
# @param  str_function   Target function
# @param  list_table     Data table that want to store
# @return list_tableinfo List of data table
def __store_list_err(list_tableinfo, str_function, list_table):
	for int_index, class_tableinfo in enumerate(list_tableinfo):
		if (class_tableinfo.str_function == str_function):
			list_tableinfo[int_index].list_err = deepcopy(list_table)
	return list_tableinfo


## This function is used to create an empty data table of log from present table
# @param  list_table Data table that want prefer
# @return lsit_log   List of data table of log
def __get_list_empty_log(list_table):
	# Create empty table
	list_log = []

	# Create empty table with format of pcl table
	for int_row, list_col in enumerate(list_table):
		if (int_row == 0):
			list_log.append(deepcopy(list_col))
		else:
			list_row = []
			for int_col, str_cell in enumerate(list_col):
				if (int_col != 0):
					str_cell = "-"
				list_row.append(str_cell)
			list_log.append(deepcopy(list_row))

	return list_log


## This function is used to create an empty data table of error
# @param  list_testcaseid List of testcaseid information
# @return lsit_err        List of data table of error info
def __get_list_empty_err(list_testcaseid):
	list_err = []

	list_testcaseid = [obj.str_value for obj in list_testcaseid]
	list_err.append(["testcaseid", ])
	for str_testcaseid in list_testcaseid:
		list_err.append([str_testcaseid, ])

	return list_err


## This function is used to select table
# @param  class_tableinfo Data table information
# @param  int_tabletype   Table macro
# @return list_data       Data table selected
def __select_table(class_tableinfo, int_tabletype):
	if int_tabletype == int_PCLTABLE:
		list_data = class_tableinfo.list_pcl
	elif int_tabletype == int_LOGTABLE:
		list_data = class_tableinfo.list_log
	elif int_tabletype == int_INFOTABLE:
		list_data = class_tableinfo.list_info
	elif int_tabletype == int_ERRTABLE:
		list_data = class_tableinfo.list_err
	else:
		list_data = None
	return list_data


## This function is used to fill data to Data table information
# @param  class_tableinfo Data table information
# @param  int_tabletype   Table macro
# @param  str_description Test validation description
# @param  str_testcaseid  Test case ID
# @param  str_value       Data value
# @return list_data       Data table selected
def __fill_data_to_tableinfo(class_tableinfo, int_tabletype, str_description, str_testcaseid, str_value):
	str_function = class_tableinfo.str_function
	list_data = __select_table(class_tableinfo, int_tabletype)

	int_max_row = len(list_data)
	int_max_col = len(list_data[0])

	int_cur_col = -1
	int_cur_row = -1

	for int_col in range(1, int_max_col):
		str_curr_description = list_data[0][int_col]
		if str_curr_description == str_description:
			int_cur_col = int_col

	if int_cur_col == -1:
		raise Exception("[E_LOG] func {}: not found in PCL".format(str_function))

	for int_row in range(1, int_max_row):
		str_curr_testcaseid = list_data[int_row][0]
		if str_curr_testcaseid == str_testcaseid:
			int_cur_row = int_row

	if int_cur_row == -1:
		raise Exception("[E_LOG] func {}: testcaseid {} not found in PCL".format(str_function, str_testcaseid))

	list_data[int_cur_row][int_cur_col] = str(str_value)

	return class_tableinfo


## This function is used to get data from Data table information
# @param  class_tableinfo Data table information
# @param  int_tabletype   Table macro
# @param  str_description Test validation description
# @param  str_testcaseid  Test case ID
# @return str_value       Data value
def __get_data_from_tableinfo(class_tableinfo, int_tabletype, str_description, str_testcaseid):
	str_function = class_tableinfo.str_function
	list_data = __select_table(class_tableinfo, int_tabletype)

	int_max_row = len(list_data)
	int_max_col = len(list_data[0])

	int_cur_col = -1
	int_cur_row = -1

	for int_col in range(1, int_max_col):
		str_cur_description = list_data[0][int_col]
		if str_cur_description == str_description:
			int_cur_col = int_col

	if int_cur_col == -1:
		raise Exception("[E_LOG] func {}: not found in PCL".format(str_function))

	for int_row in range(1, int_max_row):
		str_cur_testcaseid = list_data[int_row][0]
		if str_cur_testcaseid == str_testcaseid:
			int_cur_row = int_row

	if int_cur_row == -1:
		raise Exception("[E_LOG] func {}: testcaseid {} not found in PCL".format(str_function, str_testcaseid))

	str_value = list_data[int_cur_row][int_cur_col]

	return str_value


## This function is used to insert a column to data table
# @param  class_tableinfo Data table information
# @param  int_tabletype   Table macro
# @param  str_description Test validation description
# @param  str_testcaseid  Test case ID
# @return class_tableinfo New Data table information
def __insert_column_to_tableinfo(class_tableinfo, int_tabletype, str_description, str_testcaseid):
	list_data = __select_table(class_tableinfo, int_tabletype)

	int_max_row = len(list_data)
	int_max_col = len(list_data[0])
	int_cur_col = -1

	for int_col in range(1, int_max_col):
		str_cur_description = list_data[0][int_col]
		if str_cur_description == str_description:
			int_cur_col = int_col

	if int_cur_col == -1:
		list_data[0].append(str_description)
		for int_row in range(1, int_max_row):
			list_data[int_row].append("-")

	return class_tableinfo


## This function is used to get information from Log Information and create data table
# @param  class_tableinfo  Data table information
# @param  list_logdatainfo List of class LogDataInfo
# @return class_tableinfo  New Data table information
def __get_list_log(class_tableinfo, list_logdatainfo):
	# Append log data into log table
	for class_logdatainfo in list_logdatainfo:
		str_testcaseid = class_logdatainfo.str_testcaseid
		str_testresult = class_logdatainfo.str_testresult
		for class_validateinfo in class_logdatainfo.list_validateinfo:
			str_description = class_validateinfo.str_description
			str_value = class_validateinfo.str_macro

			if str_description in class_tableinfo.list_log[0]:
				class_tableinfo = __fill_data_to_tableinfo(class_tableinfo, int_LOGTABLE, str_description, str_testcaseid, str_value)
			else:
				class_tableinfo = __insert_column_to_tableinfo(class_tableinfo, int_ERRTABLE, str_description, str_testcaseid)
				class_tableinfo = __fill_data_to_tableinfo(class_tableinfo, int_ERRTABLE, str_description, str_testcaseid, str_value)

		class_tableinfo = __fill_data_to_tableinfo(class_tableinfo, int_INFOTABLE, "testresult", str_testcaseid, str_testresult)

	return class_tableinfo


## This function is used to a row of exist Test Case ID in Test Report
# @param  class_worksheet Worksheet object
# @param  list_testcase   List of TestCase information
# @return list_testcase   New list of TestCase information
def __testreport_find_testcaseid_row(class_worksheet, list_testcase):
	# Find testcaseid row
	# iter_row = class_worksheet.iter_rows(
	# 	g_int_TR_TESTCASEID_ROW + 1, # min_row
	# 	class_worksheet.max_row, # max_row
	# 	g_int_TR_TESTCASEID_COL, #min_col
	# 	g_int_TR_TESTCASEID_COL) #max_col
	# for iter_col in iter_row:
	# 	for class_cell in iter_col:
	# 		# Loop for list_testcase func
	# 		for int_index, class_testcaseid in enumerate(list_testcase):
	# 			if list_testcase[int_index].int_row is None:
	# 				if str(class_cell.value) == class_testcaseid.str_testcaseid:
	# 					list_testcase[int_index].int_row = class_cell.row
	# 					break
	int_index = 0
	first_value = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW + 1, g_int_TR_TESTCASEID_COL)
	if first_value == None:
		for col in class_worksheet.iter_cols(
			min_row = g_int_TR_TESTCASEID_ROW + 1, # min_row
			max_row = class_worksheet.max_row, # max_row
			min_col = g_int_TR_TESTCASEID_COL, #min_col
			max_col = g_int_TR_TESTCASEID_COL #max_col
		):
			int_index = 0
			for class_cell in col:
				list_testcase[int_index].int_row = class_cell.row
				int_index += 1
				break
	return list_testcase


## This function is used to insert a row for a Test Case ID that not exist in Test Report
# @param  class_worksheet Worksheet object
# @param  list_testcase   List of TestCase information
# @return list_testcase   New list of TestCase information
def __testreport_insert_row(class_worksheet, list_testcase):
	int_next_row = g_int_TR_TESTCASEID_ROW + 1
	# Loop for list_testcase func
	for class_testcase in list_testcase:
		class_testcase.int_row = int_next_row
		int_next_row = class_testcase.int_row + 1
		# int_row = class_testcase.int_row
		# if (int_row is None):
		# 	class_worksheet.insert_rows(int_next_row)
		# 	# Write testcase id to Test report
		# 	class_worksheet.cell(int_next_row, g_int_TR_TESTCASEID_COL, class_testcase.str_testcaseid)
		# 	# Update list_testcase
		# 	# list_testcase = __testreport_find_testcaseid_row(class_worksheet, list_testcase)
		# 	break
		# else:
		# 	int_next_row = class_testcase.int_row + 1

	return list_testcase


## This function is used to check information of list TestCase, Is Test Case ID had a row number in Test Report?
# @param  list_testcase List of TestCase information
# @return bool          True/False
def __testreport_is_all_testcaseid(list_testcase):
	for class_testcase in list_testcase:
		if (class_testcase.int_row is None):
			return False
	return True


## This function is used to write result to Test Report with matched Test Case ID
# @param  list_tableinfo  List of data table
# @param  class_worksheet Worksheet object
# @param  int_result_col  Result column find by the matched Revision
# @return None
def __testreport_write_match_testcaseid(list_tableinfo, class_worksheet, int_result_col):
	list_testcase = []
	int_logfile_col = int_result_col + 1
	int_remarks_col = int_result_col + 2
	int_testcaseid_col = g_int_TR_TESTCASEID_COL
	int_feature_col = g_int_TR_TESTCASEID_COL + 1
	int_testtype_col = g_int_TR_TESTCASEID_COL + 2
	int_device_col = g_int_TR_TESTCASEID_COL + 3

	for class_tableinfo in list_tableinfo:
		for int_row in range(1, len(class_tableinfo.list_info)):
			# Find testcaseid row in Test_Report
			class_testcase = TestCaseInfo()
			class_testcase.str_function = class_tableinfo.str_function
			class_testcase.str_testcaseid = class_tableinfo.list_info[int_row][0]
			class_testcase.str_feature = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "feature", class_testcase.str_testcaseid)
			class_testcase.str_testtype = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "testtype", class_testcase.str_testcaseid)
			class_testcase.str_testresult = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "finalresult", class_testcase.str_testcaseid)
			class_testcase.str_logfiles = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "logfiles", class_testcase.str_testcaseid)
			class_testcase.str_remarks = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "remarks", class_testcase.str_testcaseid)
			class_testcase.str_device = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "device", class_testcase.str_testcaseid)
			class_testcase.int_row = None
			list_testcase.append(deepcopy(class_testcase))

	print("#### [Test Comparison] Writing to test report ...") 
	while True:
		# __testreport_find_testcaseid_row(class_worksheet, list_testcase)
		__testreport_insert_row(class_worksheet, list_testcase)
		bool_is_done = __testreport_is_all_testcaseid(list_testcase)
		if bool_is_done:
			break

	for class_testcase in list_testcase:
		class_worksheet.cell(class_testcase.int_row, int_testcaseid_col, class_testcase.str_testcaseid)
		class_worksheet.cell(class_testcase.int_row, int_feature_col, class_testcase.str_feature)
		class_worksheet.cell(class_testcase.int_row, int_testtype_col, class_testcase.str_testtype)
		class_worksheet.cell(class_testcase.int_row, int_device_col, class_testcase.str_device)
		class_worksheet.cell(class_testcase.int_row, int_result_col, class_testcase.str_testresult)
		class_worksheet.cell(class_testcase.int_row, int_logfile_col, class_testcase.str_logfiles)
		class_worksheet.cell(class_testcase.int_row, int_remarks_col, class_testcase.str_remarks)

	print("#### [Test Comparison] Done writing")


## This function is used to write sequence of number in No. column of Test_Report
# @param  class_worksheet Worksheet object
# @return None
def __testreport_fill_number(class_worksheet):
	int_i = 1
	for int_row in range(g_int_TR_TESTCASEID_ROW + 1, class_worksheet.max_row + 1):
		class_cell = class_worksheet.cell(int_row, g_int_TR_TESTCASEID_COL - 1)
		if class_worksheet.cell(class_cell.row, g_int_TR_TESTCASEID_COL).value != None:
			class_worksheet.cell(class_cell.row, g_int_TR_TESTCASEID_COL - 1, int_i)
			int_i += 1


## This function is used to get the Result column with matched Revision
# @param  str_revision    Source test revision
# @param  class_worksheet Worksheet object
# @return int_col         Result column
def __testreport_get_result_col(str_revision, class_worksheet):
	int_row = g_int_TR_REVISION_ROW
	for int_col in range(g_int_TR_REVISION_COL + 1, class_worksheet.max_column + 1):
		class_cell = class_worksheet.cell(int_row, int_col)
		if search(r"^{}".format(str_revision), str(class_cell.value)):
			int_col = class_cell.column
			# Check format of Test Report
			str_result = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, int_col)
			if str_result != g_str_RESULT:
				raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(int_col), g_str_RESULT))

			str_logfiles = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, int_col + 1)
			if str_logfiles != g_str_LOGFILES:
				raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(int_col) + 1, g_str_LOGFILES))

			str_remarks = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, int_col + 2)
			if str_remarks != g_str_REMARKS:
				raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(int_col) + 2, g_str_REMARKS))

			return int_col

	return None


## This function is used to fill border of Test Report
# @param  class_worksheet Worksheet object
# @return None
def __testreport_fill_border(class_worksheet):
	int_row = g_int_TR_TESTCASEID_ROW + 1
	for int_row in range(g_int_TR_TESTCASEID_ROW + 1, class_worksheet.max_row + 1):
		for int_col in range(g_int_TR_TESTCASEID_COL - 1, class_worksheet.max_column + 1):
			class_cell = class_worksheet.cell(int_row, int_col)
			class_thin = Side(border_style="thin", color="000000")
			class_cell.border = Border(top=class_thin, left=class_thin, right=class_thin, bottom=class_thin)


## This function is used to check format of Test Report
# @param  class_worksheet Worksheet object
# @return None
def __testreport_check_format(class_worksheet):
	str_testcaseid = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, g_int_TR_TESTCASEID_COL)
	if (str_testcaseid != g_str_TESTCASEID):
		raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(g_int_TR_TESTCASEID_COL), g_str_TESTCASEID))

	str_feature = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, g_int_TR_TESTCASEID_COL + 1)
	if (str_feature != g_str_FEATURE):
		raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(g_int_TR_TESTCASEID_COL) + 1, str_feature))

	str_testtype = testdata_get_cell_value(class_worksheet, g_int_TR_TESTCASEID_ROW, g_int_TR_TESTCASEID_COL + 2)
	if (str_testtype != g_str_TESTTYPE):
		raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_TESTCASEID_ROW, get_column_letter(g_int_TR_TESTCASEID_COL) + 2, str_testtype))

	str_revision = testdata_get_cell_value(class_worksheet, g_int_TR_REVISION_ROW, g_int_TR_REVISION_COL)
	if (str_revision != g_str_REVISION):
		raise Exception("[E_UTS_PCL] Sheet {}: row {}: column {}: not found {}".format(class_worksheet.title, g_int_TR_REVISION_ROW, get_column_letter(g_int_TR_REVISION_COL), str_revision))


## This function is used to get the information for test comparison from the config file.
# @param  str_config_path   Path to Config file
# @return class_config_info Data contain in class ConfigInfo
def ut_cmp_configparser(str_config_path):
	str_config_path = abspath(str_config_path)
	if (not exists(str_config_path)):
		raise Exception("[E_UTS_FILE] File {}: is not exist".format(str_config_path))
	uts_TestExtract.testdata_check_file_extension(str_config_path, [".ini"])

	class_config = ConfigParser()
	try:
		class_config.read(str_config_path)
	except:
		raise Exception("[E_UTS_FILE] Could not open file: {}".format(str_config_path))

	# Check exist section/option in config file
	for str_section, list_option in g_dict_OPTION.items():
		for str_option in list_option:
			if not class_config.has_option(str_section, str_option):
				raise Exception("[E_UTS_CONF] Must contain option {} in section [{}]".format(str_option, str_section))

	str_specpath      = uts_TestCode.testcode_convert_python_type(class_config["PATH"]["InputPath"], str)
	uts_TestExtract.testdata_check_file_extension(str_specpath, [".xlsx", ".xlsm", ".xltx", ".xltm"])
	if (not exists(str_specpath)):
		raise Exception("[E_UTS_FILE] File {}: is not exist".format(str_specpath))
	str_logpath       = uts_TestCode.testcode_convert_python_type(class_config["COMPARE"]["LogPath"], str)
	uts_TestExtract.testdata_check_file_extension(str_logpath, [".log", ".txt"])
	if (not exists(str_logpath)):
		raise Exception("[E_UTS_FILE] File {}: is not exist".format(str_logpath))
	str_revision      = uts_TestCode.testcode_convert_python_type(class_config["COMPARE"]["Revision"], str)
	str_reportsheet   = uts_TestCode.testcode_convert_python_type(class_config["COMPARE"]["TestReportSheet"], str)
	bool_findbyprefix = uts_TestCode.testcode_convert_python_type(class_config["TARGET_OF_TESTING"]["FindByPrefix"], bool)
	list_prefix       = uts_TestCode.testcode_convert_python_type(class_config["TARGET_OF_TESTING"]["Prefix"], list)
	list_sheet        = uts_TestCode.testcode_convert_python_type(class_config["TARGET_OF_TESTING"]["SheetNameContainData"], list)

	class_workbook = __get_workbook(str_specpath)
	list_sheet = uts_TestExtract.testdata_get_sheetname(class_workbook.sheetnames, list_prefix, bool_findbyprefix, list_sheet)

	class_workbook.close()
	if (list_sheet == []):
		raise Exception("[E_UTS_CONF] File {}: No target sheet defined".format(basename(str_config_path)))

	class_config_info = ConfigInfo()
	class_config_info.str_specpath = str_specpath
	class_config_info.str_logpath = str_logpath
	class_config_info.str_revision = str_revision
	class_config_info.str_reportsheet = str_reportsheet
	class_config_info.list_sheet = list_sheet

	return class_config_info


## This function is used to Analyze test spec and create expected of output data information from test spec.
# @param  class_config_info Data contain in class ConfigInfo
# @return list_tableinfo    List of class TableInfo. TableInfo contains: Information from Test specification
def ut_cmp_analyze_spec(class_config_info):
	list_tableinfo = []

	class_workbook = __get_workbook(class_config_info.str_specpath)

	for str_sheet in class_config_info.list_sheet:
		class_worksheet = __get_worksheet(class_workbook, str_sheet)

		list_testcaseid = uts_TestExtract.testdata_getTestCaseID(class_worksheet, False)

		class_dataspecinfo = __get_list_dataspecinfo(class_worksheet)

		list_pcl = __get_list_pcl(class_worksheet, list_testcaseid, class_dataspecinfo)
		list_tableinfo = __store_list_expected(list_tableinfo, str_sheet, list_pcl, True)

		list_log = __get_list_empty_log(list_pcl)
		list_tableinfo = __store_list_expected(list_tableinfo, str_sheet, list_log, False)

		list_info = __get_list_info(class_worksheet, list_testcaseid)
		list_tableinfo = __store_list_info(list_tableinfo, str_sheet, list_info)

		list_err = __get_list_empty_err(list_testcaseid)
		list_tableinfo = __store_list_err(list_tableinfo, str_sheet, list_err)

	class_workbook.close()
	return list_tableinfo

## This function is used to Analyze test log and create expected data information from test log.
# @param  class_config_info Data contain in class ConfigInfo
# @param  list_tableinfo    List of class TableInfo
# @return list_tableinfo    List of class TableInfo. TableInfo contains: Information from Test Log
def ut_cmp_analyze_log(class_config_info, list_tableinfo):
	for class_loginfo in __get_list_loginfo(class_config_info.str_logpath):
		str_function = class_loginfo.str_function
		if (str_function not in class_config_info.list_sheet):
			print("[Warning] Test target {}: not found in PCL".format(str_function))
			continue
			# raise Exception("[E_UTS_PCL] Test target {}: not found in PCL".format(str_function))

		for int_index, class_tableinfo in enumerate(list_tableinfo):
			if class_tableinfo.str_function == str_function:
				list_tableinfo[int_index] = __get_list_log(class_tableinfo, class_loginfo.list_logdatainfo)
				break

	return list_tableinfo


## This function is used to Compare expected data information of test spec and test log.
# @param  class_config_info Data contain in class ConfigInfo
# @param  list_tableinfo    List of class TableInfo
# @return list_tableinfo    List of class TableInfo. TableInfo contains: Result of test comparison
def ut_cmp_compare(class_config_info, list_tableinfo):
	# Get target function in log file
	list_func_in_log = []
	list_loginfo = __get_list_loginfo(class_config_info.str_logpath)
	for class_loginfo in list_loginfo:
		str_function = class_loginfo.str_function
		if str_function not in list_func_in_log:
			list_func_in_log.append(str_function)

	for int_index, class_tableinfo in enumerate(list_tableinfo):
		str_function = class_tableinfo.str_function
		list_pcl = class_tableinfo.list_pcl
		list_log = class_tableinfo.list_log
		int_max_row = len(list_pcl)
		int_max_col = len(list_pcl[0])

		# Only target function exist in log file is going to compare
		if str_function in list_func_in_log:
			print("#### [Test Comparison] Start: {}".format(str_function))

			for int_row in range(1, int_max_row):
				str_testcaseid = list_pcl[int_row][0]
				str_testresult = __get_data_from_tableinfo(class_tableinfo, int_INFOTABLE, "testresult", str_testcaseid)

				if str_testresult == "N/A":
					# If testresult of test log is N/A then the test case has not been executed
					# print("Comparison: {}: SKIPPED".format(str_testcaseid))
					str_compareresult = "N/A"
					str_finalresult = "N/A"
					str_logfiles = ""
					str_remarks = ""
				else:
					# Else compare the PCL and Test log
					str_compareresult = "PASSED"
					list_err = []
					int_max_description = 10
					int_max_pcl = 10
					int_max_log = 10

					for int_col in range(1, int_max_col):
						str_description = list_pcl[0][int_col]
						str_pcl = list_pcl[int_row][int_col]
						str_log = list_log[int_row][int_col]

						if str_pcl.isnumeric():
							str_log = __convert_hexa_to_int(str_log)
							list_log[int_row][int_col] = str_log

						if str_pcl != str_log:
							str_compareresult = "FAILED"

							list_err.append({
								"str_description": str_description,
								"str_pcl": str_pcl,
								"str_log": str_log
							})
							if int_max_description < len(str_description):
								int_max_description = len(str_description)
							if int_max_pcl < len(str_pcl):
								int_max_pcl = len(str_pcl)
							if int_max_log < len(str_log):
								int_max_log = len(str_log)

					if str_compareresult == "FAILED":
						str_finalresult = str_compareresult
						str_remarks = "Log not match PCL expected data"
					else:
						str_finalresult = str_testresult
						str_remarks = "Log match PCL expected data"

					print("Comparison: {}: {}".format(str_testcaseid, str_finalresult))

					if str_compareresult == "FAILED":
						print("\tLog not match PCL expected data:")
						print("\t{:<{}} {:<{}} {:<{}}".format("validate", int_max_description, "pcl_value", int_max_pcl, "log_value", int_max_log))
						# print to console compare unmatch
						for dict_err in list_err:
							print("\t{:<{}} {:<{}} {:<{}}".format(dict_err["str_description"], int_max_description, dict_err["str_pcl"], int_max_pcl, dict_err["str_log"], int_max_log))

					str_logfiles = basename(class_config_info.str_logpath)

					list_tableinfo[int_index] = __fill_data_to_tableinfo(class_tableinfo, int_INFOTABLE, "compareresult", str_testcaseid, str_compareresult)
					list_tableinfo[int_index] = __fill_data_to_tableinfo(class_tableinfo, int_INFOTABLE, "finalresult", str_testcaseid, str_finalresult)
					list_tableinfo[int_index] = __fill_data_to_tableinfo(class_tableinfo, int_INFOTABLE, "logfiles", str_testcaseid, str_logfiles)
					list_tableinfo[int_index] = __fill_data_to_tableinfo(class_tableinfo, int_INFOTABLE, "remarks", str_testcaseid, str_remarks)

			print("#### [Test Comparison] End  : {}".format(str_function))


	print("#### [Test Comparison] Summary Compare Result")
	for int_index, class_tableinfo in enumerate(list_tableinfo):
		str_function = class_tableinfo.str_function
		list_info = class_tableinfo.list_info
		int_max_row = len(list_info)
		int_max_col = len(list_info[0])
		int_ok = 0
		int_ng = 0
		int_skip = 0
		int_total = 0

		for int_col in range(1, int_max_col):
			if list_info[0][int_col] == "finalresult":
				break

		for int_row in range(1, int_max_row):
			if list_info[int_row][int_col] == "PASSED":
				int_ok += 1
			elif list_info[int_row][int_col] == "N/A":
				int_skip += 1
			else:
				int_ng += 1

		int_total = int_ok + int_ng + int_skip
		print("{}: Total/OK/NG/SKIP = {}/{}/{}/{}".format(str_function, int_total, int_ok, int_ng, int_skip))

	return list_tableinfo


## This function is used to Write the result to test report.
# @param  class_config_info Data contain in class ConfigInfo
# @param  list_tableinfo    List of class TableInfo
# @return None
def ut_cmp_fill_result(class_config_info, list_tableinfo):
	str_specpath = class_config_info.str_specpath
	str_revision = class_config_info.str_revision
	str_reportsheet = class_config_info.str_reportsheet

	# Write internal data to csv
	if g_bool_DEBUG:
		__testreport_write_to_csv(list_tableinfo)

	if (str_revision == ""):
		print("[Warning] Sheet {}: Not write result because of empty Revision.".format(str_reportsheet))
		return

	class_workbook  = __get_workbook(str_specpath)
	class_worksheet = __get_worksheet(class_workbook, str_reportsheet)

	__testreport_check_format(class_worksheet)

	int_result_col = __testreport_get_result_col(str_revision, class_worksheet)
	if (int_result_col is None):
		raise Exception("[E_UTS_PCL] Sheet {}: Revision {}: is not found".format(str_reportsheet, str_revision))

	__testreport_write_match_testcaseid(list_tableinfo, class_worksheet, int_result_col)
	__testreport_fill_number(class_worksheet)
	__testreport_fill_border(class_worksheet)

	try:
		class_workbook.save(str_specpath)
		class_workbook.save(str_specpath.replace(".xlsx", "_tmp.xlsx"))
		print("#### [Test Comparison] Workbook {} is saved.".format(basename(str_specpath)))
	except:
		print("[Warning] Can not write result to PCL. Because {} is open.".format(basename(str_specpath)))
	class_workbook.close()
	print("------------------------------------------------------")
	print("[E_UTS_OK] Sucessfully comparison")


## This function is an entry of Test Comparison
# @param  str_config_path Path to Config file
# @return None
def uts_cmp_main(str_config_path):
	class_config_info = ut_cmp_configparser(str_config_path)
	list_tableinfo = ut_cmp_analyze_spec(class_config_info)
	list_tableinfo = ut_cmp_analyze_log(class_config_info, list_tableinfo)
	list_tableinfo = ut_cmp_compare(class_config_info, list_tableinfo)
	ut_cmp_fill_result(class_config_info, list_tableinfo)


if __name__ == "__main__":
	if (len(argv) != 2):
		raise Exception("[E_UTS_FILE] Please input config file path")
	uts_cmp_main(argv[1])

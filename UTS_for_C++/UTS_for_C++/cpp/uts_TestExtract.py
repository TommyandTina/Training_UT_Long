##
# @file uts_TestExtract.py
#
# @brief Below is brief information of Test data generation.
#
# @section Description
# This is a sub-module of Test Extraction module.
# It supports generating test data, test case table based on the info in PCL and user inputs.
#
# @section Notes
# - None
#
# @section Restriction
# - None.
#
# @section History
# - 2020.07.06  Create Baseline
#
# Copyright(c) 2020 Renesas Electoronics Corporation

# Imports
import os
import re
import copy
import itertools
import uts_TestCode

# Insert test code API
from uts_TestCode import uts_tsc_write_test_code
from uts_TestCode import testcode_remove_pointer
from uts_TestCode import testcode_remove_array
from uts_TestCode import testcode_convert_string_to_macro
from uts_TestCode import testcode_check_pointer_name

# Library to read excel file
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook

# Library to parsing argument from configure file
import configparser

# Library to parsing arguments
from sys import argv
# Library to set recursive limit of file
from sys import setrecursionlimit

########### Macro used in Test Tool ###########

# String Macros
str_DATA_SPECIFICATION			= "Data specification"
str_TYPE                        = "Type"
str_PARAMETER_NAME				= "Parameter name"
str_FUNCTION_NAME				= "Function name"
str_MEANING						= "Meaning"
str_ITEM						= "Item"
str_IO							= "I/O"
str_PARAMETER_IN 				= "Parameter(In)"
str_PARAMETER_OUT 				= "Parameter(Out)"
str_LOCAL_FUNCTION 				= "Local function"
str_RELATED_GLOBAL_PARAMETER	= "Related global parameter"
str_RETURN 						= "Return"
str_INPUT_AND_OUTPUT_DATA 		= "Input and Output data"
str_TEST_CASE_ID 				= "Test Case ID"
str_No 							= "No."
str_FLAG_TYPE_STR 				= "uint32_t"
str_IS_IO_INPUT 				= "I/"
str_IS_IO_OUTPUT 				= "/O"
str_MACRO_LOCAL_TRUE 			= ", TRUE }"
str_MACRO_LOCAL_FALSE 			= ", FALSE }"
str_VOID_TYPE  					= "void"
str_NAMEFILE_MACRO				= "macro.h"
str_NAMEFILE_FUNCTION			= "functions.h"
str_NAMEFILE_TESTCASE			= "test_case.h"
str_TARGET_DEVICE				= "Device"

int_INIT_VALUE                  = -1
int_MAX_CELL_LENGTH             = 4096

# Macros used to determine I/O type 
int_PARAM_TYPE_INPUT  			= 1	# I/
int_PARAM_TYPE_EXPECT 			= 2	# /O

# Macros used to determine the item type
int_INPUT_TYPE_PARAM  			= 1 # (Input) Parameter (In)
int_INPUT_TYPE_LOCAL_RET  		= 2 # (Input) Local function
int_INPUT_TYPE_RELATED_GLOBAL 	= 3 # (Input) Related global
int_INPUT_TYPE_MEMBER 			= 4 # (Input) Item type is None
int_ITEM_TYPE_LOCAL 			= 1 # (Input) Item type is None, but previous is Local function

int_OUTPUT_TYPE_RET  			= 1 # (Output) Return
int_OUTPUT_TYPE_LOCAL_FUNC  	= 2 # (Output) Local funtion
int_OUTPUT_TYPE_PARAMETER_OUT  	= 3 # (Output) Parameter (Out)
int_OUTPUT_TYPE_RELATED_GLOBAL  = 4 # (Output) Related global
int_OUTPUT_TYPE_MEMBER 			= 5 # (Output) Item type is None

# Macros used to determine the variable type of parameter
int_TYPE_POINTER 				= 1 # Type pointer
int_TYPE_ARRAY 	 				= 2 # Type array
int_TYPE_ARRAYSTRUCT 			= 3 # Type array struct
int_TYPE_STRUCTPOINTER 			= 4 # Type struct pointer
int_TYPE_STRUCT  				= 5 # Type struct
int_TYPE_ARRAYPOINTER			= 6 # Type array pointer
int_TYPE_RETVAL  				= 7 # Type value
int_TYPE_ARR_POINTER_STRUCT		= 8 # Type array pointer struct

# Macros used for handling whitespaces mode of cell's value
int_MODE_STRIP                  = 1 # Remove leading and trailing whitespaces
int_MODE_REPLACE                = 2 # Remove all whitespaces

# Macros used for checking duplicate value 
int_MODE_RETURN 				= 1

# Init global var and get value from config file later
g_SKIPED_COLUMNN    			= 2
g_OFFSET_ROW        			= 2
g_OFFSET_COL        			= 1
g_INT_PARAM_NAME_ROW			= 0
g_INT_PARAM_TYPE_ROW			= 0
g_INT_PARAM_MEANING_ROW			= 0
g_INT_PARAM_ITEM_ROW			= 0
g_INT_PARAM_IO_ROW				= 0
g_BOOL_ENABLE_VALIDATOR			= uts_TestCode.g_bool_enabled_validator
g_NOT_NULL_VALUE 	  			= ''
g_NULL_VALUE 		  			= ''
g_NOT_STUB_VALUE 	  			= ''
g_NOT_CHANGE_VALUE				= ''
g_STR_TARGET_MODULE				= ''
g_INT_TARGET_DEVICE_COL			= 0
g_BOOL_ENABLE_FILE_GENERATE		= True

# List to store information to report global variable used 
# Has format [["type of global variable", "global variable"], ["type of global variable", "global variable"]]
g_LIST_GLOBAL_VARIABLE 	= []
g_OUTPUT_POINTER_NAMES	= []

# Require option in config file
g_DICT_OPTION = {
	'PATH'                      : ['InputPath', 'OutputPath'],
	'TARGET_OF_TESTING'         : ['TargetModule', 'FindByPrefix', 'Prefix', 'SheetNameContainData'],
	'DATA_TYPE'                 : ['ListPointerType'],
	'SPECIAL_STRING'            : ['NOT_NULL_VALUE', 'NULL_VALUE', 'NOT_STUB_VALUE', 'NOT_CHANGE_VALUE'],
	'OUTPUT_GENERATION_CONTROL' : ['EnableImportData', 'EnableTestCode', 'EnableTestData', 'INIT_ADDR_DATA', 'EnableFileGenerate'],
	'VALIDATOR'                 : ['TABLE_VALIDATOR_BY_TYPE', 'TABLE_VALIDATOR_BY_NAME'],
}

# Setting maximum limit for recursive, default is 4000 => 10^6 
setrecursionlimit(10**6)

## This class store information of a cell.
# @param self			The object pointer
# @param int_row		The cell's row
# @param int_col		The cell's column
# @param str_value		The cell's value
class CellInfo:
	## The constructor
	def __init__(self, int_row, int_col, str_value):
		self.int_row = int_row
		self.int_col = int_col
		self.str_value = str_value

## This class store information obtained from config file.
# @param self		The object pointer
class ConfigureInfo:
	## The constructor.
	def __init__(self):
		self.input_path           = ""        # Input path to PCL
		self.output_path          = ""        # Output path to base folder (ut_framework)
		self.TargetModule         = ""        # Target test module (imp, imr, imprtt, impfw, ...)
		self.Prefix               = []        # Prefix of target test module to find sheetnames
		self.FindByPrefix         = True      # Condition to choose whether find all sheetnames by prefix or specified sheetnames
		self.SheetNameContainData = []        # List contain specified sheetnames need to proceed
		self.pointer_type		  = []        # List of pointer typedef in target test module
		self.NOT_NULL_VALUE 	  = ""        # String define valid pointer in PCL
		self.NULL_VALUE 		  = ""        # String define invalid pointer in PCL
		self.NOT_STUB_VALUE 	  = ""        # String define use real function in PCL
		self.NOT_CHANGE_VALUE 	  = ""        # String define not change macro in PCL
		self.EnableTestCode		  = True      # Enable flag to write test code
		self.EnableTestData		  = True      # Enable flag to write test pattern
		self.EnableFileGenerate	  = True	  # Enable flag to generate macro.h, functions.h, testcase.h

	## This method parse data from config file into the class's attributes
	#  @param self					The object pointer
	#  @param config				A ConfigureInfo class object to store information from config file 
	#  @param str_file_name			File name of the config file
	#  @return None
	def __testdata_parse_data(self, config, str_file_name):
		# Get configuration from config file
		try:
			config.read(str_file_name)
		except:
			raise Exception("[E_UTS_FILE] Could not open file: {}".format(str_file_name))

		testdata_check_exist_config_option(config, g_DICT_OPTION)
		self.input_path           = uts_TestCode.testcode_convert_python_type(config['PATH']['InputPath'], str)
		self.output_path          = uts_TestCode.testcode_convert_python_type(config['PATH']['OutputPath'], str)
		self.TargetModule         = uts_TestCode.testcode_convert_python_type(config['TARGET_OF_TESTING']['TargetModule'], str)
		self.FindByPrefix         = uts_TestCode.testcode_convert_python_type(config['TARGET_OF_TESTING']['FindByPrefix'], bool)
		self.EnableTestCode 	  = uts_TestCode.testcode_convert_python_type(config['OUTPUT_GENERATION_CONTROL']['EnableTestCode'], bool)
		self.EnableTestData 	  = uts_TestCode.testcode_convert_python_type(config['OUTPUT_GENERATION_CONTROL']['EnableTestData'], bool)
		self.Prefix               = uts_TestCode.testcode_convert_python_type(config['TARGET_OF_TESTING']['Prefix'], list)
		self.SheetNameContainData = uts_TestCode.testcode_convert_python_type(config['TARGET_OF_TESTING']['SheetNameContainData'], list)
		self.pointer_type 		  = uts_TestCode.testcode_convert_python_type(config['DATA_TYPE']['ListPointerType'], list)
		self.NOT_NULL_VALUE 	  = uts_TestCode.testcode_convert_python_type(config['SPECIAL_STRING']['NOT_NULL_VALUE'], str)
		self.NULL_VALUE 		  = uts_TestCode.testcode_convert_python_type(config['SPECIAL_STRING']['NULL_VALUE'], str)
		self.NOT_STUB_VALUE 	  = uts_TestCode.testcode_convert_python_type(config['SPECIAL_STRING']['NOT_STUB_VALUE'], str)
		self.NOT_CHANGE_VALUE	  = uts_TestCode.testcode_convert_python_type(config['SPECIAL_STRING']['NOT_CHANGE_VALUE'], str)
		self.EnableFileGenerate   = uts_TestCode.testcode_convert_python_type(config['OUTPUT_GENERATION_CONTROL']['EnableFileGenerate'], bool)

	## This method calling the method __testdata_parse_data to get information from config file and store those info into an object ConfigInfo
	#  @param self					The object pointer
	#  @param class_config_info		A ConfigureInfo class object to store information from config file 
	#  @param config_parser			A class object of ConfigParser class implementing interpolation (import module)
	#  @param str_config_path		Path leads to config file located
	#  @return class_config_info	An Ob
	def testdata_parsing_argument(self, class_config_info, config_parser, str_config_path):
		# Parsing data from config file
		self.__testdata_parse_data(config_parser, str_config_path)
		return class_config_info

## This class store information for each parameter defined in PCL.
# @param self		The object pointer
class ParamInfo:
	## The constructor
	def __init__(self):
		self.col              			= 0			# Column define data of param
		self.io_type          			= 0			# Indicate param is input/output of test function
		self.type             			= ""		# String store type of param in PCL
		self.name             			= ""		# String store name of param in PCL
		self.local_param_name 			= ""		# String store name for "Local function" out with format {local_func}_{name}
		self.is_pointer_format      	= False		# Indicate pointer
		self.is_func_param_format   	= False		# Indicate function parameter
		self.is_array_format        	= False		# Indicate array
		self.is_struct_format       	= False		# Indicate struct
		self.flag_large_index			= False		# Indicate whether it is large array
		self.input_type       			= 0			# 
		self.output_type      			= 0			#
		self.openBracket      			= 0			# Num of "{" need to add when write
		self.closeBracket     			= 0			# Num of "}" need to add when write
		self.type_row         			= 0			# Row contain "Type" value
		self.number_of_struct_member 	= 0			# Num of member of struct (only for array)
		self.func_param_type  			= 0			# 
		self.item_type        			= 0			# 
		self.list_large_index			= []		# If large index array, this list will store the index for designated initializers
		self.structure_access           = ""        # Full structure to access to parameter
		self.array_struct_member_range	= 0			# Total length of array struct member

## This class get/write data from PCL into header file (*_PCL.h or *.h).
# @param self				The object pointer
# @param worksheet			The object class: worksheet
# @param str_name			worksheet name
# @param list_pointer_type	Pointer type defined in config file
class WorksheetInfo:
	## The constructor
	def __init__(self, worksheet, str_name, list_pointer_type):
		self.worksheet    		= worksheet         # Object class: worksheet
		self.name         		= str_name          # String: worksheet name
		self.list_pointer_type	= list_pointer_type # List: Pointer type defined in config file
		self.length_validator   = 0                 # Integer: Validator length
		self.int_max_column		= 0

	## This method add open/close brackets for array struct itself or its members
	#  @param self						The object pointer
	#  @param list_param_info			List of ParamInfo
	#  @param list_array_param			List of ParamInfo for array type
	#  @param class_info				ParamInfo class of specific parameter
	#  @param int_offset_range			Offset range to be skipped
	#  @param bool_flag_open_close		Boolean to add open/close bracket at beginning/closing list_param_info 
	#  @return list_param_info			List of ParamInfo after adding brackets
	#  @return list_array_param			List of ParamInfo for array type
	#  @return class_info				ParamInfo class of specific parameter
	def __testdata_add_bracket_for_param_info(self, class_param_parent, list_param_info, list_array_param, class_info, int_offset_range, bool_flag_open_close, bool_in_arrstr):
		# This block will add bracket for member of array struct
		int_next_col = 0

		if len(list_param_info) > 0 and (class_info.col > list_param_info[-1].col):
			class_previous_info = list_param_info[-1]
		else:
			class_previous_info = None

		if bool_flag_open_close:
			# Add open/close bracket for whole array
			# check if first info is struct type
			int_bracket = 1
			for i in range(len(list_param_info)):
				if not list_param_info[i].is_struct_format:
					list_param_info[i].openBracket += int_bracket
					int_bracket = 0
					break
			
			list_param_info[0].openBracket += int_bracket
			list_param_info[-1].closeBracket += 1
		else:
			# Check previous param is struct type and add open bracket if true
			if (class_previous_info != None and class_previous_info.is_struct_format and class_info not in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER, int_TYPE_ARR_POINTER_STRUCT]):
				if (class_info.input_type == int_INPUT_TYPE_MEMBER) or (class_info.output_type == int_OUTPUT_TYPE_MEMBER):	
					class_info.openBracket += 1
			
			# Check next column offset row to add close bracket if this ParamInfo is ending Param
			int_next_col = class_info.col + 1 + int_offset_range
			int_next_type_row, str_next_type = testdata_finding_type_row(self.worksheet, int_next_col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)

			if bool_in_arrstr and (int_next_type_row != class_info.type_row) and \
				(class_param_parent != None and int_next_type_row <= class_param_parent.type_row):
				int_next_offset = class_param_parent.type_row + 1 - class_info.type_row
			else:
				int_next_offset = int_next_type_row - class_info.type_row

				# check if the current param is the last param.
				# In case the parent param has a minus closeBracket, it means the parent is value of a same pointer
				if class_param_parent != None and int_next_type_row < class_param_parent.type_row:
					int_next_offset -= class_param_parent.closeBracket

			int_num_closeBracket = 0
			# Offset smaller than 0 means the end of param
			if int_next_offset < 0:
				if str_next_type != None:
					int_num_closeBracket -= int_next_offset
				# End of PCL
				else:
					int_num_closeBracket -= int_next_offset +  g_INT_PARAM_TYPE_ROW

			if class_previous_info != None:
				if class_previous_info.func_param_type == int_TYPE_STRUCT:
					class_info.openBracket += class_previous_info.openBracket
			class_info.closeBracket += int_num_closeBracket

		# check to add bracket for array param
		if len(list_array_param) > 0:
			# check if first info is struct type
			int_bracket = class_info.openBracket
			if class_info.openBracket > 0:
				for i in range(len(list_array_param)):
					if not list_array_param[i].is_struct_format:
						list_array_param[i].openBracket += int_bracket
						int_bracket = 0
						break
			list_array_param[0].openBracket += int_bracket
			list_array_param[-1].closeBracket += int_num_closeBracket

		# If pointer to array/value, remove redundant bracket
		if class_previous_info != None:
			if len(list_array_param) > 0:
				str_name1 = testcode_remove_pointer(testcode_remove_array(list_array_param[0].name))
				str_name2 = testcode_remove_pointer(testcode_remove_array(list_param_info[-1].name))
				if str_name1 == str_name2:
					int_num_decrease = 1
					if class_info.output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
						str_local_name1 = testcode_remove_pointer(testcode_remove_array(class_info.local_param_name))
						str_local_name2 = testcode_remove_pointer(testcode_remove_array(list_param_info[-1].local_param_name))
						if str_local_name1 != str_local_name2:
							int_num_decrease = 0
					
					if int_num_decrease > 0:
						int_bracket = int_num_decrease
						for i in range(len(list_array_param)):
							if not list_array_param[i].is_struct_format:
								list_array_param[i].openBracket   -= int_bracket
								int_bracket = 0
								break
						
						list_array_param[0].openBracket   -= int_bracket
						list_array_param[-1].closeBracket -= int_num_decrease
			else:
				str_name1 = testcode_remove_pointer(testcode_remove_array(class_info.name))
				str_name2 = testcode_remove_pointer(testcode_remove_array(class_previous_info.name))
				if (str_name1 == str_name2) and class_previous_info.is_pointer_format:
					class_info.openBracket  -= 1
					class_info.closeBracket -= 1

		# Add open bracket for local function input array
		if class_info.input_type == int_INPUT_TYPE_LOCAL_RET and class_info.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER]:
			class_info.openBracket += 1

		return list_param_info, list_array_param, class_info

	## This method checks if parameter type is a pointer or not.
	#  @param self						The object pointer
	#  @param str_type					String parameter type
	# @return is_pointer	Check result in boolean
	# If True, input param type is a pointer. If FaLse, input param type is not a pointer.
	def __testdata_check_pointer_type(self, param_type):
		is_pointer = param_type in self.list_pointer_type
		return is_pointer

	## This method record/report global variable
	#  @param self						The object pointer
	#  @param str_type					String parameter type
	#  @param str_cellValue				String cell's value 
	#  @param bool_get_from_testcode	Boolean to indicate if global variable is obtained in testcode
	#  @return None
	def __testdata_obtain_global_variable(self, str_type, str_cellValue, bool_get_from_testcode):
		# Bool_flag is for indicate whether it is global variable for test code 
		# if bool_flag is True then Test data will self-transform the global variable for reporting 
		if bool_get_from_testcode:
			if ("*" in str_type and "*" in str_cellValue) or (str_cellValue.count("*") > 1) or (str_type.count("*") > 1) or (self.__testdata_check_pointer_type(str_type) and "*" in str_cellValue):
				str_cellValue = "{}{}".format(uts_TestCode.g_str_ENV_GLOBAL_POINTER_VAR_PREFIX, testcode_remove_pointer(testcode_remove_array(str_cellValue)))
			else:
				str_cellValue = "{}{}".format(uts_TestCode.g_str_ENV_GLOBAL_VAR_PREFIX, testcode_remove_pointer(testcode_remove_array(str_cellValue)))
		# List for storing temporary type and global var
		list_temp = []
		# Replace "&" 
		str_global_var = str_cellValue.replace("&", "")
		# If str_cellValue is array type, there are a lot of global variable are stored in str_cellValue
		# So split the str_cellValue by a comma to obtain single value
		list_str = str_global_var.split(",")
		# Loop through the splited string (Now is a list)
		for str_item in list_str:
			# Replace "{" and blank space 
			str_item = str_item.strip("{ ").strip()
			# If the item in splited list has pattern "g_uts" which is a global var, check for existance in the list_temp
			# If it has appended to list_temp, append it
			# This action is for avoiding repeated appending global variable
			if (uts_TestCode.g_str_ENV_GLOBAL_POINTER_VAR_PREFIX in str_item) or (uts_TestCode.g_str_ENV_GLOBAL_VAR_PREFIX in str_item):
				if str_item not in list_temp:
					# Append type into created list 
					list_temp.append(str(str_type).strip())
					list_temp.append(str_item)
		# If the global variable has not appeared in global list, store the list_temp into global list 
		if list_temp not in g_LIST_GLOBAL_VARIABLE and len(list_temp) > 0:
			g_LIST_GLOBAL_VARIABLE.append(list_temp)

	## This method set category type for ParamInfo of each parameter
	#  @param self						The object pointer
	#  @param class_param				Class of ParamInfo
	#  @param int_col					Current column
	#  @return None
	def __testdata_set_category_type(self, list_param_info, class_param, int_col):
		# Init variable 
		str_io_type_value 	= testdata_get_cell_value(self.worksheet, g_INT_PARAM_IO_ROW, int_col, int_MODE_STRIP)
		str_item_value 		= testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_col, int_MODE_STRIP)
		# I/ Type
		if str_io_type_value == str_IS_IO_INPUT:
			class_param.io_type = int_PARAM_TYPE_INPUT
			# Parameter In
			if str_item_value == str_PARAMETER_IN:
				class_param.input_type = int_INPUT_TYPE_PARAM
				class_param.is_func_param_format = True
			# Local function input
			elif str_item_value == str_LOCAL_FUNCTION:
				class_param.input_type = int_INPUT_TYPE_LOCAL_RET
			# Related Global Variable Input
			elif str_item_value == str_RELATED_GLOBAL_PARAMETER:
				class_param.input_type = int_INPUT_TYPE_RELATED_GLOBAL
			# Parameter is a member (not defined "Item")
			elif (str_item_value == None) and (class_param.name != None):
				class_param.input_type = int_INPUT_TYPE_MEMBER
				# Loop backward from the current member column to find whether it is value of a local function input
				int_idx = 1
				while(True):
					# Obtain the previous column "Item"
					str_temp_value = testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_col - int_idx, int_MODE_STRIP)
					# If found value in previous column in "Item" row
					if str_temp_value != None:
						# If the value is "Local Function" => The current column is the value defined for local function input
						if str_temp_value == str_LOCAL_FUNCTION:
							class_param.item_type = int_ITEM_TYPE_LOCAL
							break
						# If the value is not "Local Function" => Break the loop
						else:
							int_current_type_row, _  = testdata_finding_type_row(self.worksheet, class_param.col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
							int_previous_type_row, _ 	 = testdata_finding_type_row(self.worksheet, int_col - int_idx, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
							if int_current_type_row == int_previous_type_row:
								raise Exception("[E_UTS_PCL] The Item information in column {} is missing".format(get_column_letter(class_param.col))) 
							break
					# If not found any value 
					else:
						int_idx += 1
			# If the "Item" value is not belong to any above type, raise Exception 
			else:
				raise Exception("[E_UTS_PCL] Wrong type for item value {} {} at row {} column {}".format(str_item_value, self.name, g_INT_PARAM_IO_ROW, get_column_letter(int_col)))
		
		# /O Type
		elif str_io_type_value == str_IS_IO_OUTPUT:
			class_param.io_type = int_PARAM_TYPE_EXPECT
			# Return value
			if str_item_value == str_RETURN:
				class_param.output_type = int_OUTPUT_TYPE_RET
				bool_check_duplicate_return = testdata_check_duplicate(list_param_info, int_OUTPUT_TYPE_RET, int_MODE_RETURN)
				if bool_check_duplicate_return:
					raise Exception("[E_UTS_PCL] Duplicate Return Value at column {}".format(get_column_letter(class_param.col)))
			# Related Global Variable Output
			elif str_item_value == str_RELATED_GLOBAL_PARAMETER:
				class_param.output_type = int_OUTPUT_TYPE_RELATED_GLOBAL
			# Parameter Out
			elif str_item_value == str_PARAMETER_OUT:
				class_param.output_type = int_OUTPUT_TYPE_PARAMETER_OUT
			# Local Function Output
			elif str_item_value == str_LOCAL_FUNCTION:
				class_param.output_type = int_OUTPUT_TYPE_LOCAL_FUNC
			# Member of parameter
			elif (str_item_value == None) and (class_param.name != None):
				class_param.output_type = int_OUTPUT_TYPE_MEMBER
				int_idx = 1
				while(True):
					# Obtain the previous column "Item"
					str_temp_value = testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_col - int_idx, int_MODE_STRIP)
					# If found value in previous column in "Item" row
					if str_temp_value != None:
						int_current_type_row, _  = testdata_finding_type_row(self.worksheet, class_param.col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
						int_previous_type_row, _ = testdata_finding_type_row(self.worksheet, int_col - int_idx, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
						if int_current_type_row == int_previous_type_row:
							raise Exception("[E_UTS_PCL] The Item information in column {} is missing".format(get_column_letter(class_param.col))) 
						break
					# If not found any value 
					else:
						int_idx += 1
			# If the value is not belonged to above types, raise Exception
			else:
				raise Exception("[E_UTS_PCL] Wrong type for item value {} at row {} column {}".format(str_item_value, g_INT_PARAM_IO_ROW, get_column_letter(int_col)))
		# If I/O value is not "I/" or "/O", raise Exception
		else:
			raise Exception("[E_UTS_PCL] Unrecognized IO type at row {} column {} !!\n".format(g_INT_PARAM_IO_ROW, get_column_letter(int_col)))

	## This method obatin the format of ParamInfo
	#  @param self						The object pointer
	#  @param list_param_info			List of ParamInfo
	#  @param class_param				Class of ParamInfo
	#  @return None
	def __testdata_set_basic_param_type(self, list_param_info, class_param):
		# Boolean flag to check if the parameter is a pointer to array
		bool_check_pointer_to_array  = False
		# Check if param is pointer
		if class_param.name == None:
			raise Exception("[E_UTS_PCL] Not found parameter name at column {}".format(get_column_letter(class_param.col)))

		if class_param.input_type != int_INPUT_TYPE_MEMBER and class_param.output_type != int_OUTPUT_TYPE_MEMBER and \
				re.search(r'^(\[\w*\])+$', class_param.name) != None:
			raise Exception("[E_UTS_PCL] Invalid parameter name at column {}. Name is valid if \"Item\" is empty.".format(get_column_letter(class_param.col)))
		
		# Check if param is array
		# Array format:     a[2], a2[2][2], a[], a[v]
		# Non-array format: a, a[2].b, a2[2][2].b, (incase *a[], *a[5], judge as pointer type)
		# Normal:   a[2], a2[2][2], a[] (function param)
		# Abnormal: a[] (member), a[v]
		obj_match = re.search(r'\w+(((->|\.)\w+)*)((\[(\w*|\w+(\.{3}|\…)\w+)\])+)$', class_param.name)

		# If the value found has array format and it is function parameter (Parameter In)
		if obj_match != None and class_param.is_func_param_format:
			bool_check_pointer_to_array = True
		# If the parameter name starts with "*" or it is pointer to array or it is defined in config file
		# Judge is as a pointer
		if class_param.name.startswith("*") or bool_check_pointer_to_array or class_param.type in self.list_pointer_type or class_param.type.count("*") > 0:
			class_param.is_pointer_format = True
		# Raise exception if it has type a[] (pointer to array) but it is member
		if (class_param.name.endswith("[]") or bool_check_pointer_to_array) and (class_param.input_type == int_INPUT_TYPE_MEMBER or class_param.output_type == int_OUTPUT_TYPE_MEMBER):
			raise Exception("[E_UTS_PCL] Wrong input type for {}".format(class_param.name))
		
		# If it is not pointer to string and has array format
		if obj_match != None:
			# Find all [] in name
			# Array format: a[], a[v], a[v][v]
			list_match = re.findall(r'\[(\w*|\w+(\.{3}|\…)\w+)\]', obj_match.group())
			# Check if array length is number
			for str_match in list_match:
				# If the value inside is a number(integer) and not a pointer to array 
				if re.sub(r'^\[|\]$|\.|\…', "", str_match[0]).isnumeric() and not bool_check_pointer_to_array:
					class_param.is_array_format = True
				# If don't find any value inside square brackets and is is function paramter (Parameter In)
				# Or it is pointer to array
				elif (str_match == "[]" and class_param.is_func_param_format) or bool_check_pointer_to_array:
					pass
				else:
					raise Exception("[E_UTS_PCL] Wrong input information for array length {}".format(class_param.name))

		# Get next param offset row
		int_next_col = class_param.col + 1
		int_offset = testdata_finding_type_row(self.worksheet, int_next_col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)[0]
		int_offset = int_offset - class_param.type_row if int_offset != 0 else 0
		# If next parameter is 1-level lower than current paramter
		# The parameter has a struct format
		if int_offset == 1:
			class_param.is_struct_format = True
		# Next parameter is over 1-level defined type (Not a member format)
		elif int_offset > 1:
			raise Exception("[E_UTS_PCL] Wrong input for struct information {} at column {}".format(class_param.name, get_column_letter(class_param.col)))
		else:
			pass
	
	## This method set macro for array format
	#  @param self						The object pointer
	#  @param class_param				Class of ParamInfo
	#  @param list_param_info			List of ParamInfo
	#  @return None
	def __testdata_set_array_type(self, class_param, list_param_info):
		# If parameter is struct type
		if class_param.is_struct_format and not class_param.is_pointer_format:
			class_param.func_param_type = int_TYPE_ARRAYSTRUCT

			if (class_param.output_type == int_OUTPUT_TYPE_MEMBER) and (len(list_param_info) > 1):
				str_name1 = testcode_remove_pointer(testcode_remove_array(class_param.name))
				str_name2 = testcode_remove_pointer(testcode_remove_array(list_param_info[-1].name))
				if not ((str_name1 == str_name2) and (list_param_info[-1].func_param_type == int_TYPE_STRUCTPOINTER)):
					raise Exception("[E_UTS_PCL] Not support validating for member is Array of struct at column {}".format(get_column_letter(class_param.col)))

		elif class_param.is_pointer_format and not class_param.is_struct_format:
			class_param.func_param_type = int_TYPE_ARRAYPOINTER
		elif class_param.is_struct_format and class_param.is_pointer_format:
			if class_param.input_type == int_INPUT_TYPE_MEMBER or class_param.output_type == int_OUTPUT_TYPE_MEMBER:
				raise Exception("[E_UTS_PCL] Not support for Array of struct pointer at column {}".format(get_column_letter(class_param.col)))
			else:
				class_param.func_param_type = int_TYPE_ARR_POINTER_STRUCT
		else:
			class_param.func_param_type = int_TYPE_ARRAY

	## This method set macro for pointer format
	#  @param self						The object pointer
	#  @param class_param				Class of ParamInfo
	#  @return None		
	def __testdata_set_pointer_type(self, class_param, list_param_info):
		if (len(list_param_info) > 0):
			class_previous_param = list_param_info[-1]
		else:
			class_previous_param = None
		# If parameter is struct type
		if class_param.is_struct_format:
			# Not support for struct pointer is an output member
			if (class_param.output_type == int_OUTPUT_TYPE_MEMBER):
				raise Exception("[E_UTS_PCL] Not supported for output member is a struct pointer at column {}".format(get_column_letter(class_param.col)))
			else:
				class_param.func_param_type = int_TYPE_STRUCTPOINTER
				if (class_previous_param != None):
					if (class_previous_param.func_param_type == int_TYPE_STRUCTPOINTER) and \
						(class_previous_param.input_type == int_INPUT_TYPE_MEMBER or class_previous_param.output_type == int_OUTPUT_TYPE_MEMBER) and \
							(class_param.input_type == int_INPUT_TYPE_MEMBER or class_param.output_type == int_OUTPUT_TYPE_MEMBER):	
						raise Exception("[E_UTS_PCL] Not supported for 2 adjacent member pointer to struct at column {}".format(get_column_letter(class_param.col)))
		else:
			class_param.func_param_type = int_TYPE_POINTER

	## This method create list of ParamInfo for array/array struct/array pointer/ array pointer struct in Designated Initialize format
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param int_arrayLength				Array length
	#  @param list_temp_array_index			List of specific index defined in PCl
	#  @param list_temp_array_index_col		List of specific index's column defined in PCl
	#  @param list_member_param_info		List of member of ParamInfo
	#  @param int_range_of_member			Range of member inside array struct		
	#  @param bool_arr_ptr_struct			Boolean for indicate array pointer struct
	#  @return list_next_param				List of ParamInfo 
	def __testdata_create_param_info_large_index_array(self, class_param, int_arrayLength, list_temp_array_index, list_temp_array_index_col, list_member_param_info, int_range_of_member, bool_arr_ptr_struct):
		list_param_info = []
		if (class_param.func_param_type in [int_TYPE_ARRAYSTRUCT, int_TYPE_ARR_POINTER_STRUCT]) and (len(list_temp_array_index) == 1) and isinstance(list_temp_array_index[0], range) and (list_temp_array_index[0].stop > 1):
			if list_temp_array_index[0].stop > 2:
				list_temp_array_index.append(range(1, list_temp_array_index[0].stop))
			else:
				list_temp_array_index.append(list_temp_array_index[0].start)
			list_temp_array_index[0] = 0
			list_temp_array_index_col.append(list_temp_array_index_col[0])

		list_range = list_temp_array_index

		# Create param info for each range
		for int_idx in range(len(list_range)):
			if class_param.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER] or bool_arr_ptr_struct:
				# Create new ParamInfo based on parent ParamInfo
				class_next_param = copy.deepcopy(class_param)
				class_next_param.type = '' if int_idx != 0 else class_param.type 
				class_next_param.list_large_index = list_range[int_idx]
				list_param_info.append(class_next_param)
			# Array struct
			else:
				# Loop to add ParamInfo for member of array struct
				int_index = 0
				list_temporary_array_struct_param = []
				bool_is_clear_data = False
				while (int_index < int_range_of_member):
					list_temp_next_param = []
					# Create new ParamInfo based on parent ParamInfo
					int_temp_offset_range = 0 
					class_next_param = copy.deepcopy(class_param)
					class_next_param.list_large_index = list_range[int_idx]
					class_next_param.col = class_param.col + int_index + 1

					# Keeping the first index to write, else reset to initial value
					if int_index != 0 and not bool_is_clear_data:
						for class_param_check in list_temporary_array_struct_param:
							if not class_param_check.is_struct_format:
								bool_is_clear_data = True
								break

					if bool_is_clear_data:
						class_next_param.list_large_index = []

					# Initialize again
					class_next_param.name = testdata_get_cell_value(self.worksheet, g_INT_PARAM_NAME_ROW, class_next_param.col, int_MODE_REPLACE)
					if (class_next_param.name != None) and (class_next_param.name.startswith("[")):
						raise Exception("[E_UTS_PCL] Invalid param name at column {}".format(get_column_letter(class_next_param.col)))
					class_next_param.is_array_format = False
					class_next_param.is_pointer_format = False
					class_next_param.is_struct_format = False
					class_next_param.is_func_param_format = False
					class_next_param.number_of_struct_member = 0
					class_next_param.flag_large_index = False if class_next_param.list_large_index == 0 else True

					class_next_param.type_row, class_next_param.type = testdata_finding_type_row(self.worksheet, class_next_param.col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
					int_offset = class_next_param.type_row - class_param.type_row
					# Member of struct
					if int_offset >= 1:
						self.__testdata_set_category_type(list_temporary_array_struct_param, class_next_param, class_next_param.col)
						self.__testdata_set_basic_param_type(list_temporary_array_struct_param, class_next_param) 
						# Set type of param
						class_next_param, list_temp_next_param, int_temp_offset_range = self.__testdata_set_param_type(class_next_param, list_param_info, list_temp_next_param)
						# If first member of array struct is an array, raise Exception for not supporting this case (limitation of the software)
						if (int_index == 0) and (class_next_param.is_array_format):
							raise Exception("[E_UTS_PCL] Not support for array structure has first member is an array format at column {}".format(get_column_letter(class_next_param.col)))
					list_temporary_array_struct_param, list_temp_next_param, class_next_param = self.__testdata_add_bracket_for_param_info(class_param, \
												list_temporary_array_struct_param, list_temp_next_param, class_next_param, int_temp_offset_range, False, True)

					# Append into list for member of a struct
					if len(list_temp_next_param) > 0:
						for item in list_temp_next_param:
							list_temporary_array_struct_param.append(item)
					else:
						list_temporary_array_struct_param.append(class_next_param)
					
					int_index += int_temp_offset_range + 1

				# Addding open/close for each range
				list_temporary_array_struct_param, _, class_param = self.__testdata_add_bracket_for_param_info(class_param, \
					list_temporary_array_struct_param, [], class_param, int_temp_offset_range, True, True)

				# Add member of array struct into large list_param_info
				for class_param_info in list_temporary_array_struct_param:
					list_param_info.append(class_param_info)
		
		# Update column at specific index
		i = 0
		while i < len(list_param_info):
			# Update column at specific index
			if class_param.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER] or bool_arr_ptr_struct:
				int_arr_idx = list_temp_array_index.index(list_param_info[i].list_large_index)
				list_param_info[i].col = list_temp_array_index_col[int_arr_idx]
			else:
				int_arr_idx = list_temp_array_index.index(list_param_info[i].list_large_index)
				for int_index in range(len(list_temporary_array_struct_param)):
					int_offset = list_param_info[i + int_index].col - class_param.col
					list_param_info[i + int_index].col = list_temp_array_index_col[int_arr_idx] + int_offset
				i += len(list_temporary_array_struct_param)
				continue
			i += 1

		# Only add open/close bracket if list_range has more than 1 range 
		if len(list_range) > 1 or bool_arr_ptr_struct or \
			(class_param.func_param_type == int_TYPE_ARRAY or class_param.func_param_type == int_TYPE_ARRAYPOINTER):
			list_param_info, _, class_param = self.__testdata_add_bracket_for_param_info(class_param, list_param_info, [], class_param, 0, True, True)
		# TODO: close bracket is added here, and open braket is added in common write for "list large index"
		elif len(list_range) == 1:
			list_param_info[-1].closeBracket += 1

		return list_param_info

	## This method create list of ParamInfo for array
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @return list_next_param				List of ParamInfo 
	#  @return int_offset_range				Range of column to be skipped
	def __testdata_create_param_info_array(self, class_param):
		list_param_info = []
		# This variable is used to eliminate open bracket when generating Output expected data 2D array

		# Variable count is used for determine dimensions of input array
		obj_match = re.search(r'\w+(((->|\.)\w+)*)((\[(\w*|\w+(\.{3}|\…)\w+)\])+)$', class_param.name)
		match_name = obj_match.group()
		int_count = match_name.count("[")

		# If 2D, 3D, ... array => Generate exact cell value
		if int_count > 1:
			class_param.func_param_type = int_TYPE_RETVAL
			class_param.is_array_format = False
			return [], 0

		# Get the list of defined index array element and its column
		list_temp_array_index = []
		list_temp_array_index_col = []
		list_temp_array_index, list_temp_array_index_col, int_arrayLength = self.__testdata_get_index_array(class_param, class_param.number_of_struct_member)
		
		# Range to skip 
		int_offset_range = len(list_temp_array_index) - 1

		# Indicate large index => Generate following Designated Initializers
		class_param.flag_large_index = True
		list_next_param = self.__testdata_create_param_info_large_index_array(class_param, int_arrayLength, list_temp_array_index, list_temp_array_index_col, list_param_info, 0, False)
		
		return list_next_param, int_offset_range

	## This method create list of ParamInfo for array struct
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @return list_next_param				List of ParamInfo 
	#  @return int_offset_range				Range of column to be skipped
	def __testdata_create_param_info_array_struct(self, class_param):
		# Get the length of struct param if it is a array struct
		int_temp_next_col = class_param.col + 1
		int_temp_idx = 0
		int_temp_idx2 = 0
		int_struct_mem = 0
		bool_end_param_flag = False

		# Finding final column of the parameter
		while(True):
			int_current_type_row = testdata_finding_type_row(self.worksheet, int_temp_next_col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)[0]
			str_param_name = testdata_get_cell_value(self.worksheet, g_INT_PARAM_NAME_ROW, int_temp_next_col, int_MODE_REPLACE)
			str_item       = testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_temp_next_col, int_MODE_STRIP)
			if (str_param_name != None) and (str_param_name.startswith("[")):
				raise Exception("[E_UTS_PCL] Invalid param name at column {}".format(get_column_letter(int_temp_next_col)))
			if str_param_name != None:
				str_check_name1 = testcode_remove_pointer(testcode_remove_array(str_param_name))
				str_check_name2 = testcode_remove_pointer(testcode_remove_array(class_param.name))

			int_offset = int_current_type_row - class_param.type_row
			# Next column is n-level lower
			if int_offset > 0:
				# If it is 1-level lower, than it is member of struct
				if not bool_end_param_flag:
					int_temp_idx += 1
					if (int_offset == 1) and (str_item == None) and (str_check_name1 != str_check_name2):
						class_param.number_of_struct_member += 1
				else:
					int_temp_idx2 += 1
					if (int_offset == 1) and (str_item == None) and (str_check_name1 != str_check_name2):
						int_struct_mem += 1
			# Not found any type		
			elif (int_offset == 0) and (str_item == None) and (str_check_name1 == str_check_name2):
				if bool_end_param_flag:
					if int_temp_idx != int_temp_idx2:
						raise Exception("[E_UTS_PCL] The ranges between array struct ({}) are mismatch!".format(class_param.name))
					if class_param.number_of_struct_member != int_struct_mem:
						raise Exception("[E_UTS_PCL] The truct members between array struct ({}) are mismatch!".format(class_param.name))
					int_temp_idx2 = 0
					int_struct_mem = 0
				else:
					bool_end_param_flag = True
			else:
				break
			
			int_temp_next_col += 1

		if bool_end_param_flag:
			if int_temp_idx != int_temp_idx2:
				raise Exception("[E_UTS_PCL] The ranges between array struct ({}) are mismatch!".format(class_param.name))
			if class_param.number_of_struct_member != int_struct_mem:
				raise Exception("[E_UTS_PCL] The struct members between array struct ({}) are mismatch!".format(class_param.name))
			
		int_final_col = int_temp_next_col	

		# Initialize variables
		list_param_info = []
		list_member_param_info = []
		list_array_index = []
		list_array_index_col = []
		int_offset = 1
		int_offset_range = 0

		if class_param.func_param_type != int_TYPE_ARR_POINTER_STRUCT:
			list_param_info.append(class_param)
		
		# Get the array index and array index column defined in PCL
		# Example [0], [1], [2] 
		list_array_index, list_array_index_col, int_arrayLength = self.__testdata_get_index_array(class_param, int_temp_idx)

		# Designated Initializes format for array struct
		# Example: data[10] = { [0 ... 3] = {a,b}, [4] = [c,d], [5 ... 9] = [a,b] }
		class_param.flag_large_index = True
		# Calculate the range of column to be skipped
		int_offset_range = int_final_col - class_param.col - 1

		# If it is array pointer struct, create list of pointer 
		if class_param.func_param_type == int_TYPE_ARR_POINTER_STRUCT:
			list_param_array_pointer = self.__testdata_create_param_info_large_index_array(class_param, int_arrayLength, list_array_index, list_array_index_col, list_member_param_info, int_temp_idx, True)
			for class_param_array_pointer in list_param_array_pointer:
				list_param_info.append(class_param_array_pointer)	
			# Append 1 more ParamInfo type array struct for Array Pointer Struct type
			class_temp_param = copy.deepcopy(class_param)
			class_temp_param.func_param_type = int_TYPE_ARRAYSTRUCT
			class_temp_param.is_pointer_format = False
			list_param_info.append(class_temp_param)
		# Normal process for array struct generation
		list_temp_param_info = self.__testdata_create_param_info_large_index_array(class_param, int_arrayLength, list_array_index, list_array_index_col, list_member_param_info, int_temp_idx, False)
		
		for class_param_info in list_temp_param_info:
			list_param_info.append(class_param_info)

		return list_param_info, int_offset_range
	
	## This method create list of specific index and its column defined in PCL
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param int_range_of_struct			Array length will be generated
	#  @return int_arrayLength				Array length
	#  @return list_array_index				List of specific index 
	#  @return list_array_index_col			List of specific index column
	def __testdata_get_index_array(self, class_param, int_range_of_struct):
		# Get the list of defined index array element and its column
		list_array_index = []
		list_array_index_col = []
		int_idx = 0
		str_last_name = class_param.name
		int_arrayLength = 0

		# Loop to find next param has the same param type and param name
		int_temp_column = class_param.col
		while(int_temp_column <= self.int_max_column):
			int_next_type_row, str_next_type = testdata_finding_type_row(self.worksheet, int_temp_column, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
			if int_next_type_row == class_param.type_row:
				str_next_name = testdata_get_cell_value(self.worksheet, g_INT_PARAM_NAME_ROW, int_temp_column, int_MODE_REPLACE)
				if (str_next_name != None) and (str_next_name.startswith("[")):
					raise Exception("[E_UTS_PCL] Invalid param name at column {}".format(get_column_letter(int_temp_column)))
				str_check_name1 = testcode_remove_pointer(testcode_remove_array(str_next_name))
				str_check_name2 = testcode_remove_pointer(testcode_remove_array(class_param.name))
				if (str_next_type == class_param.type) and (str_check_name1 == str_check_name2):
					str_next_item = testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_temp_column, int_MODE_STRIP)
					if (int_temp_column != class_param.col) and (str_next_item != None):
						break

					# Store last array index name to get array length
					str_last_name = str_next_name

					# Extract array index
					list_match = re.findall(r'\[\w+', str_next_name)
					if list_match != list([]):
						int_start_index_ = list_match[-1].replace("[","")
					else:
						raise Exception("[E_UTS_PCL] Invalid parameter name for array: {}".format(str_next_name))

					obj_match = re.search(r'\w+\]$', str_next_name)
					if obj_match != None:
						int_end_index_ = obj_match.group().replace("]","")
					else:
						raise Exception("[E_UTS_PCL] Invalid parameter name for array: {}".format(str_next_name))

					# Get array index
					if int_start_index_.isnumeric() and int_end_index_.isnumeric():
						int_start_index_ = int(int_start_index_)
						int_end_index_   = int(int_end_index_)
						if int_start_index_ == int_end_index_:
							int_array_index = int_start_index_
						elif int_start_index_ < int_end_index_:
							int_array_index = range(int_start_index_, int_end_index_ + 1)
						else:
							raise Exception("[E_UTS_PCL] Wrong array index at column {}".format(get_column_letter(int_temp_column)))
					else:
						raise Exception("[E_UTS_PCL] Index at row {} col {} not a number".format(g_INT_PARAM_NAME_ROW, get_column_letter(int_temp_column)))
					
					# Check and get array index
					if isinstance(int_array_index, int):
						list_check_array_index = list([int_array_index])
					else:
						list_check_array_index = int_array_index

					if (len(list_array_index) > 0) and \
							((isinstance(list_array_index[-1], int)   and (list_check_array_index[0] < list_array_index[-1])) or \
							 (isinstance(list_array_index[-1], range) and (list_check_array_index[0] < list_array_index[-1].stop - 1))):
						raise Exception("[E_UTS_PCL] Array index must be in ascending order at column {}".format(get_column_letter(int_temp_column)))

					bool_is_overwrite = False
					for i in range(len(list_check_array_index)):
						if list_check_array_index[i] in list_array_index:
							bool_is_overwrite = True

						if any(list_check_array_index[i] in sublist for sublist in list_array_index if isinstance(sublist, range)):
							bool_is_overwrite = True

						if bool_is_overwrite:
							raise Exception("[E_UTS_PCL] Overwrite/Dupplicate array index at column {}".format(get_column_letter(class_param.col)))
					
					list_array_index.append(int_array_index)
					list_array_index_col.append(int_temp_column)
				else:
					break
			else:
				break

			if int_range_of_struct > 0:
				int_idx += int_range_of_struct + 1
			else:
				int_idx += 1

			int_temp_column = class_param.col + int_idx
						
		# Extract the last index
		obj_match = re.search(r'\w+\]$', str_last_name)
		if obj_match != None:
			if len(list_array_index) == 1:
				int_arrayLength = int(obj_match.group().replace("]",""))
			else:
				int_arrayLength = int(obj_match.group().replace("]","")) + 1
		else:
			raise Exception("[ERROR] Invalid parameter name for array: {}".format(str_last_name))

		if (len(list_array_index) > 1) and (list_array_index[0] != 0):
			raise Exception("[E_UTS_PCL] Must contain column for start index (index 0) if specify array index at column {}".format(get_column_letter(class_param.col)))
		if int_arrayLength <= 0:
			raise Exception("[E_UTS_PCL] Invalid array length (<=0) at column {}".format(get_column_letter(class_param.col)))

		# If initialize data in one column, update the index of array
		if str_last_name == class_param.name:
			list_array_index[0] = range(int_arrayLength)

		# Update param name according to array length instead of first index
		class_param.name = "{}[{}]".format(testcode_remove_array(class_param.name), int_arrayLength)

		return list_array_index, list_array_index_col, int_arrayLength

	## This method get Parameter Name for output local function
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param int_meaning_row				Row of string "Meaning"
	#  @return class_param					Class of ParamInfo 
	def __testdata_get_param_name(self, class_param, int_meaning_row):
		# If Parameter is output local function, get the name for creating function name
		if class_param.output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
			str_meaning_value = testdata_get_cell_value(self.worksheet, int_meaning_row, class_param.col, int_MODE_STRIP)
			bool_found = False
			if str_meaning_value == None:
				raise Exception("[E_UTS_PCL] Missing meaning value for Local Function (Output) at row {} column {}".format(int_meaning_row, get_column_letter(class_param.col)))
			else:
				# Split meaning sentence into words and store into list_value
				list_value = str_meaning_value.split(" ")
				# Loop through each word in list_value
				for str_item in list_value:
					# Format: local_function_out()
					if re.search(r'\w+\(\)', str_item):
						str_local_func = str_item.replace("(","").replace(")", "")
						class_param.local_param_name = '{0}_{1}'.format(str_local_func, class_param.name.replace("*", ""))
						bool_found = True

			if not bool_found:
				raise Exception("[E_UTS_PCL] Missing local function name at row {} column {}". format(int_meaning_row, get_column_letter(class_param.col)))
		return class_param

	## This method for create struct for local function input 
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param str_local_func_data			String for storing data of local function input structure
	#  @return str_local_func_data			String for storing data of local function input structure
	def __testdata_create_struct_local_func(self, class_param, str_local_func_data):
		if class_param.input_type == int_INPUT_TYPE_LOCAL_RET:
			# Enclose previous local function struct
			if str_local_func_data != "":
				str_local_func_data += "\tbool is_stub;\n"
				str_local_func_data += "};\n\n"

			# Add new local function struct
			str_name = class_param.name.replace("*", "")
			str_type = class_param.type
			list_ele = re.findall(r'\[\d+\]', str_name)
			for str_ele in list_ele:
				str_name = str_name.replace(str_ele, "")
			str_local_func_data += "struct {}_info {{\n".format(str_name)

			# Check return not "void"
			if (class_param.type != str_VOID_TYPE) or (class_param.is_pointer_format):
				# If parameter is pointer type
				if class_param.is_pointer_format:
					str_type = str_FLAG_TYPE_STR

				str_local_func_data += "\t{} {}".format(str_type, uts_TestCode.g_str_PATTERN_STUB_RETURN_NAME)
				
				if len(list_ele) > 0:
					str_local_func_data += "{}".format(list_ele[0])

				str_local_func_data += ";\n"
		
		# Add param out
		elif class_param.item_type == int_ITEM_TYPE_LOCAL:
			pos_param_type 	= self.testdata_get_cell(str_DATA_SPECIFICATION, g_OFFSET_ROW, g_OFFSET_COL)
			if class_param.type_row == pos_param_type.int_row:
				str_local_func_data += "\t{} {};\n".format(class_param.type, class_param.name)
		# Other type, do nothing
		else:
			pass

		return str_local_func_data

	## This method for setting macro for ParamInfo
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param list_param_info				List of ParamInfo
	#  @param list_next_param				List of array ParamInfo
	#  @return class_param					Class of ParamInfo
	#  @return list_param_info				List of ParamInfo
	#  @return int_offset_range				Range of column to be skipped
	def __testdata_set_param_type(self, class_param, list_param_info, list_next_param):
		int_offset_range = 0
		# Check array
		if class_param.is_array_format:
			self.__testdata_set_array_type(class_param, list_param_info)
			# Obtain list of param for handling array and offset range is for skip struct array's member column 
			if class_param.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER]:
				list_next_param, int_offset_range = self.__testdata_create_param_info_array(class_param)
			else:
				list_next_param, int_offset_range = self.__testdata_create_param_info_array_struct(class_param)
				for idx in range(len(list_param_info)):
					if list_next_param[idx].func_param_type == int_TYPE_ARRAYSTRUCT:
						list_next_param[idx].array_struct_member_range = len(list_next_param) - idx - 1
						break
		else:
			if class_param.is_pointer_format:
				self.__testdata_set_pointer_type(class_param, list_param_info)
			else:
				if class_param.is_struct_format:
					class_param.func_param_type = int_TYPE_STRUCT
					# Finding member of structure type
					int_temp_column = class_param.col + 1
					while(True):
						int_current_type_row = testdata_finding_type_row(self.worksheet, int_temp_column, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)[0]
						int_offset = int_current_type_row - class_param.type_row
						if int_offset > 0:	
							if int_offset == 1 :
								class_param.number_of_struct_member += 1
						else:
							break		
						int_temp_column += 1	
				else: 
					class_param.func_param_type = int_TYPE_RETVAL	
		return class_param, list_next_param, int_offset_range

	## This method for setting structure access for validate output member of struct
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param int_cur_idx					Current index of ParamInfo in list_param_info
	#  @return list_param_info				List of ParamInfo
	def __testdata_set_structure_access(self, list_param_info, int_cur_idx):
		# Init variable
		class_cur_param_info = list_param_info[int_cur_idx]
		str_raw_param_name   = testcode_remove_pointer(testcode_remove_array(class_cur_param_info.name))
		str_structure_access = str_raw_param_name
		
		if int_cur_idx != 0:
			int_prev_idx            = int_cur_idx - 1
			class_prev_param_info   = list_param_info[int_prev_idx]
			str_prev_raw_param_name = testcode_remove_pointer(testcode_remove_array(class_prev_param_info.name))
			str_prev_struct_access  = class_prev_param_info.structure_access
			str_array_idx           = ""

			# Set array index for array
			if class_cur_param_info.is_array_format:
				str_array_idx = class_cur_param_info.name.replace(str_raw_param_name, "").replace("*", "")

			# Param type row equal "Type" row, do nothing with param name
			if class_cur_param_info.type_row == g_INT_PARAM_TYPE_ROW:
				str_structure_access = str_raw_param_name + str_array_idx
			# Current param type row is higher than previous, get previous structure access and append param name
			elif class_cur_param_info.type_row > class_prev_param_info.type_row:
				# Set delimeter in case pointer to struct
				if class_prev_param_info.is_pointer_format:
					# If pointer to value/array, only keep raw param name
					if str_prev_raw_param_name == str_raw_param_name:
						str_delimeter = ""
						str_prev_struct_access = ""
					else:
						str_delimeter = "->"
				else:
					str_delimeter = "."

				str_structure_access = "{}{}{}".format(str_prev_struct_access, str_delimeter, str_raw_param_name + str_array_idx)
			# Current param type row equal or smaller, find the previous param has the same level, get structure access and change the param name
			else:
				while(int_prev_idx >= 0):
					class_prev_param_info = list_param_info[int_prev_idx]
					if class_cur_param_info.type_row == class_prev_param_info.type_row:
						# Search and change the param name
						str_prev_name = re.search(r'(.+)(\.|->)(\w+$)', class_prev_param_info.structure_access)
						if str_prev_name != None:
							str_structure_access = class_prev_param_info.structure_access.rpartition(str_prev_name.group(3))[0] + str_raw_param_name + str_array_idx
						else:
							str_structure_access = str_raw_param_name + str_array_idx
						break

					int_prev_idx -= 1

		class_cur_param_info.structure_access = str_structure_access
		return list_param_info

	## This method for getting list of ParamInfo of worksheet
	#  @param self							The object pointer
	#  @return list_param_info				List of ParamInfo
	def testdata_get_param_info(self):
		list_param_info = []
		
		# Get start cell for pattern
		pos_param_type 	= self.testdata_get_cell(str_DATA_SPECIFICATION, g_OFFSET_ROW, g_OFFSET_COL)
		pos_param_name 	= self.testdata_get_cell(str_PARAMETER_NAME, 0, g_OFFSET_COL)
		pos_meaning 	= self.testdata_get_cell(str_MEANING, 0, g_OFFSET_COL)
		pos_item		= self.testdata_get_cell(str_ITEM, 0, g_OFFSET_COL)
		pos_io_type	 	= self.testdata_get_cell(str_IO, 0, g_OFFSET_COL)
		pos_input_output_data = self.testdata_get_cell(str_INPUT_AND_OUTPUT_DATA, 0, g_OFFSET_COL)

		# Check if they are in the same column, if not raise exception for wrong format
		if not (pos_param_type.int_col == pos_param_name.int_col == pos_meaning.int_col == pos_item.int_col == pos_io_type.int_col == pos_input_output_data.int_col):
			raise Exception("[E_UTS_PCL] Start cell for pattern is not in the same column")

		# Get end cell
		int_end_col = self.int_max_column
		
		int_col = pos_param_name.int_col
		# Boolean to check if first ParamInfo is 1-level
		bool_first_element = True
		# Loop from start column to end column
		while (int_col < int_end_col):
			# initialize new parameter structure
			class_param = ParamInfo()
			# Get current "type" and type row
			int_current_param_type_row, str_param_type = testdata_finding_type_row(self.worksheet, int_col, int_end_col, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)
			
			# Get value at start cell 
			str_param_name = testdata_get_cell_value(self.worksheet, g_INT_PARAM_NAME_ROW, int_col, int_MODE_REPLACE)

			if ((str_param_type == None) or (str_param_name == None)):
				raise Exception("[E_UTS_PCL] Missing param type/name at column {}".format(get_column_letter(int_col)))
			if str_param_name.startswith("["):
				raise Exception("[E_UTS_PCL] Invalid param name at column {}".format(get_column_letter(int_col)))

			# Init value for param info
			class_param.type     = str_param_type
			class_param.col      = int_col
			class_param.type_row = int_current_param_type_row
			class_param.name     = str_param_name
			list_next_param      = []

			# If not at start column and parameter type row is not "Type" row, raise Exception
			if bool_first_element:
				if pos_param_type.int_col != class_param.col or pos_param_type.int_row != class_param.type_row:
					raise Exception("[E_UTS_PCL] Start parameter is not a parameter function at row {} column {}".format(class_param.type_row, get_column_letter(class_param.col)))
				if testdata_get_cell_value(self.worksheet, g_INT_PARAM_ITEM_ROW, int_col, int_MODE_STRIP) == None:
					raise Exception("[E_UTS_PCL] Start parameter 'Item' is empty at column {}".format(get_column_letter(class_param.col)))
				bool_first_element = False
			
			# Set category, input, output type
			self.__testdata_set_category_type(list_param_info, class_param, int_col)
			self.__testdata_set_basic_param_type(list_param_info, class_param)

			# Set type of param
			class_param, list_next_param, int_offset_range = self.__testdata_set_param_type(class_param, list_param_info, list_next_param)
			
			# find class param parent
			class_param_parent = None
			for i in range(len(list_param_info)):
				if class_param.type_row > list_param_info[-1 - i].type_row:
					class_param_parent = list_param_info[-1 - i]
					break
				elif class_param.col == list_param_info[-1 - i]:
					break

			# Common process for adding bracket
			list_param_info, list_next_param, class_param = self.__testdata_add_bracket_for_param_info(class_param_parent, \
				list_param_info, list_next_param, class_param, int_offset_range, False, False)

			# If array or array struct, append the array to list param_info
			if len(list_next_param) > 0:
				for obj_item in list_next_param:
					obj_item = self.__testdata_get_param_name(obj_item, g_INT_PARAM_MEANING_ROW)
					list_param_info.append(obj_item)
					# Set structure access for parameter
					list_param_info = self.__testdata_set_structure_access(list_param_info, len(list_param_info)-1)
			else:
				class_param = self.__testdata_get_param_name(class_param, g_INT_PARAM_MEANING_ROW)
				list_param_info.append(class_param)
				# Set structure access for parameter
				list_param_info = self.__testdata_set_structure_access(list_param_info, len(list_param_info)-1)

			# Skip the struct member column
			if int_offset_range > 0:
				int_col += int_offset_range
				int_col += 1
			# If no struct member column found, jump to next column
			else: 
				int_col += 1

		return list_param_info

	## This method for writing structure header (*.h)
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param output_path					Location directory to write structure header
	#  @return None
	def testdata_write_struct_test_pattern(self, list_param_info, output_path):
		# Init variable
		str_worksheet_name = self.name
		list_temp_inputType = []
		list_temp_inputName = []
		str_local_func_data = ""
		str_header = ""
		str_header += "#ifndef {}_H_\n".format(str_worksheet_name.upper())
		str_header += "#define {}_H_\n\n".format(str_worksheet_name.upper())
		str_data = ""
		str_data_tmp = ""
		bool_is_found_input = False
		bool_is_found_expect = False
		bool_is_found_data = False

		# Write input and expect struct
		for int_io_type_check in [int_PARAM_TYPE_INPUT, int_PARAM_TYPE_EXPECT]:
			## Write input struct
			if int_io_type_check == int_PARAM_TYPE_INPUT:
				str_item = "input"
			else:
				str_item = "expect"

			# Check whether struct input has data
			if (int_io_type_check == int_PARAM_TYPE_INPUT) and (list_param_info[0].io_type == int_PARAM_TYPE_INPUT):
				str_data_tmp += "struct {}_{} {{\n".format(self.name.strip(), str_item)
			# Check whether struct expect has data, but not only "Return" and "void"
			elif (int_io_type_check == int_PARAM_TYPE_EXPECT) and (list_param_info[-1].io_type == int_PARAM_TYPE_EXPECT) and \
				 	not (list_param_info[-1].type == str_VOID_TYPE and list_param_info[-1].output_type == int_OUTPUT_TYPE_RET and not list_param_info[-1].is_pointer_format):
				str_data_tmp += "struct {}_{} {{\n".format(self.name.strip(), str_item)

			# Loop by number of param
			for int_num in range(len(list_param_info)):
				int_io_type               = list_param_info[int_num].io_type
				bool_is_func_param_format = list_param_info[int_num].is_func_param_format
				bool_is_pointer_format    = list_param_info[int_num].is_pointer_format
				int_input_type            = list_param_info[int_num].input_type
				int_output_type           = list_param_info[int_num].output_type
				int_item_type             = list_param_info[int_num].item_type
				int_func_param_type       = list_param_info[int_num].func_param_type

				# Ensure the write order: struct input -> struct output
				if int_input_type not in [0, int_INPUT_TYPE_MEMBER] or \
						int_output_type not in [0, int_OUTPUT_TYPE_MEMBER] or \
						int_item_type == int_ITEM_TYPE_LOCAL:
					if int_io_type == int_io_type_check:
						# Skip the array member
						if int_num != 0 and list_param_info[int_num].name == list_param_info[int_num-1].name:
							if list_param_info[int_num].local_param_name == list_param_info[int_num-1].local_param_name:
								continue
						# Remove const
						str_type = str(list_param_info[int_num].type).replace("const", "").strip()
						str_name = list_param_info[int_num].name
						str_local_param_name = list_param_info[int_num].local_param_name

						# Check type of I/O 	
						if int_io_type_check == int_PARAM_TYPE_INPUT:
							bool_is_found_input = True
							bool_is_found_data = True
						else:
							if not (str_type == str_VOID_TYPE and int_output_type == int_OUTPUT_TYPE_RET and not bool_is_pointer_format):
								bool_is_found_expect = True
								bool_is_found_data = True
						
						# If Local function out, name with format {local_func}_{name}
						if int_output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
							str_name = str_local_param_name
							if int_func_param_type == int_TYPE_POINTER:
								str_type = str_FLAG_TYPE_STR
													
						str_tmp_name = str_name
						# Change array into pointer
						obj_match = re.search(r'\w+(((->|\.)\w+)*)((\[\w*\])+)$', str_tmp_name)
						if obj_match != None:
							for obj_ele in re.findall(r'\[\d*\]', obj_match.group()):
								str_tmp_name = "*{}".format(str_tmp_name.replace(obj_ele, ""))
						else:
							obj_ele = ""
						
						# If function param, store it
						if bool_is_func_param_format:
							list_temp_inputType.append(str_type)
							list_temp_inputName.append(str_tmp_name)
								
						# Create flag for pointer type
						if bool_is_pointer_format and int_input_type != int_INPUT_TYPE_LOCAL_RET:
							str_name = str_tmp_name.replace("*", "")
							# If struct pointer, create a addtion flag param before it
							if int_func_param_type in [int_TYPE_STRUCTPOINTER, int_TYPE_ARR_POINTER_STRUCT]:
								if int_func_param_type == int_TYPE_ARR_POINTER_STRUCT:
									str_arr_idx = obj_ele
								else:
									str_arr_idx = ""

								str_data_tmp += "\t{} _{}{};\n".format(str_FLAG_TYPE_STR, str_name, str_arr_idx)
								# Get check name
								if int_output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
									str_check_name = testcode_remove_pointer(testcode_remove_array(list_param_info[int_num].name))
								else:
									str_check_name = str_name
								# Check if pointer to value, pointer to array, then set to next type and name
								if str_check_name == testcode_remove_pointer(testcode_remove_array(list_param_info[int_num+1].name)):
									if list_param_info[int_num+1].type != "":
										str_type = list_param_info[int_num+1].type

									if int_output_type != int_OUTPUT_TYPE_LOCAL_FUNC:
										str_name = list_param_info[int_num+1].name
									else:
										str_local_func_name = str_name[:str_name.find(str_check_name)]
										int_index = list_param_info[int_num+1].name.find(str_check_name)
										str_name  = list_param_info[int_num+1].name[:int_index] + str_local_func_name + list_param_info[int_num+1].name[int_index:]

									if int_func_param_type == int_TYPE_ARR_POINTER_STRUCT:
										str_name = str_name.replace("*", "")

							elif int_func_param_type == int_TYPE_ARRAYPOINTER:
								str_type = str_FLAG_TYPE_STR
								str_name = "_{}".format(self.__test_data_update_expected_variable_name(str_name, False) + obj_ele)
							# If pointer, change to a flag
							else:
								str_type = str_FLAG_TYPE_STR
								str_name = "_{}".format(str_name)

						# convert for expected output value
						# var_name->element[n] or var_name[n].element --> var_name_element[n] or var_name_n_element
						# var_name[m][n] --> var_name_m_n
						if int_output_type == int_OUTPUT_TYPE_RELATED_GLOBAL or \
							int_output_type == int_OUTPUT_TYPE_PARAMETER_OUT:
							# update expected variable name and keep the array info
							str_name = self.__test_data_update_expected_variable_name(str_name, False)

						# Return (not pointer)
						elif int_output_type == int_OUTPUT_TYPE_RET:
							if str_type == str_VOID_TYPE:
								continue
						# Create struct for local function in
						elif int_input_type == int_INPUT_TYPE_LOCAL_RET or int_item_type == int_ITEM_TYPE_LOCAL:
							str_local_func_data = self.__testdata_create_struct_local_func(list_param_info[int_num], str_local_func_data)
							if int_input_type == int_INPUT_TYPE_LOCAL_RET:
								str_name = re.sub(r'\*|\[\d+\]', "", str_name)
								str_type = "struct {}_info".format(str_name)
							else:
								continue
						# Other, do nothing
						else:
							pass

						str_data_tmp += "\t{0} {1};\n".format(str_type, str_name)

			# Check whether add struct input/expect
			if (int_io_type_check == int_PARAM_TYPE_INPUT) and bool_is_found_input:
				str_data_tmp += "};\n\n"
			if (int_io_type_check == int_PARAM_TYPE_EXPECT) and bool_is_found_expect:
				str_data_tmp += "};\n\n"

		# Check if any data found
		if bool_is_found_data:
			str_data += str_data_tmp
		
		# Handle for the last local function struct
		if str_local_func_data != "": 
			str_local_func_data += "\tbool is_stub;\n"
			str_local_func_data += "};\n\n"

		## Write struct func_param
		if len(list_temp_inputName) != 0:
			str_data += "struct {}_params {{\n".format(self.name.strip())
			for index in range (len(list_temp_inputName)):
				str_data +=  "\t{0} {1};\n".format(list_temp_inputType[index], list_temp_inputName[index])

			str_data += "};\n\n"

		## Write struct test pattern
		str_data += "struct TEST_{}_Pattern {{\n".format(self.name.strip())
		if bool_is_found_input:
			str_data += "\tstruct {}_input input;\n".format(self.name.strip())
		if bool_is_found_expect:
			str_data += "\tstruct {}_expect expected;\n".format(self.name.strip())
			if g_BOOL_ENABLE_VALIDATOR and self.length_validator > 0:
				str_data += "\t{} validator[{}];\n".format(uts_TestCode.g_str_ENV_VALIDATOR_TYPE, self.length_validator)
		str_data += "};\n\n"

		str_data += "#endif /* {}_H_ */".format(str_worksheet_name.upper())
		str_data += "\n"

		## Write to file
		str_struct_pattern_file = os.path.join(output_path, str_worksheet_name + ".h")
		fp = testdata_get_file(str_struct_pattern_file, "w")
		fp.write(str_header)
		fp.write(str_local_func_data)
		fp.write(str_data)
		fp.close()

	## This method for convert don't care value to 0
	#  @param self							The object pointer
	#  @param str_value						String to be converted
	#  @return str_new_value				String after convert
	def __testdata_convert_dontcare_value(self, str_value):
		obj_match = re.search(r'(\+|-|->)(\s*)(\w+)(\s*)', str_value)
		if obj_match == None:
			str_new_value = str_value.replace("-", "0")
		else:
			str_new_value = str_value

		return str_new_value

	## This method for checking if array's data define in one cell
	#  @param self							The object pointer
	#  @param str_value						String cell's value read from PCL
	#  @param bool_check_large_index		Boolean to check large index type
	#  @return bool_is_define_one_cell		True if array data define in one cell, else False
	def __testdata_is_array_data_one_cell(self, class_info, str_value, bool_check_large_index):
		# Data defined in one cell, conditions:
		# 1. Not have open/close bracket
		# 2. Has multiple value and seperated by comma 
		# 3. Starting/ending with a pair of quotes
		# 4. Don't care value
		str_tmp_value = str_value.replace(" ", "")
		bool_is_initial_array_data  = True if re.search(r'^\{.+\}$', str_tmp_value) != None else False
		bool_is_multiple_value      = True if re.search(r'\w+,\w+', str_tmp_value) != None else False
		bool_is_multiple_value     |= True if re.search(r'(\'|\")\w*(\'|\"),(\'|\")\w*(\'|\")', str_tmp_value) != None else False
		bool_is_multiple_value     |= True if re.search(r'(\'|\")\w*(\'|\"),\w+', str_tmp_value) != None else False
		bool_is_multiple_value     |= True if re.search(r'\w+,(\'|\")\w*(\'|\")', str_tmp_value) != None else False
		bool_is_define_one_cell = (not bool_is_initial_array_data or bool_is_multiple_value) and class_info.is_array_format
		if bool_check_large_index:
			bool_is_define_one_cell &= isinstance(class_info.list_large_index, range) and (list(class_info.list_large_index)[0] == 0)

		return bool_is_define_one_cell

	## This method for writing open/close bracket for input/expect structure
	#  @param self							The object pointer
	#  @param file							File to be written
	#  @param list_to_write					List of data to be written
	#  @param bool_write_open_bracket		Boolean to write open bracket
	#  @param bool_check_input_expect		Boolean to check input or expect type
	#  @return None
	def __testdata_write_bracket_for_struct_input_expect(self, file, list_to_write, bool_write_open_bracket, bool_check_input_expect):
		# Variable to check input or expect type
		# 1: Expect type
		# 0: Input type
		int_temp_var = 1
		if bool_check_input_expect:
			int_temp_var = 0
		# Write open bracket
		if bool_write_open_bracket == True:
			# If struct have more than 1 element
			if (len(list_to_write[0]) > int_temp_var):
				file.write("{ ")
			
		# Write close bracket
		else:
			# If struct have more than 1 element
			if len(list_to_write[0]) > int_temp_var:
				file.write(" }")

	## This method for writing local function input structure
	#  @param self							The object pointer
	#  @param class_info					Class of ParamInfo
	#  @param class_previous_param			Class of previous ParamInfo
	#  @param list_param_info				List of ParamInfo
	#  @param int_index						Current index if ParamInfo in list_param_info
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param list_temporary_input			List to store multiple data #TODO: Remove this list
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @param int_row						Current test case's row
	#  @return str_cellValue				String value to be written into file
	#  @return str_value					String cells' value for later process
	#  @return list_temporary_input			List to store multiple data #TODO: Remove this list
	def __testdata_handle_local_function_input(self, class_info, class_previous_param, list_param_info, int_index, str_value, str_cellValue, list_temporary_input, int_min_row, int_max_row, int_row):
		# Bool_flag == True for Local function input has data
		if str_value != g_NOT_STUB_VALUE:
			str_macro = str_MACRO_LOCAL_TRUE
		else:
			str_macro = str_MACRO_LOCAL_FALSE

		if class_info.type == str_VOID_TYPE and not class_info.is_pointer_format:
			str_value = ""

		if class_info.func_param_type in [int_TYPE_STRUCTPOINTER, int_TYPE_ARR_POINTER_STRUCT]:
			raise Exception("[E_UTS_PCL] Not support for Pointer define data at column {}".format(get_column_letter(class_info.col)))
		# Struct
		elif class_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT]:
			str_cellValue += "{ "
			if list_param_info[int_index + 1].item_type == int_ITEM_TYPE_LOCAL:
				str_cellValue, list_temporary_input = self.__testdata_local_func_find_value(list_param_info, str_value, str_cellValue, list_temporary_input, int_index, int_row, int_min_row, int_max_row, str_macro)
			else:
				str_cellValue += str_macro
		#Array type
		elif class_info.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER]:
			str_cellValue, list_temporary_input = self.__testdata_process_cell_value(list_param_info, class_info, class_previous_param, str_value, str_cellValue, list_temporary_input, int_min_row, int_max_row, int_index)
			# Check next ParamInfo to add close brace
			str_name1 = testcode_remove_pointer(testcode_remove_array(class_info.name))
			str_name2 = testcode_remove_pointer(testcode_remove_array(list_param_info[int_index + 1].name))
			if (str_value != g_NOT_STUB_VALUE) and self.__testdata_is_array_data_one_cell(class_info, str_value, True):
				pass
			else:
				if str_name1 != str_name2:
					str_cellValue += ' }'
			# If next ParamInfo is value of defined local function input, process for adding value to the function
			if list_param_info[int_index + 1].item_type == int_ITEM_TYPE_LOCAL:
				str_cellValue += ", "
				list_temporary_input.clear()
				str_cellValue, list_temporary_input = self.__testdata_local_func_find_value(list_param_info, str_value, str_cellValue, list_temporary_input, int_index, int_row, int_min_row, int_max_row, str_macro.replace("}",""))
			else:
				if str_name1 != str_name2:
					str_cellValue += str_macro.replace("}","")
		# Pointer type
		elif class_info.func_param_type == int_TYPE_POINTER:
			str_cellValue = '{ '
			str_cellValue = self.__testdata_handle_pointer_value(class_info, str_value, str_cellValue, int_min_row, int_max_row)
			if list_param_info[int_index + 1].item_type == int_ITEM_TYPE_LOCAL:
				str_cellValue += ", "
				str_cellValue, list_temporary_input = self.__testdata_local_func_find_value(list_param_info, str_value, str_cellValue, list_temporary_input, int_index, int_row, int_min_row, int_max_row, str_macro)
			else:
				str_cellValue += str_macro
		# Value
		else:
			str_cellValue += "{ "
			str_cellValue += str_value if str_value != '-' else '0'
			if list_param_info[int_index + 1].item_type == int_ITEM_TYPE_LOCAL:
				str_cellValue += ", "
				str_cellValue, list_temporary_input = self.__testdata_local_func_find_value(list_param_info, str_value, str_cellValue, list_temporary_input, int_index, int_row, int_min_row, int_max_row, str_macro)
			else:
				str_cellValue += str_macro

		if class_info.type == str_VOID_TYPE and not class_info.is_pointer_format:
			str_cellValue = str_cellValue.replace(", ", "", 1)

		list_tmp = str_cellValue.split(',')
		for i in range(len(list_tmp)):
			if re.search(r'\[[0-9]*\]\s*=\s*-', list_tmp[i]):
				list_tmp[i] = list_tmp[i].replace('-', '0')
		str_cellValue = ','.join(list_tmp) if len(list_tmp) > 0 else str_cellValue

		return str_cellValue, str_value, list_temporary_input
	
	## This method for writing data header (*_PCL.h)
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param str_output_pathData			Location directory to write data header
	#  @return None
	def testdata_write_test_pattern(self, list_param_info, str_output_pathData):
		# Load workbook and sheet name contains PCL _data
		worksheet_name   = self.name
		list_data_input  = []
		list_data_expect = []

		# Get function name by input argument name_file
		str_define = "#ifndef {0}_PCL_H\n".format(worksheet_name.upper())
		str_define += "#define {0}_PCL_H\n\n".format(worksheet_name.upper())
		str_prefix= '#include "{0}.h"\n\n'.format(worksheet_name)
		str_prefix +=  "static struct TEST_{0}_Pattern {0}_PCL[] =\n{{\n".format(worksheet_name)
		str_postfix="};\n\n" + "#endif\n"

		# Write to Header file (*.h)
		str_file_name = worksheet_name + "_PCL.h"
		str_pattern_file = os.path.join(str_output_pathData, str_file_name)
		_FOUT = testdata_get_file(str_pattern_file, "w")
		_FOUT.write(str_define)
		_FOUT.write(str_prefix)

		# Get start cell
		class_start_param_data = self.testdata_get_cell(str_INPUT_AND_OUTPUT_DATA, g_OFFSET_ROW, g_OFFSET_COL)

		# Get list of test case ID
		list_testcase_id = testdata_getTestCaseID(self.worksheet, True)

		# Get number of test
		int_number_of_test_case = len(list_testcase_id) - 1
		
		# Init variables
		int_min_row = class_start_param_data.int_row
		int_max_row = int_min_row + int_number_of_test_case 
		list_name_input = []
		list_name_expect = []
		bool_tmp = True
		int_validator_index = 0
		int_count_validator_name = 0

		# Loop through each test case
		for int_row in range (int_min_row, int_max_row):
			list_input = []
			list_expect = []
			list_validator = []
			# Reset validator variables for each test case
			int_validator_index = 0
			int_count_validator_name = 0
			# Loop through each param 
			list_temporary_input = []
			int_index = -1
			int_offset_range = 0
			while (int_index < len(list_param_info) - 1):
				# Jumping column
				if int_offset_range > 0:
					int_index += int_offset_range + 1
					int_offset_range = 0
				else:
					int_index += 1	
				# Initial a string variable to store cell value
				str_cellValue = ""
				
				class_info = copy.deepcopy(list_param_info[int_index])
				str_value = testdata_get_cell_value(self.worksheet, int_row, class_info.col, int_MODE_STRIP)

				if int_index != 0:
					class_previous_param = list_param_info[int_index - 1]
				else:
					class_previous_param = None

				# Check the type of cell I/ or O/
				# Input type
				if class_info.io_type == int_PARAM_TYPE_INPUT:
					if class_info.input_type == int_INPUT_TYPE_PARAM or class_info.input_type == int_INPUT_TYPE_RELATED_GLOBAL or class_info.input_type == int_INPUT_TYPE_MEMBER:
						# Share the same flow of handling data
						if class_info.func_param_type == int_TYPE_STRUCT or class_info.item_type == int_ITEM_TYPE_LOCAL or (class_info.input_type == int_INPUT_TYPE_MEMBER and class_info.func_param_type == int_TYPE_STRUCTPOINTER):
							list_param_info[int_index+1].flag_large_index = False
							continue	
						else:
							str_cellValue, list_temporary_input = self.__testdata_process_cell_value(list_param_info, class_info, class_previous_param, str_value, str_cellValue, list_temporary_input, int_min_row, int_max_row, int_index)
						
					elif class_info.input_type == int_INPUT_TYPE_LOCAL_RET:
						str_cellValue, str_value, list_temporary_input = self.__testdata_handle_local_function_input(class_info, class_previous_param, list_param_info, int_index, str_value, str_cellValue, list_temporary_input, int_min_row, int_max_row, int_row)
						
					else:
						raise Exception("[E_UTS_PCL] Wrong type for input function param {} at row {} col {} !!!".format(class_info.input_type, int_row, get_column_letter(class_info.col)))

				# Expect type	
				elif class_info.io_type == int_PARAM_TYPE_EXPECT:
					if (class_info.output_type == int_OUTPUT_TYPE_RET and (class_info.type == str_VOID_TYPE) and (not class_info.is_pointer_format)) or \
						(class_info.output_type == int_OUTPUT_TYPE_MEMBER and ((class_info.func_param_type == int_TYPE_RETVAL and class_info.item_type == int_ITEM_TYPE_LOCAL) or (class_info.func_param_type == int_TYPE_STRUCTPOINTER ))):
							continue
					str_cellValue, list_temporary_input = self.__testdata_process_cell_value(list_param_info, class_info, class_previous_param, str_value, str_cellValue, list_temporary_input, int_min_row, int_max_row, int_index)

				else:
					raise Exception("[E_UTS_PCL] Wrong type {} for {} !!!!!".format(class_info.io_type, class_info.name))
				
				# Skip not write
				if len(str_cellValue) == 0:
					if (class_info.func_param_type in [int_TYPE_STRUCTPOINTER, int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT]) or \
						(class_info.type == str_VOID_TYPE):
						continue
				# Common function to process cell value
				str_cellValue, int_offset_range = self.__testdata_common_Write_action(list_param_info, class_info, str_value, str_cellValue, int_index, int_row)
				if not str_cellValue and (class_info.input_type != int_INPUT_TYPE_MEMBER \
					and (class_info.func_param_type != int_TYPE_STRUCT or class_info.func_param_type != int_TYPE_STRUCTPOINTER)) :
					raise Exception("[E_UTS_PCL] Not found cell value row {} column {}".format(int_row, get_column_letter(class_info.col)))
				# Append cell value
				if str_cellValue:
					if class_info.io_type == int_PARAM_TYPE_INPUT:
						list_input.append(str_cellValue)
					else:
						list_expect.append(str_cellValue)
				# Skip value will be turned into space
				if int_offset_range > 0:
					for index in range(int_offset_range):
						class_temp_info = list_param_info[int_index + index]
						# Not add space for struct because struct/struct pointer type is skipped, not write
						if class_temp_info.func_param_type not in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER, int_TYPE_ARRAYSTRUCT]:
							if class_info.io_type == int_PARAM_TYPE_INPUT:
								list_input.append(" ")
							else:
								list_expect.append(" ")
								list_validator.append(" ")
			# Get name list 
			for i in range(len(list_param_info)):
				class_param = list_param_info[i]
				# Boolean flag for getting name once and not item value of local function input and not a member type
				if bool_tmp and class_param.item_type == 0 and \
						(class_param.input_type != int_INPUT_TYPE_MEMBER or (class_param.input_type == int_INPUT_TYPE_MEMBER and class_param.func_param_type not in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER, int_TYPE_ARR_POINTER_STRUCT])) \
							and (class_param.output_type != int_OUTPUT_TYPE_MEMBER or (class_param.output_type == int_OUTPUT_TYPE_MEMBER and class_param.func_param_type not in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER, int_TYPE_ARR_POINTER_STRUCT])):
					if class_param.io_type == int_PARAM_TYPE_INPUT:
						if class_param.func_param_type not in [int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT] or class_param.input_type == int_INPUT_TYPE_LOCAL_RET:
							list_name_input.append(class_param.name)
					else:
						# Store the output name if:
						# Not "VOID" type or "Void pointer" or "Void type" and not return value
						# Function param is not a struct
						# Member type is not struct/structpointer
						if (class_param.type != str_VOID_TYPE or (class_param.type == str_VOID_TYPE and class_param.func_param_type == int_TYPE_POINTER) or (class_param.type == str_VOID_TYPE and class_param.output_type != int_OUTPUT_TYPE_RET)) and \
							class_param.func_param_type not in [int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT]:
							list_name_expect.append(class_param.name)
				else:
					pass
			
			# Get validator list
			list_validator, int_validator_index, int_count_validator_name = self.__testdata_get_validator_list(list_param_info, int_row)		

			# This flag is used for appending the list of parameter name once
			bool_tmp = False
			list_data_input.append(list_input)
			list_data_expect.append(list_expect)
			if g_BOOL_ENABLE_VALIDATOR and len(list_validator) > 0:
				for item in list_validator:
					list_expect.append(item)
		if g_BOOL_ENABLE_VALIDATOR:
			self.length_validator = int_validator_index
			for i in range(int_count_validator_name):
				list_name_expect.append("validator")
			
		# Insert list of parameter name 
		list_data_input.insert(0, list_name_input)
		list_data_expect.insert(0, list_name_expect)
		# Insert list of test case ID
		for int_temp_index in range(len(list_testcase_id)):
			list_data_input[int_temp_index].insert(0, list_testcase_id[int_temp_index])
		# Find maximum character length for each column => Padding and Aligment
		list_length_alignment_input = [len(max(str_value, key=len)) for str_value in itertools.zip_longest(*list_data_input, fillvalue = " ")] 
		list_length_alignment_expect = [len(max(str_value, key=len)) for str_value in itertools.zip_longest(*list_data_expect, fillvalue = " ")] 
		for int_idx_row in range (len(list_data_input)):
			# Write comment parameter name
			if int_idx_row == 0:
				_FOUT.write("\t/* ")
				# Write test case id
				int_tmp_range = list_length_alignment_input[0] - len(list_data_input[int_idx_row][0])   	# Range of maximum character length in a column
				while int_tmp_range > 0:
					list_data_input[int_idx_row][0] += " "
					int_tmp_range -= 1
				_FOUT.write('"{}", '.format(list_data_input[int_idx_row][0]))
			# Write test cases content
			else:
				_FOUT.write("\t/* ")
				# Write test case id
				int_tmp_range = list_length_alignment_input[0] - len(list_data_input[int_idx_row][0]) 	# Range of maximum character length in a column
				# Add space to make alignment
				while int_tmp_range > 0:
					list_data_input[int_idx_row][0] += " "
					int_tmp_range -= 1
				_FOUT.write("{}".format(list_data_input[int_idx_row][0]))
				_FOUT.write(" */ { ")
				
				# Write open bracket for input struct
				self.__testdata_write_bracket_for_struct_input_expect(_FOUT, list_data_input, True, False)

			# Write input type
			self.__testdata_write_list_to_file(_FOUT, list_length_alignment_input, list_data_input, int_idx_row, 0, True)
			# Write close bracket for input struct
			self.__testdata_write_bracket_for_struct_input_expect(_FOUT, list_data_input, False, False)
			if (len(list_data_input[0]) > 1) and (len(list_data_expect[0]) > 0):
				_FOUT.write(", ")
			# Write open bracket for expect struct
			self.__testdata_write_bracket_for_struct_input_expect(_FOUT, list_data_expect, True, True)				
			# Write expect type
			self.__testdata_write_list_to_file(_FOUT, list_length_alignment_expect, list_data_expect, int_idx_row, int_count_validator_name,False)
			# Write end comment parameter name
			if int_idx_row == 0:
				_FOUT.write(" */\n")
			# Write test case content
			else:
				if int_idx_row == len(list_data_input) - 1:
					_FOUT.write(' }\n')
				else:
					_FOUT.write(' },\n')
			
		_FOUT.write(str_postfix)
		_FOUT.write("\n\n")
		_FOUT.close()

	## This method for writing table of test cases (*.cpp)
	#  @param self							The object pointer
	#  @param str_output_pathTable			Location directory to write table of test cases 
	#  @return None
	def testdata_getTableFile(self, str_output_pathTable, src_language):
		str_name_file    = self.name
		int_idx = 1
		
		# Get list of test case ID
		list_testcase_id = testdata_getTestCaseID(self.worksheet, True)
		list_device		 = testdata_getDevice(self.worksheet)

		# Remove first element which is "Testcase ID"
		list_testcase_id.pop(0)
		if (len(list_testcase_id) != len(list_device)):
			raise Exception("[E_UTS_PCL] Miss-match between number of test case ID and number of Device at sheet {}".format(str_name_file))

		str_prefix = '#include "testenv.h"\n\n'
		str_struct_fix = '\nstruct TestCase UT_{0}_All_Tests[] = {{\n'.format(str_name_file)
		str_postfix = '\tTEST_CASE_END\n};\n'

		# Write to Cpp file (*.cpp)
		if src_language == ".c":
			str_table_file = os.path.join(str_output_pathTable, str_name_file + "_table.c")
		if src_language == ".cpp":
			str_table_file = os.path.join(str_output_pathTable, str_name_file + "_table.cpp")
		FOUT = testdata_get_file(str_table_file, "w")
		FOUT.write(str_prefix)
		for testcase_id in list_testcase_id:
			str_device = (list_device[int_idx - 1].strip())
			# Replace all non-alphanumeric characters in string with underscore
			str_device = re.sub('[^0-9a-zA-Z]+', "_", str_device).upper()
			str_content = 'TEST_CASE_F(UT, {}, {}, {}) '.format(str_name_file, str_device, re.sub("[^0-9a-zA-Z_]", "_", testcase_id))
			str_content += "{{ EXPECT_EQ( true, TEST_{}".format(str_name_file)
			str_content += '("PCL", {})); '.format(int_idx)
			str_content += "}\n"
			FOUT.write(str_content)
			int_idx += 1
		FOUT.write(str_struct_fix)
		int_idx = 0
		for testcase_id in list_testcase_id:
			str_device = (list_device[int_idx].replace(", ", "_")).upper()
			str_content = '\tTEST_CASE_T(UT, {}, {}, {}, {}),\n'.format(str_name_file, str_device, re.sub("[^0-9a-zA-Z_]", "_", testcase_id), testcase_id)
			FOUT.write(str_content)
			int_idx += 1
		FOUT.write(str_postfix)
		FOUT.close()
	
	## This method for create list of dictionary information for TestCode
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param list_param				    List of ParamInfo
	#  @param int_cur_idx					Current index of ParamInfo in list_param
	#  @param dict_pattern_info				Dictionary of data providing to test code
	#  @param str_key_value				    Key of dictionary to be written
	#  @param str_name						Name of ParamInfo
	#  @return dict_pattern_info			Dictionary of data providing to test code
	def __testdata_create_list_of_dict(self, class_param, list_param, int_cur_idx, dict_pattern_info, str_key_value, str_name):
		str_type = class_param.type.replace("const ", "")
		
		if class_param.input_type == int_INPUT_TYPE_LOCAL_RET:
			if testcode_check_pointer_name(class_param.name) or class_param.type != str_VOID_TYPE:
				list_member = [uts_TestCode.g_str_PATTERN_STUB_RETURN_NAME]
			else:
				list_member = []
		else:
			list_member = []

		# Add member
		if (class_param.output_type == int_OUTPUT_TYPE_MEMBER) or (class_param.item_type == int_ITEM_TYPE_LOCAL):
			# Skip in case member is struct format
			if class_param.is_struct_format:
				return dict_pattern_info

			# Get the last key to append
			str_key = list(dict_pattern_info[str_key_value][-1].keys())[0]
			# Get structure access expected output
			if str_key_value == uts_TestCode.g_str_PATTERN_EXPECTED_PARAM_NAME:
				if dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE][-1] == uts_TestCode.g_str_PATTERN_ITEM_LOCAL_FUNC:
					str_first_parent = testcode_remove_pointer(str_key)
				else:
					str_first_parent = testcode_remove_pointer(dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_NAME][-1])
			# Local function input
			else:
				obj_match = re.search(r'((\.|->)\w+)', class_param.structure_access)
				if obj_match != None:
					str_first_parent = class_param.structure_access[:class_param.structure_access.index(obj_match.group(1))]
					if testcode_remove_pointer(testcode_remove_array(str_first_parent)) == testcode_remove_pointer(testcode_remove_array(str_key)):
						str_first_parent == uts_TestCode.g_str_PATTERN_STUB_RETURN_NAME
				else:
					str_first_parent = class_param.structure_access

			str_structure_access = class_param.structure_access.replace(str_first_parent, "", 1)

			obj_match = re.search(r'^(\.|->)', str_structure_access)
			if obj_match != None:
				str_structure_access = str_structure_access.replace(obj_match.group(), "", 1)

			if class_param.item_type == int_ITEM_TYPE_LOCAL:
				str_structure_access = str_first_parent

			if class_param.item_type == int_ITEM_TYPE_LOCAL:
				str_member = "{}".format(str_structure_access)
			else:
				str_member = "{} {}".format(str_type, str_structure_access)

			str_name_check = testcode_remove_pointer(testcode_remove_array(str_name))
			str_key_check  = testcode_remove_pointer(testcode_remove_array(str_key))
			if (str_member not in dict_pattern_info[str_key_value][-1][str_key]) and (str_key_check != str_name_check):
				dict_pattern_info[str_key_value][-1][str_key].append(str_member)
		# Add parent
		else:
			dict_pattern_info[str_key_value].append({str_name : list_member})

		return dict_pattern_info

	## This method for create dictionary information for TestCode
	#  @param self							The object pointer
	#  @param class_param					Class of ParamInfo
	#  @param list_param				    List of ParamInfo
	#  @param int_cur_idx					Current index of ParamInfo in list_param
	#  @param dict_pattern_info				Dictionary of data providing to test code
	#  @return dict_pattern_info			Dictionary of data providing to test code
	def __testdata_create_info_out(self, class_param, list_param, int_cur_idx, dict_pattern_info):
		str_name = class_param.name
		str_type = class_param.type

		if class_param.output_type != int_OUTPUT_TYPE_MEMBER:
			class_next_param = list_param[int_cur_idx + 1] if int_cur_idx < len(list_param) - 1 else None
			bool_is_pointer_array = False
			if class_next_param != None:
				str_check_name1 = testcode_remove_pointer(testcode_remove_array(str_name))
				str_check_name2 = testcode_remove_pointer(testcode_remove_array(class_next_param.name))
				bool_is_pointer_array = (class_next_param.output_type == int_OUTPUT_TYPE_MEMBER) and class_next_param.is_array_format and (str_check_name1 == str_check_name2)

			if bool_is_pointer_array:
				str_name = "*" + class_next_param.name

			# Append param name
			if class_param.output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
				str_params_name      = testcode_remove_pointer(testcode_remove_array(str_name))
				str_local_param_name = testcode_remove_pointer(testcode_remove_array(class_param.local_param_name))
				tmp_idx = re.search(r'_{}$'.format(str_params_name), str_local_param_name).span()[0]
				local_func_name = class_param.local_param_name[:tmp_idx]
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_NAME].append(local_func_name)
			else:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_NAME].append(str_name)

			# Append param type
			dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_TYPE].append(str_type)

			# Get "not change" data type
			bool_notchange_data_type = False
			for int_row in self.worksheet.iter_rows(min_col = self.worksheet.min_column, max_col = self.int_max_column, values_only=True):
				if int_row[class_param.col - 1] == g_NOT_CHANGE_VALUE:
					bool_notchange_data_type = True
					break

			# Get output data type
			if bool_notchange_data_type:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_OUT_NOT_CHANGE)
			elif class_param.is_pointer_format:
				if bool_is_pointer_array:
					dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_OUT_POINTER_ARRAY)
				elif class_param.is_struct_format:
					dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_OUT_POINTER_VALUE)
				else:
					dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_OUT_POINTER)
			else:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_OUT_VALUE)

			# Get item type
			if class_param.output_type == int_OUTPUT_TYPE_LOCAL_FUNC:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_LOCAL_FUNC)
			elif class_param.output_type == int_OUTPUT_TYPE_PARAMETER_OUT:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_PARAMETER)
			elif class_param.output_type == int_OUTPUT_TYPE_RET:
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_RETURN_VALUE)
			else: # Related global out
				dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_GLOBAL_PARAMETER)

		# Get expected output variable name for output parameters and related global parameters and return value
		if class_param.output_type != int_OUTPUT_TYPE_LOCAL_FUNC:
			str_name = self.__test_data_update_expected_variable_name(str_name, True)

		# set expected parameter name to pattern info
		dict_pattern_info = self.__testdata_create_list_of_dict(class_param, list_param, int_cur_idx, dict_pattern_info, uts_TestCode.g_str_PATTERN_EXPECTED_PARAM_NAME, str_name)

		return dict_pattern_info

	## This method mainly for create dictionary information for TestCode
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @return dict_pattern_info			Dictionary of data providing to test code
	def testdata_create_testcode_info(self, list_param_info):
		dict_pattern_info = {
						 	uts_TestCode.g_str_PATTERN_INPUT_PARAM_TYPE    :[], \
						 	uts_TestCode.g_str_PATTERN_INPUT_PARAM_NAME    :[], \
							uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE     :[], \
							uts_TestCode.g_str_PATTERN_INPUT_ITEM_TYPE     :[], \
							uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_TYPE   :[], \
							uts_TestCode.g_str_PATTERN_OUTPUT_PARAM_NAME   :[], \
							uts_TestCode.g_str_PATTERN_EXPECTED_PARAM_NAME :[], \
							uts_TestCode.g_str_PATTERN_OUTPUT_DATA_TYPE    :[], \
							uts_TestCode.g_str_PATTERN_OUTPUT_ITEM_TYPE    :[], \
							uts_TestCode.g_str_PATTERN_OUTPUT_POINTER_NAME :[]
						}

		dict_pattern_info[uts_TestCode.g_str_PATTERN_OUTPUT_POINTER_NAME] = g_OUTPUT_POINTER_NAMES
		int_num_member = int_INIT_VALUE

		for int_num in range(len(list_param_info)):
			class_info = list_param_info[int_num]
			
			# Decrease number of loop for struct member of array struct
			if int_num_member != int_INIT_VALUE:
				int_num_member -= 1

			if int_num < len(list_param_info) - 1:
				next_class_info = list_param_info[int_num + 1]
				# get name of next param
				str_next_param_name = testcode_remove_pointer(testcode_remove_array(next_class_info.name))
			else:
				str_next_param_name = ""
				next_class_info = None

			# Skip only array member
			if (int_num != 0) and (list_param_info[int_num].name == list_param_info[int_num-1].name):
				if list_param_info[int_num].local_param_name == list_param_info[int_num-1].local_param_name:
					continue
			
			# Initialize int_num_member for array struct
			if (list_param_info[int_num].func_param_type == int_TYPE_ARRAYSTRUCT or list_param_info[int_num].func_param_type == int_TYPE_ARR_POINTER_STRUCT) and (list_param_info[int_num].io_type == int_PARAM_TYPE_EXPECT):
				if list_param_info[int_num].number_of_struct_member == 0:
					continue
				elif int_num_member == int_INIT_VALUE:
					int_num_member = list_param_info[int_num].number_of_struct_member
			
			# Skip for pointer to value or pointer to array	
			if (int_num != 0) and (list_param_info[int_num-1].func_param_type == int_TYPE_STRUCTPOINTER) and \
				(list_param_info[int_num].name in list_param_info[int_num-1].name):
				continue
			
			if ((class_info.input_type not in [0, int_INPUT_TYPE_MEMBER]) or \
				(class_info.output_type != 0) or (class_info.item_type == int_ITEM_TYPE_LOCAL)):

				str_params_name = class_info.name
				str_params_type = class_info.type.replace("const ", "")
				str_raw_param_name = testcode_remove_array(testcode_remove_pointer(str_params_name))

				# Parameter in
				if class_info.input_type == int_INPUT_TYPE_PARAM:
					# handle member to set input parameter name for test code
					if class_info.is_pointer_format and next_class_info != None:
						# get next parameter info
						str_pointer_name = testcode_remove_pointer(testcode_remove_array(class_info.name))

						# handle input param is an array or a pointer of array types
						if class_info.name.endswith("]"):
							if (str_pointer_name == str_next_param_name) and class_info.name.startswith("*"):
								# pattern format: *param[], *param[3] ONLY, and PCL defined data for these pointer
								str_params_name = "*" + next_class_info.name
							else:
								# pattern format: *param[], *param[3]
								# pattern format: parram[], param[3], and PCL has not defined data for these pointer
								# Change array into pointer
								obj_match = re.search(r'\w+(((->|\.)\w+)*)((\[\w*\])+)$', class_info.name)
								if obj_match != None:
									for obj_ele in re.findall(r'\[\d*\]', obj_match.group()):
										str_params_name = "*{}".format(str_params_name.replace(obj_ele, ""))

					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_PARAM_NAME].append(str_params_name)
					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_PARAM_TYPE].append(str_params_type)

					if class_info.func_param_type == int_TYPE_POINTER and str_next_param_name != str_raw_param_name:
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_IN_GLOBAL)
						self.__testdata_obtain_global_variable(str_params_type, str_params_name, True)
					else:
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_IN_LOCAL)

					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_PARAMETER)

				# Related global in
				elif class_info.input_type == int_INPUT_TYPE_RELATED_GLOBAL:
					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_PARAM_NAME].append(str_params_name)
					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_PARAM_TYPE].append(str_params_type)

					if class_info.func_param_type == int_TYPE_POINTER and str_next_param_name != str_raw_param_name:
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_IN_GLOBAL)
						self.__testdata_obtain_global_variable(str_params_type, str_params_name, True)
					else:
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_IN_LOCAL)

					dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_GLOBAL_PARAMETER)

				# Local function in and value
				elif class_info.input_type == int_INPUT_TYPE_LOCAL_RET or class_info.item_type == int_ITEM_TYPE_LOCAL:
					# only append param type, item type and data type for local function name, not local func variables
					if class_info.item_type != int_ITEM_TYPE_LOCAL:
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_PARAM_TYPE].append(str_params_type)
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_DATA_TYPE].append(uts_TestCode.g_str_PATTERN_DATA_IN_LOCAL)
						dict_pattern_info[uts_TestCode.g_str_PATTERN_INPUT_ITEM_TYPE].append(uts_TestCode.g_str_PATTERN_ITEM_LOCAL_FUNC)

					str_params_name = re.sub(r'\*|\[\d+\]', "", str_params_name)
					dict_pattern_info = self.__testdata_create_list_of_dict(class_info, list_param_info, int_num, dict_pattern_info, \
						uts_TestCode.g_str_PATTERN_INPUT_PARAM_NAME, str_params_name)

				# Return value, Parameter out, Related global out, Local function out, member out
				else:
					dict_pattern_info = self.__testdata_create_info_out(class_info, list_param_info, int_num, dict_pattern_info)

		return dict_pattern_info

	## This method for casting type
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param class_previous_param_info		Class of previous ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param int_row						Current test case's row
	#  @param int_index						Current index if ParamInfo in list_param_info
	#  @return str_newCellValue				String value to be written into file
	#  @return int_offset_range				Range of columns to be skipped
	#  @return bool_null					Boolean for indicate NULL member is a structure pointer
	def __testdata_casting_type(self, list_param_info, class_previous_param_info, class_info, int_row, int_index):
		str_newCellValue = ""
		bool_null = False
		int_offset_range = 0
		for _ in range(class_previous_param_info.openBracket):
			str_newCellValue += "{ "
			
		# Take the previous value to indicate whether casting type for current ParamInfo or not
		str_previous_value = testdata_get_cell_value(self.worksheet, int_row, class_previous_param_info.col, int_MODE_STRIP)

		# If the value is "not null" or don't care => Cast type for member struct pointer
		if (str_previous_value == g_NOT_NULL_VALUE or str_previous_value == '-'):
			# Check name of previous and current ParamInfo
			str_name1 = testcode_remove_pointer(testcode_remove_array(class_previous_param_info.name))
			str_name2 = testcode_remove_pointer(testcode_remove_array(class_info.name))
			# Initialize
			str_previous_type = ""
			str_to_cast_previous = ""
			str_arr_length = ""

			if str_name1 == str_name2:
				# Format: abc[2], abc[2][3]
				obj_match = re.search(r'(\[\d+\])+$', class_info.name)
				if obj_match != None:
					str_arr_length = obj_match.group()
				# Search to add "*" in previous type 
				int_count_pointer_pattern = class_previous_param_info.name.count("*")
				for _ in range (int_count_pointer_pattern):
					str_to_cast_previous += "*"
				# Search to add array pattern in previous type
				if re.search(r'(\[\d+\])+$', class_previous_param_info.name):
					str_to_cast_previous += re.search(r'(\[\d+\])+$', class_previous_param_info.name).group()

				str_temp_type = "{} {}".format(class_previous_param_info.type, str_to_cast_previous)
				str_previous_type = "({})".format(str_temp_type) if len(str_temp_type) > 0 else ""
				str_newCellValue += "{}&({}{})".format(str_previous_type, class_info.type, str_arr_length)
			# Different name
			else:
				str_newCellValue += "&({})".format(class_previous_param_info.type)
			
		# If the value is "null", generate only NULL value
		elif str_previous_value == g_NULL_VALUE:
			bool_null = True
			str_newCellValue += g_NULL_VALUE.upper()
			# Loop to skip struct pointer member
			int_idx = 1
			int_sum_openBracket = 0
			int_sum_closeBracket = 0
			while(True):
				class_next_param = list_param_info[int_index + int_idx]
				if class_next_param.type_row > class_previous_param_info.type_row:
					# Accumulate brackets
					if not class_next_param.is_struct_format:
						int_sum_openBracket  += class_next_param.openBracket
						int_sum_closeBracket += class_next_param.closeBracket

					int_offset_range += 1	
					int_idx += 1
				else:
					# Open bracket
					int_openBracket  = int_sum_openBracket  + class_info.openBracket
					int_closeBracket = int_sum_closeBracket + class_info.closeBracket

					offset = int_closeBracket - int_openBracket
					for _ in range(offset):
						str_newCellValue += " }"

					break

		return str_newCellValue, int_offset_range, bool_null
	
	## This method for common process for str_cellValue
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param int_index						Current index if ParamInfo in list_param_info
	#  @param int_row						Current test case's row
	#  @return str_newCellValue				String value to be written into file
	#  @return int_offset_range				Range of columns to be skipped
	def __testdata_common_Write_action(self, list_param_info, class_info, str_value, str_cellValue, int_index, int_row):
		str_newCellValue = ''
		bool_range_array = False
		bool_null_value = False
		int_offset_range = 0

		if len(list_param_info) > 0:
			class_previous_param_info = list_param_info[int_index - 1]
		else:
			class_previous_param_info = None

		# Check if it is pointer to array struct that is skipped in the write_test_pattern
		if (class_previous_param_info != None) and (class_previous_param_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT]) and \
			(class_previous_param_info.input_type == int_INPUT_TYPE_MEMBER or class_previous_param_info.output_type == int_OUTPUT_TYPE_MEMBER):
			class_previous_previous_param_info = list_param_info[int_index - 2]
			if class_previous_previous_param_info.func_param_type == int_TYPE_STRUCTPOINTER and \
					(class_previous_previous_param_info.input_type == int_INPUT_TYPE_MEMBER or class_previous_previous_param_info.output_type == int_OUTPUT_TYPE_MEMBER) :
				str_newCellValue, int_offset_range, bool_null_value = self.__testdata_casting_type(list_param_info, class_previous_previous_param_info, class_previous_param_info, int_row, int_index - 1)
				if bool_null_value:
					return str_newCellValue, int_offset_range - 1

		# Check to cast type for struct pointer(not function param)
		elif class_previous_param_info != None and \
			(class_previous_param_info.func_param_type in [int_TYPE_STRUCTPOINTER, int_TYPE_ARR_POINTER_STRUCT]) and (not class_previous_param_info.is_func_param_format) and \
			(class_previous_param_info.input_type == int_INPUT_TYPE_MEMBER or class_previous_param_info.output_type == int_OUTPUT_TYPE_MEMBER):
			# Casting type	
			str_newCellValue, int_offset_range, bool_null_value = self.__testdata_casting_type(list_param_info, class_previous_param_info, class_info, int_row, int_index)
			if bool_null_value:
				return str_newCellValue, int_offset_range
		
		# Loop to add "{" 
		# The condition prevents adding bracket for designated element of array struct
		if isinstance(class_info.list_large_index, range):
			bool_range_array = True

		int_openBracket  = class_info.openBracket
		int_closeBracket = class_info.closeBracket
		if (str_value != g_NOT_STUB_VALUE) and self.__testdata_is_array_data_one_cell(class_info, str_value, True):
			int_openBracket  -= 1
			if class_info.input_type != int_INPUT_TYPE_LOCAL_RET:
				int_closeBracket -= 1

		if (not class_info.flag_large_index) or (class_info.flag_large_index and class_info.func_param_type in [int_TYPE_ARR_POINTER_STRUCT, int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER]) or \
			(class_info.flag_large_index and bool_range_array and class_info.list_large_index.start == 0):
			for _ in range(int_openBracket):
				str_newCellValue += "{ "

		# Replace "-" -> 0
		if str_cellValue != None:
			if str_value == g_NOT_STUB_VALUE:
				str_newCellValue += str_cellValue.replace(g_NOT_STUB_VALUE, "0")
			else:
				str_newCellValue += self.__testdata_convert_dontcare_value(str_cellValue)

		# Loop to add "}"
		for _ in range(int_closeBracket):
			str_newCellValue += " }"

		return str_newCellValue, int_offset_range

	# This method for getting increase value for validator index
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param int_idx						Current index of parent's ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param number_of_struct_member		Total amount of member of structure
	#  @param list_validator				List of validator
	#  @param int_validator_index			Current index of validator
	#  @return int_increase_value			Increase value for validator index
	def __testdata_get_validator_increase_value(self, list_param_info, int_idx, class_info, number_of_struct_member, list_validator, int_validator_index):
		
		if isinstance(class_info.list_large_index, int):
			int_start_idx = class_info.list_large_index
		elif isinstance(class_info.list_large_index, range):
			int_start_idx = class_info.list_large_index.start
		else:
			raise Exception("[ERROR] Invalid list_large_index")

		bool_is_not_start_idx_arr  = isinstance(class_info.list_large_index, int)   and (class_info.list_large_index != 0)
		bool_is_not_start_idx_arr |= isinstance(class_info.list_large_index, range) and (class_info.list_large_index.start != 0)
		bool_is_not_start_idx_arr &= (number_of_struct_member > 0)

		class_prev_info = list_param_info[int_idx - 1] if int_idx > 0 else None
		str_name1 = testcode_remove_array(testcode_remove_pointer(class_info.name))
		str_name2 = testcode_remove_array(testcode_remove_pointer(class_prev_info.name))
		bool_is_not_start_idx_arr &= (str_name1 != str_name2)

		if bool_is_not_start_idx_arr:
			class_prev_info = list_param_info[int_idx - number_of_struct_member] if int_idx > number_of_struct_member else None
			str_name1 = testcode_remove_array(testcode_remove_pointer(class_info.name))
			str_name2 = testcode_remove_array(testcode_remove_pointer(class_prev_info.name))

		if (str_name1 == str_name2) and isinstance(class_prev_info.list_large_index, int):
			int_increase_value = int_start_idx - class_prev_info.list_large_index
		elif (str_name1 == str_name2) and isinstance(class_prev_info.list_large_index, range):
			int_increase_value = int_start_idx - class_prev_info.list_large_index[-1]
		else:
			int_increase_value = 1

		if bool_is_not_start_idx_arr and (int_increase_value > 1):
			obj_match = re.search(r'\d+ ]', list_validator[-number_of_struct_member])
			int_prev_validator_idx = int(obj_match.group().replace("]","").strip()) if obj_match != None else int_increase_value - 1
			int_increase_value = int_validator_index - int_prev_validator_idx + 1

		return int_increase_value

	# This method for getting validator value of output value
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param list_validator				List of validator
	#  @param int_validator_index			Current index of validator
	#  @param int_count_validator_name		Validator name
	#  @param class_info					Class of ParamInfo
	#  @param int_validator_index			Current index of validator
	#  @param int_count_validator_name		Validator name
	#  @param int_row						Current row
	#  @param int_idx						Current index of parent's ParamInfo
	#  @param number_of_struct_member		Total amount of member of structure
	#  @return int_validator_index			Current index of validator
	#  @return int_count_validator_name		Validator name
	#  @return list_validator				List of validator
	def __testdata_get_validator(self, list_param_info, list_validator, int_validator_index, int_count_validator_name, class_info, int_row, int_idx, number_of_struct_member):
		int_range = 0
		# Not write validator for struct type or member is a struct/structpointer 
		if not isinstance(class_info.list_large_index, range) and (class_info.func_param_type == int_TYPE_STRUCT or \
				((class_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER]) and (class_info.output_type == int_OUTPUT_TYPE_MEMBER))):
			return list_validator, int_validator_index, int_count_validator_name, int_range

		# Add validator 
		str_bool = self.__testdata_get_boolean_validator(class_info, int_row)
		if isinstance(class_info.list_large_index, range):
			if class_info.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARR_POINTER_STRUCT, int_TYPE_ARRAYPOINTER]:	
				range_array = class_info.list_large_index
				int_validator_index += self.__testdata_get_validator_increase_value(list_param_info, int_idx, class_info, number_of_struct_member, list_validator, int_validator_index)
				str_temp_validator = "[ {} ... {} ] = {}".format(int_validator_index - 1, (range_array.stop - 1) + (int_validator_index - 1) - range_array.start, str_bool)
				list_validator.append(str_temp_validator)
				int_validator_index = range_array.stop + (int_validator_index - 1) - range_array.start
				int_count_validator_name += 1
			# Array struct
			else:
				int_range_to_loop = class_info.list_large_index.stop - class_info.list_large_index.start 
				for j in range (int_range_to_loop):
					for i in range (number_of_struct_member):
						class_next_info = list_param_info[int_idx + i]
						range_array = class_next_info.list_large_index

						# Obtain the validate value
						str_bool = self.__testdata_get_boolean_validator(class_next_info, int_row)
						if isinstance(class_next_info.list_large_index, range):
							int_increase_value = self.__testdata_get_validator_increase_value(list_param_info, int_idx + i, class_next_info, number_of_struct_member, list_validator, int_validator_index)

							if class_next_info.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARR_POINTER_STRUCT, int_TYPE_ARRAYPOINTER]:	
								int_validator_index += int_increase_value
								str_temp_validator = "[ {} ... {} ] = {}".format(int_validator_index - 1, (range_array.stop - 1) + (int_validator_index - 1) - range_array.start, str_bool)
								list_validator.append(str_temp_validator)
								int_validator_index = range_array.stop + (int_validator_index - 1) - range_array.start
								int_count_validator_name += 1
							elif class_next_info.func_param_type == int_TYPE_STRUCT or ((class_next_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER]) and (class_next_info.output_type == int_OUTPUT_TYPE_MEMBER)):
								pass
							else:
								if j == 0:
									int_validator_index += int_increase_value
								else:
									int_validator_index += 1

								str_temp_validator = "[ {} ] = {}".format(int_validator_index - 1, str_bool)
								list_validator.append(str_temp_validator)
								int_count_validator_name += 1	
						# Not write validator for struct type or member is a struct/structpointer 
						elif class_next_info.func_param_type == int_TYPE_STRUCT or ((class_next_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER]) and (class_next_info.output_type == int_OUTPUT_TYPE_MEMBER)):
							pass
						else:
							list_validator.append(str_bool)
							int_validator_index += 1
							int_count_validator_name += 1
				int_range = number_of_struct_member - 1
		else:
			if class_info.list_large_index != list([]):
				int_validator_index += self.__testdata_get_validator_increase_value(list_param_info, int_idx, class_info, number_of_struct_member, list_validator, int_validator_index)
				str_temp_validator = "[ {} ] = {}".format(int_validator_index - 1, str_bool)
			else:
				int_validator_index += 1
				str_temp_validator = str_bool

			int_count_validator_name += 1
			list_validator.append(str_temp_validator)
			
		return list_validator, int_validator_index, int_count_validator_name, int_range

	# This method for getting boolean validator value of output value
	#  @param self							The object pointer
	#  @param class_info					Class of ParamInfo
	#  @param int_row						Current row in PCL						
	#  @return str_bool						Boolean for validator
	def __testdata_get_boolean_validator(self, class_info, int_row):
		# Obtain the validate value
		str_value = testdata_get_cell_value(self.worksheet, int_row, class_info.col, int_MODE_REPLACE)
		if re.search(r"^{-}$|^-$", str_value):
			str_bool = "FALSE"
		else:
			str_bool = "TRUE"
		return str_bool

	# This method for getting validator list
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param int_row						Current row in PCL
	#  @return list_validator				List of validator
	#  @return int_count_validator_name		Validator name
	def __testdata_get_validator_list(self, list_param_info, int_row):
		list_validator = []
		int_validator_index = 0 
		int_count_validator_name = 0
		int_index = 0
		int_range_to_skip = 0
		while int_index < (len(list_param_info)):
			class_info = list_param_info[int_index]
			int_range = 0
			int_range_member = 0
			if class_info.io_type == int_PARAM_TYPE_EXPECT and class_info.output_type != int_OUTPUT_TYPE_RET:
				if class_info.func_param_type == int_TYPE_ARRAYSTRUCT:
					int_temp_idx = 1
					int_temp_next_col = class_info.col + 1
					# Finding range member of array struct
					while(True):
						int_current_type_row = testdata_finding_type_row(self.worksheet, int_temp_next_col, self.int_max_column, g_INT_PARAM_TYPE_ROW, g_INT_PARAM_NAME_ROW)[0]
						int_offset = int_current_type_row - class_info.type_row
						# Next column is n-level lower
						if int_offset > 0:
							int_range_member += 1
						else:
							break
						
						int_temp_next_col += 1

					while int_temp_idx <= class_info.array_struct_member_range:
						class_info_member = list_param_info[int_index + int_temp_idx]
						list_validator, int_validator_index, int_count_validator_name, int_range_to_skip = self.__testdata_get_validator(list_param_info, 
							list_validator, int_validator_index, int_count_validator_name, class_info_member, int_row, int_index + int_temp_idx, int_range_member)
						int_temp_idx += 1 + int_range_to_skip		
					int_range += class_info.array_struct_member_range
				# Not array struct
				elif class_info.func_param_type == int_TYPE_STRUCT or ((class_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_STRUCTPOINTER]) and (class_info.output_type == int_OUTPUT_TYPE_MEMBER)):
					pass
				
				else:
					# Obtain the validate value
					list_validator, int_validator_index, int_count_validator_name, int_range_to_skip = self.__testdata_get_validator(list_param_info, 
						list_validator, int_validator_index, int_count_validator_name, class_info, int_row, int_index, int_range_member)
			
			int_index += int_range + 1
					
		return list_validator, int_validator_index, int_count_validator_name		

	## This method for common processing value for str_cellValue
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param list_temporary				List to store multiple data #TODO: Remove this list
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @return str_cellValue				String value to be written into file
	#  @return list_temporary				List to store multiple data #TODO: Remove this list
	def __testdata_handle_cell_value(self, list_param_info, class_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index):
		bool_is_contain_large_index = class_info.list_large_index != list([])
		bool_is_define_one_cell = self.__testdata_is_array_data_one_cell(class_info, str_value, True)

		int_openBracket  = class_info.openBracket
		int_closeBracket = class_info.closeBracket
		if (class_info.item_type == int_ITEM_TYPE_LOCAL) and bool_is_define_one_cell:
			int_openBracket  -= 1
			int_closeBracket -= 1

		# Add array index (if any)
		str_large_index = ""
		if bool_is_contain_large_index:
			# Get index, If range_array is a specific index For example ([0])
			if isinstance(class_info.list_large_index, int):
				if class_info.list_large_index == 0:
					str_large_index = ""
				else:
					str_large_index = "[{}] = ".format(class_info.list_large_index)
			# If range_array is a range object For example (0 ... n)
			else:
				str_large_index = "[ {} ... {} ] = ".format(list(class_info.list_large_index)[0], list(class_info.list_large_index)[-1])

			if bool_is_define_one_cell and str_value != g_NOT_STUB_VALUE:
				str_large_index = ""

		# Add string large index
		if ((bool_is_define_one_cell or not class_info.is_array_format) and str_value != g_NOT_STUB_VALUE):
			str_cellValue += str_large_index

		# Add open braces in case input of local function
		if ((not isinstance(class_info.list_large_index, int)) and len(class_info.list_large_index) > 0 and list(class_info.list_large_index)[0] != 0) or \
				(isinstance(class_info.list_large_index, int) and class_info.list_large_index > 0) or \
				(isinstance(class_info.list_large_index, int) and class_info.list_large_index == 0 and class_info.item_type == int_ITEM_TYPE_LOCAL) or \
				(not bool_is_define_one_cell and class_info.item_type == int_ITEM_TYPE_LOCAL):
			for _ in range (int_openBracket):
				str_cellValue += "{ "

		if not ((bool_is_define_one_cell or not class_info.is_array_format) and str_value != g_NOT_STUB_VALUE):
			str_cellValue += str_large_index

		# Process cell's value based on their data type
		if class_info.func_param_type in [int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER, int_TYPE_ARR_POINTER_STRUCT]:
			if not bool_is_define_one_cell:
				str_value = re.sub(r'^{|}$', "", str_value)

			if class_info.is_pointer_format:
				str_cellValue = self.__testdata_handle_pointer_value(class_info, str_value, str_cellValue, int_min_row, int_max_row)
			else:
				str_cellValue += str_value

		elif class_info.func_param_type in [int_TYPE_STRUCT, int_TYPE_ARRAYSTRUCT]:
			pass

		elif class_info.func_param_type in [int_TYPE_STRUCTPOINTER, int_TYPE_POINTER]:
			str_cellValue = self.__testdata_handle_pointer_value(class_info, str_value, str_cellValue, int_min_row, int_max_row)

		elif class_info.func_param_type == int_TYPE_RETVAL:
			# If cell's value is "not stub", generate 0	
			if str_cellValue == g_NOT_STUB_VALUE: # TODO This checking shall only done in "Local Function" type
				str_cellValue += '0'
			else:
				# If that column has defined "not change" value, generate a struct of zero instead
				if self.__testdata_scan_for_finding_pattern(self.worksheet, class_info, int_min_row, int_max_row, g_NOT_CHANGE_VALUE):
					str_cellValue += '{ 0 }'
				else:
					str_cellValue += str_value
		else:
			raise Exception("[E_UTS_PCL] Wrong type for input function param {}!!!".format(class_info.func_param_type))

		# Add close braces in case input of local function
		if (not bool_is_define_one_cell and class_info.item_type == int_ITEM_TYPE_LOCAL):
			for _ in range (int_closeBracket):
				str_cellValue += " }"

		return str_cellValue, list_temporary

	## This method for common processing value for str_cellValue for Designated Initialize format
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param list_temporary				List to store multiple data #TODO: Remove this list
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @return str_cellValue				String value to be written into file
	#  @return list_temporary				List to store multiple data #TODO: Remove this list
	def __testdata_handle_large_array(self, list_param_info, class_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index):
		bool_is_return = False

		# Process for array value
		str_value, list_temporary = self.__testdata_handle_cell_value(list_param_info, class_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index)

		# Not write value for struct
		if str_value == '':
			bool_is_return = True

		if not bool_is_return:
			if class_info.func_param_type == int_TYPE_ARRAY and class_info.type != '':
				str_cellValue = str_value
				bool_is_return = True

		if not bool_is_return:			
			if class_info.func_param_type in [int_TYPE_ARR_POINTER_STRUCT, int_TYPE_ARRAY, int_TYPE_ARRAYPOINTER]:
				str_cellValue += str_value
			# Array struct
			else:
				str_cellValue = ""
				pass

		return str_cellValue, list_temporary
	
	## This method for common processing value for str_cellValue pointer type
	#  @param self							The object pointer
	#  @param class_info					Class of ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @return str_cellValue				String value to be written into file
	def __testdata_handle_pointer_value(self, class_info, str_value, str_cellValue, int_min_row, int_max_row):
		# Pattern match for: &aaa, -, {-}, (aa)bb
		obj_searching_pattern = re.search(r'&|-|(\{\-|\d+)\}', str_value)
		
		# If ParamInfo is a member and its value is not "not null" => Generate value
		if (class_info.input_type == int_INPUT_TYPE_MEMBER or class_info.output_type == int_OUTPUT_TYPE_MEMBER) and str_value != g_NOT_NULL_VALUE:
			# If cell's value is "null" => Generate "NULL"
			if str_value == g_NULL_VALUE:
				if class_info.item_type == int_ITEM_TYPE_LOCAL:
					str_cellValue += g_NULL_VALUE.upper()
				else:
					str_cellValue = g_NULL_VALUE.upper()
			else:
				str_cellValue += str_value
		
		# If ParamInfo is not a member and not belongs to "not stub, -" value => Generate macro flag
		elif class_info.input_type != int_INPUT_TYPE_MEMBER and class_info.output_type != int_OUTPUT_TYPE_MEMBER and \
				str_value != g_NOT_STUB_VALUE and str_value != "-":
			str_cellValue += testcode_convert_string_to_macro(str_value)
			# Output pointer will be append to global list for reporting to console/terminal
			if (class_info.io_type == int_PARAM_TYPE_EXPECT) and (str_value not in g_OUTPUT_POINTER_NAMES) and \
				(class_info.func_param_type in [int_TYPE_POINTER, int_TYPE_ARRAYPOINTER, int_TYPE_ARR_POINTER_STRUCT, int_TYPE_STRUCTPOINTER]):
				g_OUTPUT_POINTER_NAMES.append(str_value)
		
		# If cell's value is "not change" and a pointer type => Generate struct of 0
		elif str_value == g_NOT_CHANGE_VALUE and class_info.func_param_type == int_TYPE_POINTER:
			str_cellValue = '{ 0 }'

		# If cell's value is "not null" and a member
		elif str_value == g_NOT_NULL_VALUE and (class_info.input_type == int_INPUT_TYPE_MEMBER or class_info.output_type == int_OUTPUT_TYPE_MEMBER):
			# Create global variable
			str_cellValue += "&{}".format(uts_TestCode.g_str_ENV_GLOBAL_VAR_PREFIX)
			str_cellValue += testcode_remove_pointer(testcode_remove_array(class_info.name))
		
		# Matched Pattern
		elif obj_searching_pattern:
			if self.__testdata_scan_for_finding_pattern(self.worksheet, class_info, int_min_row, int_max_row, g_NOT_CHANGE_VALUE):
				str_cellValue += '{ 0 }'
			else:
				if str_cellValue == g_NOT_STUB_VALUE:
					str_cellValue += '0'
				else:
					str_cellValue += str_value
		else:
			str_cellValue += str_value
			
		return str_cellValue

	## This method for common processing value for str_cellValue for all type of value
	#  @param self							The object pointer
	#  @param list_param_info				List of ParamInfo
	#  @param class_info					Class of ParamInfo
	#  @param class_previous_info			Class of previous ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param list_temporary				List to store multiple data #TODO: Remove this list
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @return str_cellValue				String value to be written into file
	#  @return list_temporary				List to store multiple data #TODO: Remove this list
	def __testdata_process_cell_value(self, list_param_info, class_info, class_previous_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index):
		# Clear list stored array defined in one cell
		if (class_previous_info == None) or (class_previous_info.name != class_info.name):
			list_temporary.clear()

		if len(str_value) > 0:
			# Designated Initializer for array type
			if class_info.is_array_format:
				str_cellValue, list_temporary = self.__testdata_handle_large_array(list_param_info, class_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index)
				
			else:
				# Struct
				if class_info.func_param_type == int_TYPE_STRUCT:
					pass
				# Pointer/Pointer Struct/Value
				else:
					str_cellValue, list_temporary = self.__testdata_handle_cell_value(list_param_info, class_info, str_value, str_cellValue, list_temporary, int_min_row, int_max_row, int_current_index)

		return str_cellValue, list_temporary

	## This method for writing data to target file
	#  @param self							The object pointer
	#  @param file_to_write					Target file to be written
	#  @param list_alignment_length			List of alignment length for each element
	#  @param list_to_write					List of data to be written
	#  @param int_idx_row					Current row of data in PCL
	#  @param length_array_validator		Validator's array length
	#  @param bool_is_input_flag			Boolean to indicate whether is writing for input struct
	#  @return None
	def __testdata_write_list_to_file(self, file_to_write, list_alignment_length, list_to_write, int_idx_row, length_array_validator, bool_is_input_flag):
		# Flag to determine start index of list Input type
		if bool_is_input_flag:
			int_temp_index = 1
		# Flag to determine start index of list Expect type
		else:
			int_temp_index = 0
		# TODO: Remove when release
		list_length_2d_list = set([len(list_element) for list_element in list_to_write])
		if len(list_length_2d_list) > 1:
			raise Exception("[Error] Error at adding space for alignment")
		
		for int_idx_col in range (int_temp_index, len(list_alignment_length)):
			int_tmp_range = list_alignment_length[int_idx_col] + 1 - len(list_to_write[int_idx_row][int_idx_col]) 	# Range of maximum character length in a column
			# Add space to make alignment
			while int_tmp_range > 0:
				list_to_write[int_idx_row][int_idx_col] += " "
				int_tmp_range -= 1
			str_value = list_to_write[int_idx_row][int_idx_col]
			obj_match = re.search(r'^ +$', str_value)
			# If end of row, don't add comma
			# Write input type
			if bool_is_input_flag:
				if int_idx_col == len(list_alignment_length) - 1: 
					file_to_write.write(str_value)
				elif obj_match:
					file_to_write.write(str_value)
					# Add 2 space for ", " 
					file_to_write.write("  ")
				else:
					int_idx = int_idx_col + 1
					bool_space = False
					while(int_idx < len(list_to_write[int_idx_row])):
						str_temp_value = list_to_write[int_idx_row][int_idx]
						if not re.search(r'^ +$', str_temp_value):
							break
						else:
							if int_idx != len(list_to_write[int_idx_row]) - 1:
								int_idx += 1	
							else:
								bool_space = True
								break
					if bool_space:
						file_to_write.write(str_value)
						file_to_write.write("  ")
					else:
						file_to_write.write(str_value + ', ')
			# Write expect type
			else:
				# Write last element (Not validator)
				if int_idx_col == (len(list_alignment_length) - length_array_validator - 1):
					# _FOUT.write("{ ")
					if int_idx_row == 0 or obj_match:
						file_to_write.write(str_value + ', ')
					else:
						file_to_write.write(str_value)
						# Write close bracket for expect struct
						self.__testdata_write_bracket_for_struct_input_expect(file_to_write, list_to_write, False, True)
						if length_array_validator > 0:
							file_to_write.write(", { ")
				# Write validator
				else:
					if int_idx_col == len(list_alignment_length) - 1:
						file_to_write.write(str_value)
						self.__testdata_write_bracket_for_struct_input_expect(file_to_write, list_to_write, False, True)
					elif obj_match:
						file_to_write.write(str_value)
						# Add 2 space for ", " 
						file_to_write.write("  ")
					else:
						file_to_write.write(str_value + ', ')

	## This method for writing data to target file
	#  @param self							The object pointer
	#  @param str_pattern					String of pattern to be looked up in PCL
	#  @param int_ofs_row					Offset of row(not used rows in PCL)
	#  @param int_ofs_col					Offset of column(not used columns in PCL)
	#  @return class_cell					Class of location of found str_pattern (class Position)
	def testdata_get_cell(self, str_pattern, int_ofs_row, int_ofs_col):
		worksheet = self.worksheet
		# Loop whole worksheet row by row
		for int_row in worksheet.iter_rows(max_row=worksheet.max_row, min_col=g_SKIPED_COLUMNN+1, max_col=self.int_max_column):
			# Loop through each column in row
			for int_idx in range(self.int_max_column - g_SKIPED_COLUMNN):
				# Condition to find string that matched with the str_pattern
				if (int_row[int_idx].value != None ) and (str(int_row[int_idx].value).strip() == str_pattern):
					# If found matched string, store the location of it by create class CellInfo
					class_cell = CellInfo(int_row[int_idx].row + int_ofs_row, int_row[int_idx].column + int_ofs_col, str_pattern)
					return class_cell
		# If not found str_pattern, raise Exception
		raise Exception("[E_UTS_PCL] Cell contain pattern \'{}\' not found".format(str_pattern))
	
	## This method for detecting invalid characters/space/brackets 
	#  @param self							The object pointer
	#  @return None
	def testdata_detect_invalid(self):
		worksheet = self.worksheet
		# Get start cell
		class_start_param_data = self.testdata_get_cell(str_INPUT_AND_OUTPUT_DATA, g_OFFSET_ROW, g_OFFSET_COL)
		int_min_col  = self.testdata_get_cell(str_DATA_SPECIFICATION, 0, 0).int_col
		# Get end cell
		int_end_col = self.int_max_column - 1
		
		# Get number of test
		list_test_case_id = testdata_getTestCaseID(self.worksheet, True)
		int_number_of_test_case = len(list_test_case_id) - 1
		
		# Init variables
		int_min_row = class_start_param_data.int_row
		int_max_row = int_min_row + int_number_of_test_case 
		for int_row in worksheet.iter_rows(min_row = int_min_row, max_row=int_max_row - 1, min_col=int_min_col+1, max_col=int_end_col):
			for i in range(0, int_end_col - int_min_col):
				# If the cell's value is empty, then str_value will be NoneType object
				# Else the cell's value will be casted into string type
				str_value = testdata_get_cell_value(worksheet, int_row[i].row, int_row[i].column)
			
				# Find space at the start of value
				if str_value == None:
					raise Exception("[E_UTS_PCL] Empty cell at row {} column {}".format(int_row[i].row, get_column_letter(int_row[i].column)))
				
				# If it is string, don't check the value
				if ((str_value.startswith('"') and not str_value.endswith('"')) or \
					(not str_value.startswith('"') and str_value.endswith('"'))):
					raise Exception("[E_UTS_PCL] Wrong define value for string at row {} column {}".format(int_row[i].row, get_column_letter(int_row[i].column)))
				elif ((str_value.startswith("'") and not str_value.endswith("'")) or \
					(not str_value.startswith("'") and str_value.endswith("'"))):
					raise Exception("[E_UTS_PCL] Wrong define value for string at row {} column {}".format(int_row[i].row, get_column_letter(int_row[i].column)))
				elif (str_value.startswith('"') and str_value.endswith('"')) or (str_value.startswith("'") and str_value.endswith("'")):
					continue
				else:
					pass

				# Check multiple line
				if re.search(r"\n", str_value):
					raise Exception("[E_UTS_PCL] Multiple lines per cell at row {} column {}".format(int_row[i].row, get_column_letter(int_row[i].column))) 
				
				# Check Leading/Trailing white spaces in cell's value
				if re.search('^ +| +$', str_value):
					print("[Warning] Leading/Trailing white spaces found at '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))
				
				# Latin character only and operators in C
				bool_invalid_character = False
				for idx in range(len(str_value)):
					if ord(str_value[idx]) > 127:
						bool_invalid_character = True
				
				str_temp = str_value
				list_found_quotes = re.findall(r'(\".*?\")', str_temp)
				for item in list_found_quotes:
					str_temp = str_temp.replace(item, '')
				str_patn = r"[?!`@#$\\:;=]"
				obj_search_for_invalid = re.search(str_patn, str_temp)
				if bool_invalid_character or obj_search_for_invalid:
					raise Exception("[E_UTS_PCL] Invalid value at row {} column {}".format(int_row[i].row, get_column_letter(int_row[i].column)))
				
				# Missing close bracket
				if re.search(r"[\{]", str_value):
					if not re.search(r"[\}]", str_value):
						print("[Warning] Unmatched/Missing close bracket '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))
				if re.search(r"[\(]", str_value):
					if not re.search(r"[\)]", str_value):
						print("[Warning] Unmatched/Missing close bracket '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))
				
				# Missing open bracket
				if re.search(r"[\}]", str_value):
					if not re.search(r"[\{]", str_value):
						print("[Warning] Unmatched/Missing open bracket '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))
				if re.search(r"[\)]", str_value):
					if not re.search(r"[\(]", str_value):
						print("[Warning] Unmatched/Missing open bracket '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))
				
				# Check the encapsulation of the cell's value if it has bracket
				if str_value.count("(") != str_value.count(")") or str_value.count("[") != str_value.count("]") or str_value.count("{") != str_value.count("}"):
					print("[Warning] Unmatched number of bracket '{}' at row {} column {}".format(str_value, int_row[i].row, get_column_letter(int_row[i].column)))

	## This method for finding target pattern in target column
	#  @param self							The object pointer
	#  @param worksheet						Class of worksheet read from PCL (built-in class openpyxl)
	#  @param class_info					Class of ParamInfo
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @param str_pattern					String of pattern to be looked up in PCL
	#  @return boolean						True if found, False if not found
	def __testdata_scan_for_finding_pattern(self, worksheet, class_info, int_min_row, int_max_row, str_pattern):
		# Loop each row 
		for int_temp_row in range (int_min_row, int_max_row + 1):
			str_temp_cell_value = testdata_get_cell_value(self.worksheet, int_temp_row, class_info.col, int_MODE_STRIP)
			if str_temp_cell_value != None:
				if re.search(str_pattern, str_temp_cell_value):
					return True
		return False

	## This method for update expected variable name for giving information to testcode
	#  @param self							The object pointer
	#  @param str_name						Parameter name
	#  @param bool_remove_array				Boolean to indicate whether adding array info or not
	#  @return str_param					String name after conversion
	def __test_data_update_expected_variable_name(self, str_name, bool_remove_array):
		# Convert for expected output value: bool_remove_array = False
		# var_name->element[n] or var_name[n].element --> var_name_element[n] or var_name_n_element
		# var_name[m][n] --> var_name_m_n
		#
		# Convert for expected output value: bool_remove_array = True
		# test->state[20]       ---> test_state
		# test->state[20].type  ---> test_state_20_type
		# test->state[20]->type ---> test_state_20_type
		# test[10][16]          ---> test_10_16
		# test[10][17].type     ---> test_10_17_type
		# test[10][17]->type    ---> test_10_17_type
		str_param = ""
		str_array = ""
		# check if param name in array format
		if str_name.endswith("]"):
			# find 2D array
			match = re.search(r'(.+)(\[\d+\])(\[\d+\])$', str_name)
			if match != None:
				str_param = str_name
			else:
				# get info of param and 1D array
				match = re.search(r'(.+)(\[\d+\])$', str_name)
				if match != None:
					str_param = match.group(1)
					str_array = match.group(2)
				else:
					raise Exception("[E_UTS_PCL] Invalid array format", str_name)
		else:
			str_param = str_name

		str_param = str_param.replace("->", "_").replace(".", "_").replace("&", "").replace("*", "").replace(" ", "")
		str_param = str_param.replace("(", "").replace(")", "").replace("[", "_").replace("]", "")

		# check condition to append array info
		if not bool_remove_array:
			str_param += str_array
		
		return str_param

	## This method for appending value of local function input to its parents
	#  @param self							The object pointer
	#  @param worksheet						Class of worksheet read from PCL (built-in class openpyxl) #TODO: Remove this list
	#  @param list_param_info				List of ParamInfo
	#  @param str_value						String cell's value read from PCL
	#  @param str_cellValue					String cell's value to be written
	#  @param list_temporary				List to store multiple data #TODO: Remove this list
	#  @param int_current_index				Current index of ParamInfo in list_param_info when calling this function
	#  @param int_row						Current row of test case
	#  @param int_min_row					Min row of test case data
	#  @param int_max_row					Max row of test case data
	#  @param str_macro						String macro is_bool of local function input struct
	#  @return str_cellValue				String value to be written into file
	#  @return list_temporary				List to store multiple data #TODO: Remove this list
	def __testdata_local_func_find_value(self, list_param_info, str_value, str_cellValue, list_temporary, int_current_index, int_row, int_min_row, int_max_row, str_macro):
		int_temp_index = 1
		
		# Infinite loop to find next local function input defined data 
		while (True):
			str_temp_cellValue = ''
			# Obtain next ParamInfo
			class_temp_param = list_param_info[int_current_index + int_temp_index]
			
			# If next ParamInfo is local function input's value
			if class_temp_param.item_type == int_ITEM_TYPE_LOCAL:

				# Pointer to value/array/struct format is unsupported
				if (class_temp_param.type_row == g_INT_PARAM_TYPE_ROW) and class_temp_param.is_pointer_format and not class_temp_param.name.startswith("["):
					raise Exception("[E_UTS_PCL] Not support for Pointer at column {}".format(get_column_letter(class_temp_param.col)))

				# Obtain cell's value
				str_temp_next_cell_value = testdata_get_cell_value(self.worksheet, int_row, class_temp_param.col, int_MODE_STRIP)
				# If the value is "False/True" => turn it into uppercase FALSE/TRUE
				if str_temp_next_cell_value == 'False' or str_temp_next_cell_value == 'True':
					str_temp_next_cell_value = str_temp_next_cell_value.upper()
				else:
					str_temp_next_cell_value = self.__testdata_convert_dontcare_value(str_temp_next_cell_value)

				str_temp_cellValue, list_temporary = self.__testdata_process_cell_value(list_param_info, class_temp_param, list_param_info[int_current_index], str_temp_next_cell_value, str_temp_cellValue, list_temporary, int_min_row, int_max_row, int_current_index)
				str_cellValue += str_temp_cellValue
				# Add comma for next value
				if not class_temp_param.is_struct_format:
					str_cellValue += ", "
				int_temp_index += 1
			else:
				# Regular expression to remove last comma 
				str_cellValue = re.sub(r'\s*,\s*$', " ",str_cellValue)
				break
		# After write enough value for local function input, close the struct and add value for variable is_stub
		str_cellValue += str_macro
		
		return str_cellValue, list_temporary

## This class get/write data from PCL function.h/testcase.h/macro.h
# @param self				The object pointer
# @param workbook			The object class: workbook
class WorkbookInfo:
	## The constructor
	def __init__(self, workbook):
		self.workbook = workbook

	## This method for writing data to target file
	#  @param self								The object pointer
	#  @param str_output_path					String of destination location where the file will be written
	#  @param str_name_of_file					String name of the file 
	#  @param str_data							Data will be written
	#  @return None
	def __testdata_write_to_file(self, str_output_path, str_name_of_file, str_data):
		# Location of file
		str_name_of_file = os.path.join(str_output_path, str_name_of_file)
		file_to_write = testdata_get_file(str_name_of_file, "w")
		file_to_write.write(str_data)
		file_to_write.close()

	## This method for obtaining data for macro.h file
	#  @param self								The object pointer
	#  @param str_output_pathFunction			String of destination location where the file will be written
	#  @return None
	def testdata_write_macro_file(self, str_output_pathFunction):
		# Initialize variable index for writing to macro.h
		int_index = 0
		str_data = '#ifndef _MACRO_H_\n'
		str_data += '#define _MACRO_H_\n\n'
		
		for str_macro in uts_TestCode.g_list_test_addr_macro:
			str_data += '#define {}\t\tTEST_ADDR({})\n'.format(str_macro, int_index)
			int_index += 1
		
		str_data += '\n#define TEST_MACRO_MAX\t\t({})\n'.format(int_index)
		str_data += '\n#endif /* _MACRO_H_ */\n'
		# Write to file
		if g_BOOL_ENABLE_FILE_GENERATE:
			self.__testdata_write_to_file(str_output_pathFunction, str_NAMEFILE_MACRO, str_data)

	## This method for obtaining data for function.h file
	#  @param self								The object pointer
	#  @param list_keywords						List of function name
	#  @param str_output_pathFunction			String of destination location where the file will be written
	#  @return None
	def testdata_write_function_header(self, list_keywords, str_output_pathFunction):
		# Get list of test names API
		str_data = '#ifndef _FUNCTIONS_H_\n'
		str_data += '#define _FUNCTIONS_H_\n\n'
		str_data += '#include <stdbool.h>\n'
		str_data += '#include <stdint.h>\n\n'
		str_data += '#ifdef __cplusplus\n'
		str_data += 'extern "C" {\n'
		str_data += '#endif /* __cplusplus */\n\n'

		for str_test_name in list_keywords:
			str_data += "bool TEST_{}(const char *category, int32_t no);\n".format(str_test_name)
		str_data += '\n#ifdef __cplusplus\n}\n'
		str_data += '#endif\n#endif\n'

		# Write to file
		if g_BOOL_ENABLE_FILE_GENERATE:
			self.__testdata_write_to_file(str_output_pathFunction, str_NAMEFILE_FUNCTION, str_data)

	## This method for obtaining data for testcase.h file
	#  @param self								The object pointer
	#  @param list_keywords						List of function name
	#  @param str_target_module					String of target module defined in config file
	#  @param str_output_pathTestCase			String of destination location where the file will be written
	#  @return None
	def testdata_write_test_case_header(self, list_keywords, str_target_module, str_output_pathTestCase):
		str_data = '#ifndef _TEST_CASE_H_\n'
		str_data += '#define _TEST_CASE_H_\n\n'
		str_data += '#ifdef __cplusplus\n'
		str_data += 'extern "C" {\n'
		str_data += '#endif /* __cplusplus */\n\n'
		str_data += '#define DECLARE_TESTCASE_TABLE(func)	'
		str_data += 'extern struct TestCase UT_##func##_All_Tests[]\n'
		str_data += '#define TESTCASE_TABLE(category,func) '
		str_data += '{ #func,	category,	UT_##func##_All_Tests	}\n\n'

		for str_test_name in list_keywords: 
			str_data += 'DECLARE_TESTCASE_TABLE({0});\n'.format(str_test_name)
		str_data += '\nstatic const struct TestInfo TEST_UT_Info[] = \n{\n'
		for str_test_name in list_keywords:
			str_data += '\tTESTCASE_TABLE("{0}", {1}),\n'.format(str_target_module, str_test_name)
		
		str_data += '\t{ NULL, "", NULL }\n};\n'
		str_data += '\n#ifdef __cplusplus\n}\n#endif\n#endif\n'

		# Write to file
		if g_BOOL_ENABLE_FILE_GENERATE:
			self.__testdata_write_to_file(str_output_pathTestCase, str_NAMEFILE_TESTCASE, str_data)

## This method get cell's value and handling whitespace
#  @param obj_worksheet				Object of current worksheet
#  @param int_row					Cell's row
#  @param int_col					Cell's column
#  @param int_mode					Handling whitespaces mode for cell's value
#  @return str_cell_value			Cell's value
def testdata_get_cell_value(obj_worksheet, int_row, int_col, int_mode=int_MODE_STRIP):
	str_cell_value = obj_worksheet.cell(int_row, int_col).value

	if str_cell_value != None:
		str_cell_value = str(str_cell_value).strip()
		if len(str_cell_value) > int_MAX_CELL_LENGTH:
			raise Exception("[E_UTS_PCL] Exceed max cell length at cell {}{}".format(get_column_letter(int_col), int_row))

		if int_mode == int_MODE_REPLACE:
			str_cell_value = str(str_cell_value).replace(" ", "")
		else:
			pass

	return str_cell_value

## This function for checking file extension
	#  @param str_file		String of file name
	#  @param list_exts		List of expected extensions of file
	#  @return None
def testdata_check_file_extension(str_file, list_exts):
	str_root_ext = os.path.splitext(str_file)
	if str_root_ext[1] not in list_exts:
		raise Exception("[E_UTS_FILE] Extension of {} is {}".format(str_root_ext[0], str(list_exts).replace("[","").replace("]","")))

## This function for getting file object
	#  @param str_file		String of file path
	#  @param str_mode		String of mode when open file
	#  @return fp			File object
def testdata_get_file(str_file, str_mode):
	try:
		fp = open(str_file, mode=str_mode, encoding="utf-8")
		if str_mode == 'r':
			_ = fp.read()
			#reset pointer fp
			fp.seek(0)
	except:
		raise Exception("[E_UTS_FILE] Could not open file: {}".format(str_file))
	else:
		return fp

## This function is used to check testcaseid
# @param  obj_worksheet 					Worksheet object
# @param  list_testcaseid 					List of testcaseid
# @param bool_return_testcase_id_only		Boolean to indicate whether return list of test case ID only
# @return None
def __testdata_check_list_testcaseid(obj_worksheet, list_to_check, bool_return_testcase_id_only):
	# Check empty list
	if len(list_to_check) < 1:
		raise Exception("[E_UTS_PCL] Sheet {}: Not found any test case ID.".format(obj_worksheet.title))

	# Check duplication of testcaseid
	list_available = []
	list_duplicated = []

	for class_cell in list_to_check:
		str_to_check = class_cell.str_value if not bool_return_testcase_id_only else class_cell
		if str_to_check not in list_available:
			list_available.append(str_to_check)
		else:
			if str_to_check not in list_duplicated:
				list_duplicated.append(str_to_check)

		if list_duplicated != []:
			raise Exception("[E_UTS_PCL] Sheet {}: TestCaseID {}: is duplicated.".format(obj_worksheet.title, list_duplicated))

## This function for getting list of Test Case ID
	#  @param obj_worksheet						Object of current worksheet 
	#  @param bool_return_testcase_id_only		Boolean to indicate whether return list of test case ID only
	#  @return list_cell_info					List of location of testcase's cell information
	#  @return list_testcase_id					List of test case ID
def testdata_getTestCaseID(obj_worksheet, bool_return_testcase_id_only):
	# Initialize variables
	list_testcase_id = []
	list_cell_info   = []
	bool_start_test_case_id = False
	# Loop through worksheet
	for int_row in obj_worksheet.iter_rows(min_col = g_SKIPED_COLUMNN + 2, max_col = g_SKIPED_COLUMNN + 2):
		# str_test_case_id = str(int_row[0].value) if int_row[0].value != None else None
		str_test_case_id = testdata_get_cell_value(obj_worksheet, int_row[0].row, int_row[0].column, int_MODE_STRIP)
		# Not found string "Test case ID"
		if bool_start_test_case_id:
			if str_test_case_id != None:
				list_testcase_id.append(str(str_test_case_id))
				list_cell_info.append(CellInfo(int_row[0].row, int_row[0].column, str(str_test_case_id)))
			else:
				str_next_row_value = testdata_get_cell_value(obj_worksheet, int_row[0].row + 1, int_row[0].column, int_MODE_STRIP)
				if str_next_row_value != None:
					raise Exception("[E_UTS_PCL] Missing value at Test case ID at row {} column {}".format(int_row[0].row, get_column_letter(int_row[0].column)))
				else:
					break
		else:
			if (str_test_case_id != None) and (str_test_case_id.strip() == str_TEST_CASE_ID):
				list_testcase_id.append(str_test_case_id)
				bool_start_test_case_id = True

	# Condition to indicate what variable will be returned
	if bool_return_testcase_id_only:
		list_temp_test_case_id = copy.deepcopy(list_testcase_id)
		list_temp_test_case_id.pop(0)
		__testdata_check_list_testcaseid(obj_worksheet, list_temp_test_case_id, bool_return_testcase_id_only)	
		return list_testcase_id
	else:
		__testdata_check_list_testcaseid(obj_worksheet, list_cell_info, bool_return_testcase_id_only)	
		return list_cell_info

## This function for getting list of Device
	#  @param obj_worksheet						Object of current worksheet 
	#  @return list_device						List of device
def testdata_getDevice(obj_worksheet):
	list_device = []
	bool_start_target_device = False
	# Loop through worksheet
	for int_row in obj_worksheet.iter_rows(min_col = g_SKIPED_COLUMNN + 2, max_col = g_SKIPED_COLUMNN + 2):
		# str_test_case_id = str(int_row[0].value) if int_row[0].value != None else None
		str_target_device = testdata_get_cell_value(obj_worksheet, int_row[0].row, g_INT_TARGET_DEVICE_COL - 1, int_MODE_STRIP)
		# Not found string "Device"
		if bool_start_target_device:
			if str_target_device != None:
				list_device.append(str(str_target_device))
			else:
				str_next_row_value = testdata_get_cell_value(obj_worksheet, int_row[0].row + 1, g_INT_TARGET_DEVICE_COL - 1, int_MODE_STRIP)
				if str_next_row_value != None:
					raise Exception("[E_UTS_PCL] Missing value at Device Column at row {} column {}".format(int_row[0].row, get_column_letter(g_INT_TARGET_DEVICE_COL - 1)))
				else:
					break
		else:
			if (str_target_device != None) and (str_target_device.strip() == str_TARGET_DEVICE):
				bool_start_target_device = True
	return list_device

## This function for finding type's row and type of target column
	#  @param obj_worksheet						The object pointer
	#  @param int_target_col					Target column
	#  @param int_max_column					Maximum column, if set to 0 means that no checking int_target_col and int_max_col
	#  @param int_param_type_row				"Type" row 
	#  @param int_param_name_row				"Name" row
	#  @return int_nearest_type_row				Target column type's row
	#  @return str_tmp_type_value				Target column type
def testdata_finding_type_row(worksheet, int_target_col, int_max_column, int_param_type_row, int_param_name_row):
	# Initialize variable for storing type_row 
	int_nearest_type_row = 0
	for int_temp_idx in range(int_param_type_row, int_param_name_row):
		str_tmp_type_value = testdata_get_cell_value(worksheet, int_temp_idx, int_target_col, int_MODE_STRIP)
		if str_tmp_type_value != None:
			int_nearest_type_row =  int_temp_idx
			break
	
	# Condition for int_max_column different with 0 is for user wants to check int_target_col is in the range of PCl or not 
	# If not found any next type row and target column is still in range of PCL 
	if int_nearest_type_row == 0 and int_target_col < int_max_column and int_max_column != 0:
		raise Exception("[E_UTS_PCL] Missing parameter type at column {}!".format(get_column_letter(int_target_col)))

	return int_nearest_type_row, str_tmp_type_value

## This function for getting list of worksheet by prefix or by defined worksheet in config file
	#  @param list_sheetnames					Object of current worksheet 
	#  @param list_prefix						List of prefix defined in config file
	#  @param bool_find_by_prefix				Boolean to indicate whether list_worksheet is getting from finding prefix or by defined worksheet in config file
	#  @param list_sheet_name_contain_data		List of sheet names defined in config file
	#  @return list_worksheet					List of found worksheet
def testdata_get_sheetname(list_sheetnames, list_prefix, bool_find_by_prefix, list_sheet_name_contain_data):
	# Initialize a list to store worksheets
	list_worksheet = []

	# Get worksheet by prefix
	if bool_find_by_prefix:
		for str_sheet_name in list_sheetnames:
			for str_prefix in list_prefix:
				if str_sheet_name.startswith(str_prefix):
					list_worksheet.append(str_sheet_name)
		if len(list_worksheet) == 0:
			raise Exception("[E_UTS_CONF] Invalid input prefix {}".format(list_prefix))
	# Get worksheet by specify sheet name
	else:
		if len(list_sheet_name_contain_data) > 0:
			for sheet_name in list_sheet_name_contain_data:
				if sheet_name not in list_sheetnames:
					raise Exception("[E_UTS_CONF] Invalid sheet name '{}' in SheetNameContainData".format(sheet_name))
			list_worksheet = list_sheet_name_contain_data
		else:
			raise Exception("[E_UTS_CONF] Not found any defined worksheet")

	return list_worksheet

## This function for checking existing config option from config file
	#  @param config							Built-in class config from ConfigParser
	#  @param dict_option						Global dictionary of data obtained in config file
	#  @return None	
def testdata_check_exist_config_option(config, dict_option):
	for str_section, list_option in dict_option.items():
		for str_option in list_option:
			if not config.has_option(str_section, str_option):
				raise Exception("[E_UTS_CONF] Must contain option {} in section [{}]".format(str_option, str_section))

## This function for print out all global variables used in Test Extraction and Test code module
	#  @return None
def testdata_log_out_global_variable():
	# Print out all of global variable defined
	print("#################### GLOBAL VARIABLE ####################")
	list_title = ['Type', 'Global variable']
	list_splash = []
	g_LIST_GLOBAL_VARIABLE.insert(0, list_title)
	# Take the maximum length for each item in the global list
	list_length_alignment_global_var = [len(max(int_idx, key=len)) for int_idx in itertools.zip_longest(*g_LIST_GLOBAL_VARIABLE, fillvalue = " ")] 
	# Append "-" for better vision
	for _ in range(len(list_length_alignment_global_var)):
		list_splash.append("-")
	g_LIST_GLOBAL_VARIABLE.insert(1, list_splash)
	# Loop through global list to log out into console (terminal)
	for int_idx_row in range (len(g_LIST_GLOBAL_VARIABLE)):
		for int_idx in range (len(list_length_alignment_global_var)):
			# Range of maximum character length in a column
			int_tmp_range = list_length_alignment_global_var[int_idx] + 1 - len(g_LIST_GLOBAL_VARIABLE[int_idx_row][int_idx]) 	
			# If it is "-" row, append "-" to make alignment
			if int_idx_row == 1:
				while int_tmp_range > 0:
					g_LIST_GLOBAL_VARIABLE[int_idx_row][int_idx] += "-"
					int_tmp_range -= 1
			else:
				# Add space to make alignment
				while int_tmp_range > 0:
					g_LIST_GLOBAL_VARIABLE[int_idx_row][int_idx] += " "
					int_tmp_range -= 1
	# log out all global variable after making alignment
	for list_item in g_LIST_GLOBAL_VARIABLE:
		str_temp = ''
		for str_small_item in list_item:
			str_temp += str_small_item
		print(str_temp) 

##  Function to check duplicate value following mode
	#  @param list_to_check						List to be checked for duplicated value
	#  @param str_target_to_check				Target to be checked 
	#  @param int_mode							Mode for indicate which attributes is going to be checked
	#  @return bool_check						Boolean for checking duplicated value
def testdata_check_duplicate(list_to_check, str_target_to_check, int_mode):
	bool_check = False
	for item in list_to_check:	
		if int_mode == int_MODE_RETURN:
			if item.output_type == str_target_to_check:
				bool_check = True 
				break
	
	return bool_check			

## Main function for Test Extraction module
	#  @param str_config_path				Location of config file parsed as an argument 
	#  @return None
def uts_ext_extract_pcl_data(str_config_path):
	if (not os.path.exists(str_config_path)):
		raise Exception("[E_UTS_FILE] File {}: is not exist".format(str_config_path))
	testdata_check_file_extension(str_config_path, [".ini"])

	## Parsing argument
	config_parser = configparser.ConfigParser()
	class_config_info = ConfigureInfo()
	class_config_info = class_config_info.testdata_parsing_argument(class_config_info, config_parser, str_config_path)

	str_input_path       			= class_config_info.input_path
	if (not os.path.exists(str_input_path)):
		raise Exception("[E_UTS_FILE] File {}: not exist.".format(str_input_path))
	testdata_check_file_extension(str_input_path, [".xlsx", ".xlsm", ".xltx", ".xltm"])

	str_output_path 	 			= class_config_info.output_path
	if (not os.path.exists(str_output_path)):
		raise Exception("[E_UTS_FILE] File {}: not exist.".format(str_output_path))

	str_target_module	 			= class_config_info.TargetModule
	str_prefix	 		 			= class_config_info.Prefix
	bool_find_by_prefix  	 		= class_config_info.FindByPrefix
	list_sheet_name_contain_data	= class_config_info.SheetNameContainData
	list_pointer_type				= class_config_info.pointer_type
	bool_enable_write_test_code		= class_config_info.EnableTestCode
	bool_enable_write_test_data		= class_config_info.EnableTestData

	global g_NOT_NULL_VALUE
	global g_NULL_VALUE
	global g_NOT_STUB_VALUE
	global g_NOT_CHANGE_VALUE
	global g_LIST_GLOBAL_VARIABLE
	global g_OUTPUT_POINTER_NAMES
	global g_INT_PARAM_NAME_ROW
	global g_INT_PARAM_TYPE_ROW
	global g_INT_PARAM_MEANING_ROW
	global g_INT_PARAM_ITEM_ROW
	global g_INT_PARAM_IO_ROW
	global g_INT_FUNC_NAME_ROW
	global g_INT_TARGET_DEVICE_COL
	global g_STR_TARGET_MODULE
	global g_BOOL_ENABLE_FILE_GENERATE

	g_NOT_NULL_VALUE 	  		= class_config_info.NOT_NULL_VALUE
	g_NULL_VALUE 		  		= class_config_info.NULL_VALUE
	g_NOT_STUB_VALUE 	  		= class_config_info.NOT_STUB_VALUE
	g_NOT_CHANGE_VALUE			= class_config_info.NOT_CHANGE_VALUE
	g_BOOL_ENABLE_FILE_GENERATE	= class_config_info.EnableFileGenerate

	list_workbook = []
	list_workbook.append(str_input_path)

	# Config path following file structure
	str_path_Function = os.path.join("base", str_target_module, "include")
	str_path_TestCase = os.path.join("ut", str_target_module, "include")
	str_output_pathFunction = os.path.join(str_output_path, str_path_Function)
	str_output_pathTestCase = os.path.join(str_output_path, str_path_TestCase)
	if not os.path.exists(str_output_pathFunction): os.makedirs(str_output_pathFunction)
	if not os.path.exists(str_output_pathTestCase): os.makedirs(str_output_pathTestCase)

	# reset the list of pointer string and pointer macro
	uts_TestCode.g_list_test_addr_macro		= []
	uts_TestCode.g_list_test_addr_string	= []

	# Loop over workbook
	for name_workbook in list_workbook:
		workbook = load_workbook(name_workbook, data_only=True, keep_links=False)

		# Get workbook information
		class_wb_info = WorkbookInfo(workbook)

		# Get list of worksheet based on condition
		list_worksheet = testdata_get_sheetname(workbook.sheetnames, str_prefix, bool_find_by_prefix, list_sheet_name_contain_data)

		# Write function.h file
		class_wb_info.testdata_write_function_header(list_worksheet, str_output_pathFunction)

		# Write test_case.h file
		class_wb_info.testdata_write_test_case_header(list_worksheet, str_target_module, str_output_pathTestCase)

		for worksheet in list_worksheet:
			if (list_worksheet == []):
				workbook.close()
				raise Exception("[E_UTS_CONF] File {}: No target sheet defined".format(str_config_path))
			
			if (worksheet not in workbook.sheetnames):
				workbook.close()
				raise Exception("[E_UTS_CONF] Sheet {} not found in PCL".format(worksheet))
			
			print("============================================")
			print("#### [Test Data] Start: {}".format(worksheet))
			class_worksheet_info = WorksheetInfo(workbook[worksheet], worksheet, list_pointer_type)

			# Init global variable
			g_INT_PARAM_NAME_ROW = 0 
			g_INT_PARAM_TYPE_ROW = 0
			g_INT_PARAM_MEANING_ROW = 0
			g_INT_PARAM_ITEM_ROW = 0
			g_INT_PARAM_IO_ROW = 0
			g_INT_FUNC_NAME_ROW = 0
			# Finding max column
			int_start_col = g_OFFSET_COL + 7
			class_worksheet_info.int_max_column = int_start_col
			int_temp_col = int_start_col + 1
			# Finding position of parameter name & function name
			pos_func_name 		= class_worksheet_info.testdata_get_cell(str_FUNCTION_NAME, 0, g_OFFSET_COL)
			pos_param_name 		= class_worksheet_info.testdata_get_cell(str_PARAMETER_NAME, 0, g_OFFSET_COL)
			pos_io_type			= class_worksheet_info.testdata_get_cell(str_IO, 0, g_OFFSET_COL)
			pos_item_type 		= class_worksheet_info.testdata_get_cell(str_ITEM, 0, g_OFFSET_COL)
			pos_type 			= class_worksheet_info.testdata_get_cell(str_TYPE, 0, g_OFFSET_COL)
			pos_meaning 		= class_worksheet_info.testdata_get_cell(str_MEANING, 0, g_OFFSET_COL)
			pos_target_device 	= class_worksheet_info.testdata_get_cell(str_TARGET_DEVICE, g_OFFSET_ROW, g_OFFSET_COL)

			g_INT_PARAM_NAME_ROW = pos_param_name.int_row
			g_INT_PARAM_TYPE_ROW = pos_type.int_row
			g_INT_PARAM_MEANING_ROW = pos_meaning.int_row
			g_INT_PARAM_ITEM_ROW = pos_item_type.int_row
			g_INT_PARAM_IO_ROW = pos_io_type.int_row
			g_INT_FUNC_NAME_ROW = pos_func_name.int_row
			g_INT_TARGET_DEVICE_COL = pos_target_device.int_col
			g_STR_TARGET_MODULE = str_target_module

			# Finding function name
			str_function_name = testdata_get_cell_value(class_worksheet_info.worksheet, g_INT_FUNC_NAME_ROW, 5, int_MODE_STRIP)
			if str_function_name == None:
				raise Exception("[E_UTS_PCL] Missing function name at row {} in column E in sheet".format(g_INT_FUNC_NAME_ROW))

			# Finding ending column by searching name
			# If not found name, check the existance of other Parameter in the column
			# If not found any value, it is end column. If found, raise Exception for missing value
			while(True):
				str_temp_name_value = testdata_get_cell_value(class_worksheet_info.worksheet, g_INT_PARAM_NAME_ROW, int_temp_col, int_MODE_STRIP)
				if str_temp_name_value != None:
					int_temp_col += 1
				else:
					for int_temp_row in range (pos_item_type.int_row, pos_io_type.int_row + 1):
						str_temp = testdata_get_cell_value(class_worksheet_info.worksheet, int_temp_row, int_temp_col, int_MODE_STRIP)
						if str_temp != None:
							raise Exception("[E_UTS_PCL] Missing value at row {} in column {}".format(int_temp_row, get_column_letter(int_temp_col)))
					break

			if pos_io_type.int_col == int_temp_col:
				raise Exception("[E_UTS_PCL] Column {} is not start cell column".format(get_column_letter(int_temp_col)))
			
			# Store max column into class attribute for global usage
			class_worksheet_info.int_max_column = int_temp_col

			# Detect invalid in worksheet
			class_worksheet_info.testdata_detect_invalid()
			# Config path following file structure:
			pos_src_cell = class_worksheet_info.testdata_get_cell("Source file", 0, 0)
			int_end_col    = class_worksheet_info.testdata_get_cell(str_DATA_SPECIFICATION, 0, 0).int_col
			for int_column in range(pos_src_cell.int_col, int_end_col + 1):
				str_value = testdata_get_cell_value(workbook[worksheet], pos_src_cell.int_row, int_column, int_MODE_STRIP)
				if int_column == int_end_col:
					raise Exception("[E_UTS_FILE] Not found source file name in PCL")
				elif str_value == None:
					continue
				elif str_value.endswith(".c") or str_value.endswith(".cpp"):
					pos_src_cell.int_col = int_column
					break
				else:
					pass
			
			if str_value.endswith(".c"):
				str_src_folder   		= testdata_get_cell_value(workbook[worksheet], pos_src_cell.int_row, pos_src_cell.int_col, int_MODE_STRIP).replace(".c", "")
				src_language = ".c"
			if str_value.endswith(".cpp"):
				str_src_folder   		= testdata_get_cell_value(workbook[worksheet], pos_src_cell.int_row, pos_src_cell.int_col, int_MODE_STRIP).replace(".cpp", "")
				src_language = ".cpp"

			pos_method_cell = class_worksheet_info.testdata_get_cell("Method", 0, 0)
			method_flag = testdata_get_cell_value(workbook[worksheet], pos_method_cell.int_row, pos_method_cell.int_col+1, int_MODE_STRIP)

			pos_namespace_cell = class_worksheet_info.testdata_get_cell("namespace", 0, 0)
			namespace_value = testdata_get_cell_value(workbook[worksheet], pos_namespace_cell.int_row, pos_namespace_cell.int_col+1, int_MODE_STRIP)


			str_output_pathData     = os.path.join(str_output_path, "base", str_target_module, str_src_folder, "patterns")
			str_output_pathTable    = os.path.join(str_output_path, "ut", str_target_module, str_src_folder)
			str_output_pathCode     = os.path.join(str_output_path, "base", str_target_module, str_src_folder)
			if not os.path.exists(str_output_pathData): os.makedirs(str_output_pathData)
			if not os.path.exists(str_output_pathTable): os.makedirs(str_output_pathTable)
			if not os.path.exists(str_output_pathCode): os.makedirs(str_output_pathCode)

			# reset output pointer name
			g_OUTPUT_POINTER_NAMES = []

			# Get param location
			list_param_info = class_worksheet_info.testdata_get_param_info()

			if bool_enable_write_test_data:
				# Loop over number of test cases
				class_worksheet_info.testdata_write_test_pattern(list_param_info, str_output_pathData)

				# Write struct test pattern
				class_worksheet_info.testdata_write_struct_test_pattern(list_param_info, str_output_pathData)

				# Write test case table
				class_worksheet_info.testdata_getTableFile(str_output_pathTable, src_language)

			print("#### [Test Data] End: {}".format(worksheet))

			if bool_enable_write_test_code:
				# get test pattern info for test code generation
				list_pattern_info = class_worksheet_info.testdata_create_testcode_info(list_param_info)

				# write test code
				uts_tsc_write_test_code(str_output_pathCode, worksheet, str_function_name, list_pattern_info, config_parser, src_language, str_src_folder, method_flag, namespace_value)

	# Print out all global variable used 	
	if len(g_LIST_GLOBAL_VARIABLE) > 0:
		testdata_log_out_global_variable()

	# Write macro.h file
	class_wb_info.testdata_write_macro_file(str_output_pathFunction)

	print("------------------------------------------------------")
	print("[E_UTS_OK] Sucessfully generating Test Pattern")

if __name__ == "__main__":
	if (len(argv) != 2):
		raise Exception("[E_UTS_FILE] Please input config file path")

	uts_ext_extract_pcl_data(argv[1])

